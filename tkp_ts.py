import os
import sys

# Windows consoles often default to cp1252; emoji in print() would raise UnicodeEncodeError and kill startup.
if sys.platform == "win32":
    try:
        if hasattr(sys.stdout, "reconfigure"):
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        if hasattr(sys.stderr, "reconfigure"):
            sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

import json
import base64
import openpyxl
from datetime import datetime, timedelta

import pandas as pd
from pandas.tseries.holiday import USFederalHolidayCalendar
from pandas.tseries.offsets import CustomBusinessDay

import numpy as np
import plotly.graph_objs as go

import dash
from dash import html, dcc, dash_table
from dash.dash_table.Format import Format, Scheme, Symbol
import dash_bootstrap_components as dbc
from dash.dependencies import Input, Output, State

import yfinance as yf

import quantstats as qs
from quantstats import utils

import tearsheet_disclosure as tsd
from tearsheet_gate_ui import build_sibling_accept_gate
from tearsheet_gate_auth import (
    build_gate_password_row,
    gate_password_row_style,
    GATE_PASSWORD_ROW_ID,
    GATE_PASSWORD_INPUT_ID,
    GATE_PASSWORD_SUBMIT_ID,
    GATE_PASSWORD_PORTAL_ID,
    GATE_PASSWORD_ERROR_ID,
    GATE_PASSWORD_VISIBLE_STORE_ID,
    INVALID_PASSWORD_MESSAGE,
    ADMIN_PORTAL_PATH,
    TKP_SESSION_KEY,
    load_tkp_admin_auth_settings,
)
from tcp_admin import AdminAuthManager, configure_flask_session_secret
from tearsheet_portal import render_legacy_diagnostics_table, render_portal_page
from tearsheet_header import (
    build_header_date_label_children_from_date,
    build_tearsheet_header_row,
)
from tearsheet_date_defaults import default_add_row_date_str
from flask import session, redirect, jsonify
from collections import OrderedDict

# ==============================================================================
# 1) BUSINESS-DAY CALENDAR
#    Create a CustomBusinessDay that drops weekends & US federal holidays
# ==============================================================================
us_bd = CustomBusinessDay(calendar=USFederalHolidayCalendar())

# ==============================================================================
# 2) LOGO ENCODING & ESTHETICS
#    Read your branded logo and encode it as base64 for Dash images
# ==============================================================================
logo_path = r"C:\Users\H&CDanHughes\Hughes & Company\Hughes & Company - Documents\2_Hughes & Company Marketing\Branded Logo\Trianle-Only-Logo.png"
try:
    with open(logo_path, "rb") as f:
        logo_b64 = base64.b64encode(f.read()).decode("utf-8")
    logo_src = f"data:image/png;base64,{logo_b64}"
except FileNotFoundError:
    print(f"Error: Logo file not found at {logo_path}")
    logo_src = ""
# ─── Brand colors ────────────────────────────────
WHITE_BG     = "#ffffff"
GREY_BG      = "#EBEBEB"   # matches ggplot2's plot_bgcolor
PRIMARY_COLOR = "#0D3562"  # your "blue," user‐changeable
SECONDARY_COLOR = "#CCCCCC"

LEFT_TABLE_GAPS = "20px"
RIGHT_TABLE_GAPS = "30px"

HEADER_ROW_CLASS = "bg-light"

# ==============================================================================
# HELPER FUNCTIONS & CONSTANTS (Safe additions - won't affect current functionality)
# ==============================================================================

# Constants for magic numbers used throughout the code
TRADING_DAYS_PER_YEAR = 252
BUSINESS_DAYS_PER_YEAR = 365
BASELINE_AMOUNT = 150000
NOMINAL_ASSETS = 300000
ACCOUNT_COUNT = 4
OPEN_ACCOUNTS = 2
CLOSED_PROFITABLE = 2
CLOSED_UNPROFITABLE = 0
MIN_RETURN = 0.36
MAX_RETURN = 4.2
AVERAGE_MARGIN_USAGE = 1.77
TRANSACTION_FEE_PER_CONTRACT = 0.30

def validate_file_path(file_path):
    """Validate that file exists and is accessible"""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    if not os.access(file_path, os.R_OK):
        raise PermissionError(f"Cannot read file: {file_path}")
    return file_path

def validate_nav_data(df: pd.DataFrame) -> bool:
    """Validate NAV dataframe structure and content"""
    if df.empty:
        return False
    
    if df.index.has_duplicates:
        print("Warning: Duplicate dates found, removing duplicates")
        df = df[~df.index.duplicated(keep="first")]
    
    if df.isnull().any().any():
        print("Warning: Missing values found in NAV data")
    
    return True

def safe_logo_loading(logo_path: str) -> str:
    """Safely load logo with fallback"""
    try:
        with open(logo_path, "rb") as f:
            logo_b64 = base64.b64encode(f.read()).decode("utf-8")
        return f"data:image/png;base64,{logo_b64}"
    except FileNotFoundError:
        print(f"Error: Logo file not found at {logo_path}")
        return ""
    except Exception as e:
        print(f"Error loading logo: {e}")
        return ""

def safe_benchmark_download(symbol: str, max_retries: int = 3) -> pd.Series:
    """Download benchmark data with retry logic"""
    for attempt in range(max_retries):
        try:
            return utils.download_returns(symbol)
        except Exception as e:
            if attempt == max_retries - 1:
                print(f"Failed to download {symbol} after {max_retries} attempts: {e}")
                return pd.Series(dtype=float)
            print(f"Attempt {attempt + 1} failed for {symbol}, retrying...")
            import time
            time.sleep(1)  # Brief delay before retry
    return pd.Series(dtype=float)

def create_monthly_calendar(monthly_simple):
    """Create monthly calendar table from monthly simple returns"""
    years = sorted(monthly_simple.index.year.unique())
    months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    
    monthly_data = {"Year": [str(y) for y in years]}
    
    for idx, m in enumerate(months, start=1):
        monthly_data[m] = [
            f"{monthly_simple.get(pd.Period(f'{y}-{idx:02d}'), 0):.2f}%"
            if pd.Period(f'{y}-{idx:02d}') in monthly_simple.index
            else ""
            for y in years
        ]
    
    # Calculate year totals
    yearly_simple = monthly_simple.groupby(monthly_simple.index.year).sum()
    monthly_data["Year Total"] = [
        f"{yearly_simple.get(y, 0):.2f}%"
        for y in years
    ]
    
    return pd.DataFrame(monthly_data)

# Safe helper function for monthly calendar (alternative to inline logic)
def build_monthly_calendar_safe(monthly_simple_series):
    """Alternative monthly calendar builder with error handling"""
    try:
        years = sorted(monthly_simple_series.index.year.unique())
        months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        
        monthly_data = {"Year": [str(y) for y in years]}
        
        for idx, m in enumerate(months, start=1):
            monthly_data[m] = [
                f"{monthly_simple_series.get(pd.Period(f'{y}-{idx:02d}'), 0):.2f}%"
                if pd.Period(f'{y}-{idx:02d}') in monthly_simple_series.index
                else ""
                for y in years
            ]
        
        # Calculate year totals
        yearly_simple = monthly_simple_series.groupby(monthly_simple_series.index.year).sum()
        monthly_data["Year Total"] = [
            f"{yearly_simple.get(y, 0):.2f}%"
            for y in years
        ]
        
        return pd.DataFrame(monthly_data)
    except Exception as e:
        print(f"Warning: Error building monthly calendar: {e}")
        return pd.DataFrame()  # Return empty DataFrame as fallback

# ==============================================================================
# 3) CONFIGURATION
#    Strategy name, benchmark list, and NAV CSV path
# ==============================================================================
STRATEGY_NAME = "TKP"

# ── LAYOUT CONFIGURATION ────────────────────────────────────────────────────
# Set to True for side-by-side layout (current), False for stacked layout (new)
USE_SIDE_BY_SIDE_LAYOUT = True  # Toggle between True (side-by-side) and False (stacked)

# Set to True to show percentage axis on the right side of NAV chart
SHOW_PERCENTAGE_AXIS = True  # Toggle between True (show %) and False (hide %)

# When True, show a "Debug / Data Provenance" table at bottom of page (field name vs source).
DEBUG_PROVENANCE = False

BENCHMARKS = [
    "^SP500TR",   # S&P 500 Total Return
    "AGG",        # US Aggregate Bond
    "GLD",        # Gold ETF
    "BTC-USD",    # Bitcoin
    "ETH-USD",    # Ethereum
]

xlsx_path = (
    r"C:\Users\H&CDanHughes\Hughes & Company\Hughes & Company - Documents\3_Advisors Marketing (Tearsheets, PitchBooks, etc)\1. Tearsheet Project\TKP\VADI\Copy of tkp_alex_old1.xlsx"
)

# Persisted Daily Returns editor state (Add Row / Delete Last Row). Survives browser hard refresh.
SECRET_EDITOR_STATE_FILENAME = "daily_returns_secret_state.json"


def _secret_editor_state_path():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), SECRET_EDITOR_STATE_FILENAME)


_PCT_JSON_COLS = {"Perc. Net", "Cumm Perc. Net"}

def _save_secret_editor_state(rows):
    """Write full secret table rows to JSON so they reload after app restart / hard refresh.

    Pct columns (Perc. Net, Cumm Perc. Net) are stored as raw decimals (÷100) regardless
    of whether the in-memory value is already in display-% form (×100).  _load normalises
    them back to ×100 on the way in, giving a stable round-trip contract.
    """
    if rows is None:
        return
    path = _secret_editor_state_path()
    try:
        serializable = []
        for r in rows:
            row_out = {}
            for k, v in r.items():
                if isinstance(v, float) and v != v:
                    row_out[k] = None
                elif isinstance(v, (np.integer,)):
                    row_out[k] = int(v)
                elif isinstance(v, (np.floating,)):
                    fv = float(v)
                    # Normalise pct cols to raw-decimal form so load always gets consistent values
                    if k in _PCT_JSON_COLS:
                        fv = fv / 100.0
                    row_out[k] = fv
                elif isinstance(v, float) and k in _PCT_JSON_COLS:
                    row_out[k] = v / 100.0
                else:
                    row_out[k] = v
            serializable.append(row_out)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(serializable, f, indent=2)
        last_date = ""
        for r in reversed(serializable):
            if r.get("Date"):
                last_date = r["Date"]
                break
        print(f"💾 Saved {len(serializable)} rows to {SECRET_EDITOR_STATE_FILENAME} (last date: {last_date})")
    except OSError as e:
        print(f"⚠️ Could not save daily returns editor state to {path}: {e}")


def _load_secret_editor_state(expected_columns):
    """Load saved rows if present; normalize keys to match current column layout.
    
    Handles column name mismatches between older saved JSON and current Excel headers
    by trying both the canonical name and known aliases.
    """
    path = _secret_editor_state_path()
    if not os.path.isfile(path):
        return None
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError) as e:
        print(f"⚠️ Could not load saved daily returns state from {path}: {e}")
        return None
    if not isinstance(data, list) or len(data) == 0:
        return None

    # Map of canonical column name -> list of aliases that may appear in older JSON
    _aliases = {
        "Perc. Net":       ["Perc. Net", " Perc. Net", "%Net", "percNet"],
        "Cumm Perc. Net":  ["Cumm Perc. Net", "S net cummulative Perc.", "S net cummulative %", "S net cummulative perc"],
        "Net P&L":         ["Net P&L", "Net PL"],
        "Net P&L / Unit":  ["Net P&L / Unit", "Net PL per unit", "Net P&L / unit"],
        "Cumm Fee":        ["Cumm Fee", "cumm fee"],
        "Fee (20%)":       ["Fee (20%)", "0.2"],
    }

    cols = list(expected_columns)
    out = []
    for row in data:
        if not isinstance(row, dict):
            continue
        new_row = {}
        for c in cols:
            # Try canonical name first
            val = row.get(c, "")
            # If empty/missing and aliases exist, try them
            if (val == "" or val is None) and c in _aliases:
                for alias in _aliases[c]:
                    candidate = row.get(alias, "")
                    if candidate != "" and candidate is not None:
                        val = candidate
                        break
            new_row[c] = "" if val is None else val
        out.append(new_row)
    return out if out else None


def _load_fresh_secret_records():
    """Re-read JSON on every call so page refreshes always see the latest saved rows.

    Falls back to the module-level secret_table_records (from server startup) if the
    JSON file is missing or unreadable.
    """
    try:
        loaded = _load_secret_editor_state(full_daily_df.columns)
    except Exception:
        loaded = None
    if loaded is not None:
        for _row in loaded:
            for _col in _PCT_JSON_COLS:
                _v = _row.get(_col)
                if isinstance(_v, (int, float)):
                    _row[_col] = round(_v * 100, 6)
        return loaded
    return secret_table_records


# If Excel has blank rows or formula cells without cached values, pandas/openpyxl can
# stop "early" when inferring the used range. Set this to the last Excel row you
# want included (1-indexed, including header row). Set to None to disable.
FORCE_LAST_EXCEL_ROW = 715

# ============================================================================== 
# 4) LOAD & VALIDATE NAV DATA (Excel cols C=Date, N=nav‑x1)
# ==============================================================================

# Validate file exists and is accessible before attempting to read
if not os.path.exists(xlsx_path):
    print(f"❌ File not found: {xlsx_path}")
    sys.exit(1)

if not os.access(xlsx_path, os.R_OK):
    print(f"❌ Permission denied: Cannot read file '{xlsx_path}'")
    print("   This usually means the file is open in Excel or another program.")
    print("   Please close the file and try again.")
    sys.exit(1)

try:
    # First, find the last row with data using openpyxl to ensure we read all rows
    # This prevents pandas from stopping early at empty rows
    # Use data_only=False so formula cells count as "present" even if cached values
    # aren't stored in the workbook (common if the file wasn't recalculated/saved).
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb["Sheet1"]
    
    # Find the last row with data in columns C (Date) or N (NAV)
    # This technique matches the working approach in test_read_excel.py
    # Scan from row 2 to max_row + 1, checking only the columns we actually need
    last_row = 1  # Start with header row
    for row_idx in range(2, ws.max_row + 1):
        cell_c = ws.cell(row=row_idx, column=3)   # Column C = Date
        cell_n = ws.cell(row=row_idx, column=14)  # Column N = NAV
        if cell_c.value is not None or cell_n.value is not None:
            last_row = row_idx
    
    wb.close()

    # Ensure we use at least FORCE_LAST_EXCEL_ROW if set
    if FORCE_LAST_EXCEL_ROW is not None:
        last_row = max(last_row, int(FORCE_LAST_EXCEL_ROW))
    
    print(f"📋 Found last non-blank row: {last_row} (Excel row number, scanned columns C and N)")
    
    # Load only columns C and N, parse C as Date
    # Explicitly read up to the last row with data to ensure we don't miss anything
    read_params = {
        "sheet_name": "Sheet1",
        "usecols": "C,N",              # Excel columns C and N
        "header": 0,                   # first row is header
        "parse_dates": ["Date"],       # parse the C‑column into datetime
        "engine": "openpyxl",
    }
    # Only add nrows if we found data rows (last_row > 1 means we have data beyond header)
    if last_row > 1:
        read_params["nrows"] = last_row - 1  # Number of data rows (excluding header)
    
    NAV_df = pd.read_excel(xlsx_path, **read_params)
    print(f"📊 Raw rows read from Excel: {len(NAV_df)}")
    
    # Drop rows where BOTH Date and NAV are missing (completely empty rows)
    NAV_df = NAV_df.dropna(how='all')
    print(f"📊 Rows after dropping completely empty: {len(NAV_df)}")
    
    # Add safe validation (won't break existing functionality)
    if NAV_df.empty:
        print("❌ NAV data is empty")
        sys.exit(1)
        
    if len(NAV_df.columns) < 2:
        print("❌ NAV data must have at least 2 columns")
        sys.exit(1)
        
except PermissionError as e:
    print(f"❌ Permission denied: Cannot read file '{xlsx_path}'")
    print("   This usually means the file is open in Excel or another program.")
    print("   Please close the file and try again.")
    sys.exit(1)
except Exception as e:
    print(f"❌ Failed to load NAV data: {e}")
    print(f"   File path: {xlsx_path}")
    sys.exit(1)

# ==============================================================================
# 4b) LOAD FULL DAILY-RETURNS TABLE (Excel cols A–S) for secret "Daily Returns" view
# ==============================================================================
_secret_editor_restored_from_disk = False
try:
    full_daily_df = pd.read_excel(
        xlsx_path,
        sheet_name="Sheet1",
        usecols="A:S",
        header=0,
        nrows=last_row - 1 if last_row > 1 else None,
        engine="openpyxl",
    )
    full_daily_df = full_daily_df.dropna(how="all")

    # Normalise every column name to string so float headers (e.g. 0.2) match the rename map
    full_daily_df.columns = [str(c) if not isinstance(c, str) else c for c in full_daily_df.columns]

    # Drop the empty "Unnamed" placeholder column (Excel col D has no header)
    unnamed_cols = [c for c in full_daily_df.columns if c.startswith("Unnamed")]
    if unnamed_cols:
        full_daily_df.drop(columns=unnamed_cols, inplace=True)

    # Rename columns to clean display names (preserve Excel order, move Deposit last)
    # Both old (% symbol) and new (perc word) header variants are mapped
    col_map = {
        "cash transfers": "Deposit",
        "trading date": "#Day",
        "Date": "Date",
        "Balance (StoneX)": "StoneX",
        "Plus500 NL": "Plus500",
        "StoneX NL": "StoneX NL",
        "#": "# Trades",
        "$PL": "$PL",
        "0.2": "Fee (20%)",
        "cumm fee": "Cumm Fee",
        "Net P&L": "Net P&L",
        "Net PL": "Net P&L",
        "Net P&L / unit": "Net P&L / Unit",
        "Net PL per unit": "Net P&L / Unit",
        "nav-x1": "NAV",
        "Loss Carry": "Loss Carry",
        "%Net": "Perc. Net",
        " Perc. Net": "Perc. Net",
        "percNet": "Perc. Net",
        "S net cummulative %": "Cumm Perc. Net",
        "S net cummulative perc": "Cumm Perc. Net",
        "S net cummulative Perc.": "Cumm Perc. Net",
        "HWM": "HWM",
        "Cash": "Cash",
    }
    # Strip leading/trailing whitespace from headers before renaming
    full_daily_df.columns = [c.strip() if isinstance(c, str) else c for c in full_daily_df.columns]
    full_daily_df.rename(columns=col_map, inplace=True)

    # Permanently remove excluded columns from this UI (Plus500 is kept)
    SECRET_EXCLUDED = {"Cash", "StoneX NL"}
    full_daily_df.drop(columns=[c for c in SECRET_EXCLUDED if c in full_daily_df.columns], inplace=True)

    # Final safety: drop any remaining None/NaN column names and deduplicate
    full_daily_df = full_daily_df.loc[:, full_daily_df.columns.notna()]
    full_daily_df.columns = pd.io.common.dedup_names(list(full_daily_df.columns), is_potential_multiindex=False)

    # Parse Date column
    if "Date" in full_daily_df.columns:
        full_daily_df["Date"] = pd.to_datetime(full_daily_df["Date"], errors="coerce")
        full_daily_df = full_daily_df.dropna(subset=["Date"])

    # Sort chronologically (oldest first) for HWM asterisk calculation
    full_daily_df = full_daily_df.sort_values("Date").reset_index(drop=True)

    # HWM asterisk: mark rows where HWM exceeds the previous row's HWM (chronological)
    if "HWM" in full_daily_df.columns:
        hwm_shifted = full_daily_df["HWM"].shift(1)
        hwm_new_high = (full_daily_df["HWM"] > hwm_shifted) & hwm_shifted.notna()
        full_daily_df["HWM"] = full_daily_df["HWM"].apply(
            lambda v: f"${v:,.2f}" if pd.notna(v) else ""
        )
        full_daily_df.loc[hwm_new_high, "HWM"] = full_daily_df.loc[hwm_new_high, "HWM"] + " *"

    # Format Date for display
    full_daily_df["Date"] = full_daily_df["Date"].dt.strftime("%Y-%m-%d")

    # Move Deposit to last column
    if "Deposit" in full_daily_df.columns:
        deposit_col = full_daily_df.pop("Deposit")
        full_daily_df["Deposit"] = deposit_col

    # Ensure Plus500 sits immediately left of StoneX
    if "Plus500" in full_daily_df.columns and "StoneX" in full_daily_df.columns:
        stonex_idx = full_daily_df.columns.get_loc("StoneX")
        plus_col = full_daily_df.pop("Plus500")
        full_daily_df.insert(stonex_idx, "Plus500", plus_col)

    # Format numeric columns for display
    money_cols = [
        "Plus500", "StoneX", "$PL", "Fee (20%)",
        "Cumm Fee", "Net P&L", "Net P&L / Unit", "NAV", "Loss Carry",
    ]
    for col in money_cols:
        if col in full_daily_df.columns:
            full_daily_df[col] = full_daily_df[col].apply(
                lambda v: f"${v:,.2f}" if pd.notna(v) else ""
            )
    # Perc. Net and Cumm Perc. Net stored as display-ready percentages (e.g. 1.2345 means 1.2345%).
    # DataTable appends the % sign via Format — no further multiply needed at render.
    for col in ["Perc. Net", "Cumm Perc. Net"]:
        if col in full_daily_df.columns:
            full_daily_df[col] = (
                pd.to_numeric(full_daily_df[col], errors="coerce") * 100
            )
    if "Deposit" in full_daily_df.columns:
        full_daily_df["Deposit"] = full_daily_df["Deposit"].apply(
            lambda v: f"${v:,.0f}" if pd.notna(v) and v != 0 else ""
        )
    if "# Trades" in full_daily_df.columns:
        full_daily_df["# Trades"] = pd.to_numeric(full_daily_df["# Trades"], errors="coerce")
        full_daily_df["# Trades"] = full_daily_df["# Trades"].apply(
            lambda v: str(int(v)) if pd.notna(v) else ""
        )
    if "#Day" in full_daily_df.columns:
        full_daily_df["#Day"] = pd.to_numeric(full_daily_df["#Day"], errors="coerce")
        full_daily_df["#Day"] = full_daily_df["#Day"].apply(
            lambda v: str(int(v)) if pd.notna(v) else ""
        )

    # Add stable row identifier for CRUD operations
    full_daily_df.insert(0, "_row_id", range(len(full_daily_df)))

    secret_table_records = full_daily_df.to_dict("records")
    secret_all_columns = [c for c in full_daily_df.columns if c not in ("_row_id", "Edit", "Del")]
    secret_table_columns = [{"name": c, "id": c} for c in secret_all_columns]

    _loaded_secret = _load_secret_editor_state(full_daily_df.columns)
    if _loaded_secret is not None:
        # Pct columns are always saved as raw decimals by _save_secret_editor_state.
        # Multiply ×100 here to restore display-ready values expected by DataTable Format.
        for _row in _loaded_secret:
            for _col in _PCT_JSON_COLS:
                _v = _row.get(_col)
                if isinstance(_v, (int, float)):
                    _row[_col] = round(_v * 100, 6)
        secret_table_records = _loaded_secret
        _secret_editor_restored_from_disk = True
        print(
            f"📋 Restored {len(_loaded_secret)} daily return rows from saved editor state "
            f"({SECRET_EDITOR_STATE_FILENAME})"
        )

    print(f"📋 Secret table loaded: {len(secret_table_records)} rows, {len(secret_table_columns)} columns")
    print(f"   Columns: {secret_all_columns}")
except Exception as e:
    print(f"⚠️  Failed to load secret table data: {e}")
    import traceback; traceback.print_exc()
    _secret_editor_restored_from_disk = False
    secret_table_records = []
    secret_all_columns = []
    secret_table_columns = []

SECRET_DEFAULT_VISIBLE = [
    "#Day", "Date", "Plus500", "StoneX", "Perc. Net",
    "$PL", "NAV", "HWM", "Fee (20%)", "Deposit",
]

PUBLIC_DAILY_COLUMNS = ["#Day", "Date", "NAV", "Perc. Net", "$PL", "HWM", "Fee (20%)"]

_PCT_COL_FORMAT = Format(precision=4, scheme=Scheme.fixed, symbol=Symbol.yes, symbol_suffix="%")


def _build_table_columns(col_names):
    """Build DataTable column dicts; pct columns get numeric Format so sorting works.
    Shared by both admin and public Daily Returns tables."""
    PCT_COLS = {"Perc. Net", "Cumm Perc. Net"}
    cols = []
    for c in col_names:
        if c in PCT_COLS:
            cols.append({"name": c, "id": c, "type": "numeric", "format": _PCT_COL_FORMAT})
        else:
            cols.append({"name": c, "id": c})
    return cols


def _secret_table_columns(col_names):
    return _build_table_columns(col_names)

def _default_add_row_date_str():
    """Return previous business day (Mon -> Fri) as YYYY-MM-DD.

    Delegates to the shared tearsheet_date_defaults helper so TCP/AGM Add Row
    modals can mirror this exact method (see tearsheet_date_defaults.py)."""
    return default_add_row_date_str()

# Compute latest date from Daily Returns for status label
def _get_latest_daily_date():
    """Return latest date string from secret_table_records, or 'unavailable'."""
    try:
        dates = [r.get("Date") for r in secret_table_records if r.get("Date")]
        if dates:
            latest = max(dates)
            return pd.to_datetime(latest).strftime("%B %d, %Y")
    except Exception:
        pass
    return "unavailable"

DAILY_RETURNS_LATEST_DATE = _get_latest_daily_date()

# Make sure pandas named the second column correctly
if NAV_df.columns[1] != "nav‑x1" and NAV_df.columns[1] != "nav-x1":
    NAV_df.rename(columns={NAV_df.columns[1]: "nav-x1"}, inplace=True)

# Ensure Date column is datetime - handle cases where parse_dates didn't work
if NAV_df["Date"].dtype == 'object':
    # Try to convert string dates to datetime
    NAV_df["Date"] = pd.to_datetime(NAV_df["Date"], errors='coerce')
    # Remove rows where date conversion failed
    invalid_conversions = NAV_df["Date"].isna()
    if invalid_conversions.any():
        print(f"⚠️  Warning: {invalid_conversions.sum()} row(s) had unparseable dates and were removed")
        NAV_df = NAV_df[~invalid_conversions]
        if NAV_df.empty:
            print("❌ No valid dates remaining after date parsing")
            sys.exit(1)

# Auto-correct year typos in dates (e.g., 2025-01-06 after 2026-01-05 should be 2026-01-06)
# This fixes Excel data entry errors where the year gets typed incorrectly
corrected_dates = []
corrections_made = 0
for i in range(len(NAV_df)):
    current_date = NAV_df.iloc[i]["Date"]
    
    # Check if this date goes backwards in time compared to previous date
    if i > 0 and current_date < NAV_df.iloc[i-1]["Date"]:
        prev_date = NAV_df.iloc[i-1]["Date"]
        
        # Check if it's a year typo (same month/day but different year)
        # and the year is exactly 1 year behind
        if (current_date.month == prev_date.month and 
            current_date.day >= prev_date.day and
            prev_date.year - current_date.year == 1):
            
            # Correct the year
            corrected_date = current_date.replace(year=prev_date.year)
            corrected_dates.append((i, current_date, corrected_date))
            NAV_df.iloc[i, NAV_df.columns.get_loc("Date")] = corrected_date
            corrections_made += 1

if corrections_made > 0:
    print(f"⚠️  Auto-corrected {corrections_made} date(s) with year typos:")
    for idx, old_date, new_date in corrected_dates[:5]:  # Show first 5
        print(f"   Row {idx+2}: {old_date.date()} → {new_date.date()}")
    if len(corrected_dates) > 5:
        print(f"   ... and {len(corrected_dates) - 5} more")

# Set Date as the index
NAV_df.set_index("Date", inplace=True)

# Validate and filter out invalid dates (pandas datetime64[ns] can only handle ~1677-2262)
# This prevents overflow errors from dates like 2505-11-04
MIN_VALID_DATE = pd.Timestamp('1900-01-01')
MAX_VALID_DATE = pd.Timestamp('2260-01-01')  # Well before pandas limit

# Ensure index is datetime before comparison
if not pd.api.types.is_datetime64_any_dtype(NAV_df.index):
    print("⚠️  Warning: Index is not datetime, attempting conversion...")
    NAV_df.index = pd.to_datetime(NAV_df.index, errors='coerce')
    # Remove rows where conversion failed
    invalid_conversions = NAV_df.index.isna()
    if invalid_conversions.any():
        print(f"⚠️  Warning: {invalid_conversions.sum()} row(s) had unparseable dates and were removed")
        NAV_df = NAV_df[~invalid_conversions]
        if NAV_df.empty:
            print("❌ No valid dates remaining after date conversion")
            sys.exit(1)

# Now safe to compare dates
invalid_dates = (NAV_df.index < MIN_VALID_DATE) | (NAV_df.index > MAX_VALID_DATE)
if invalid_dates.any():
    invalid_count = invalid_dates.sum()
    print(f"⚠️  Warning: Found {invalid_count} invalid date(s) outside valid range (1900-2260)")
    print(f"   Invalid dates: {NAV_df.index[invalid_dates].tolist()}")
    print(f"   Removing invalid dates...")
    NAV_df = NAV_df[~invalid_dates]
    
    if NAV_df.empty:
        print("❌ No valid dates remaining after filtering")
        sys.exit(1)

# Drop exact duplicates so .asfreq() works
if NAV_df.index.has_duplicates:
    NAV_df = NAV_df[~NAV_df.index.duplicated(keep="first")]

# Reindex to your custom business‑day calendar (fills missing dates with NaN)
# Now safe because all dates are validated
nav_col_name = NAV_df.columns[0]  # Get the NAV column name (usually "nav-x1")
print(f"📊 Before asfreq: {len(NAV_df)} rows, date range: {NAV_df.index.min().date()} to {NAV_df.index.max().date()}")
print(f"   Last NAV value before asfreq: ${NAV_df[nav_col_name].iloc[-1]:,.2f}")

NAV_df = NAV_df.asfreq(us_bd)

print(f"📊 After asfreq: {len(NAV_df)} rows, date range: {NAV_df.index.min().date()} to {NAV_df.index.max().date()}")
print(f"   Last NAV value after asfreq: ${NAV_df[nav_col_name].iloc[-1]:,.2f}")

# Forward fill NaN values to prevent gaps in the chart
# This ensures smooth continuation between dates
NAV_df = NAV_df.ffill()

print(f"📊 After ffill: {len(NAV_df)} rows")
print(f"   Last 5 NAV values: {NAV_df[nav_col_name].tail(5).tolist()}")
print(f"   Last 5 dates: {NAV_df.index[-5:].tolist()}")

# ─────────────────────────────────────────────────────────────
# ADJUST FOR CASH TRANSFER (remove non-performance effect)
# ─────────────────────────────────────────────────────────────
# Configuration: Set to False to disable auto-detection and show raw NAV
AUTO_DETECT_CASH_TRANSFERS = True  # Set to True to enable auto-detection of deposits/withdrawals

# Manual specification (only used if AUTO_DETECT is False and values are provided)
CASH_TRANSFER_DATE = None     # e.g., "2024-01-16" or pd.Timestamp("2024-01-16"); use None to disable
CASH_TRANSFER_ROW = None       # Excel row number (1-indexed, header is row 1); use None to disable
TRANSFER_AMOUNT = None      # Positive to add back a withdrawal; negative to remove a deposit; None to skip

def _resolve_transfer_date_from_row(xlsx_path: str, excel_row: int) -> pd.Timestamp:
    if excel_row is None:
        raise ValueError("excel_row is None")
    if excel_row < 2:
        raise ValueError("Excel row number must be >= 2 (row 1 is header)")
    if not os.access(xlsx_path, os.R_OK):
        raise PermissionError(f"Cannot read file '{xlsx_path}' - file may be open in Excel")
    temp_df = pd.read_excel(
        xlsx_path,
        sheet_name="Sheet1",
        usecols="C",
        header=0,
        engine="openpyxl",
    )
    df_index = excel_row - 2  # Excel row 2 -> DataFrame index 0
    if df_index < 0 or df_index >= len(temp_df):
        raise IndexError(
            f"Excel row {excel_row} is out of range (valid: 2 to {len(temp_df) + 1})"
        )
    return pd.to_datetime(temp_df.iloc[df_index, 0])

def _read_excel_dates(xlsx_path: str) -> pd.DatetimeIndex:
    """Return a normalized DatetimeIndex of actual Excel dates (no forward-filled days)."""
    dates = pd.read_excel(
        xlsx_path, sheet_name="Sheet1", usecols="C", header=0, engine="openpyxl"
    )["Date"].dropna()
    dates = pd.to_datetime(dates).dt.normalize()
    return pd.DatetimeIndex(sorted(dates.unique()))

def _next_real_excel_date(xlsx_path: str, candidate: pd.Timestamp) -> pd.Timestamp:
    """Snap candidate to the next actual Excel date (not a generated business day)."""
    real = _read_excel_dates(xlsx_path)
    later = real[real >= pd.Timestamp(candidate).normalize()]
    return later[0] if len(later) else pd.Timestamp(candidate)

def _apply_cash_transfer_adjustment(
    nav_df: pd.DataFrame,
    nav_column: str,
    transfer_row: int,
    transfer_amount: float,
    xlsx_path: str,
):
    """
    Normalize NAV series for deposits/withdrawals to remove non-performance effects.
    
    Financial intent:
    - Withdrawal: NAV drops (e.g., 182k → 82k, delta = -100k) → ADD BACK +100k from transfer date onward
    - Deposit: NAV rises (e.g., 182k → 282k, delta = +100k) → SUBTRACT -100k from transfer date onward
    
    Args:
        nav_df: DataFrame with Date index and NAV column (already reindexed/forward-filled)
        nav_column: Name of NAV column
        transfer_row: Excel row number where transfer occurs (1-indexed, header is row 1)
        transfer_amount: Amount to adjust (positive = add back withdrawal, negative = remove deposit)
                         If None, auto-detect from Excel NAV delta
        xlsx_path: Path to Excel file for reading actual NAV values
    
    Returns:
        None (modifies nav_df in place)
    """
    # Read actual Excel data to get true before/after NAV values (not forward-filled)
    df_excel = pd.read_excel(
        xlsx_path,
        sheet_name="Sheet1",
        usecols="C,N",
        header=0,
        engine="openpyxl",
    )
    df_excel.columns = ["Date", "NAV"]
    df_excel["Date"] = pd.to_datetime(df_excel["Date"], errors='coerce')
    df_excel = df_excel.dropna(subset=["Date"])
    
    # Filter out invalid dates (pandas datetime64 limitation)
    MIN_VALID_DATE = pd.Timestamp('1900-01-01')
    MAX_VALID_DATE = pd.Timestamp('2260-01-01')
    invalid_dates = (df_excel["Date"] < MIN_VALID_DATE) | (df_excel["Date"] > MAX_VALID_DATE)
    if invalid_dates.any():
        df_excel = df_excel[~invalid_dates]
    
    df_excel = df_excel.reset_index(drop=True)
    
    # Get the Excel row indices (0-indexed)
    excel_idx = transfer_row - 2  # Excel row 2 → DataFrame index 0
    if excel_idx < 0 or excel_idx >= len(df_excel):
        raise IndexError(f"Excel row {transfer_row} is out of range")
    if excel_idx == 0:
        raise ValueError(f"Cannot use row {transfer_row} as transfer row (need previous row for comparison)")
    
    # Get actual NAV values from Excel at transfer boundary
    before_nav_excel = df_excel.iloc[excel_idx - 1]["NAV"]  # Row before transfer
    after_nav_excel = df_excel.iloc[excel_idx]["NAV"]      # Row with transfer
    delta = after_nav_excel - before_nav_excel
    
    # Get the transfer date from Excel
    transfer_date_excel = df_excel.iloc[excel_idx]["Date"]
    
    # Determine effective adjustment amount
    if transfer_amount is not None:
        # User specified amount: use it directly
        # Positive = add back (for withdrawals), Negative = remove (for deposits)
        effective_amount = transfer_amount
    else:
        # Auto-detect: negate the actual cash movement
        # If NAV dropped (withdrawal), we need to add back (positive correction)
        # If NAV rose (deposit), we need to subtract (negative correction)
        effective_amount = -delta
    
    # Find the transfer date in the reindexed NAV_df (may be forward-filled)
    if transfer_date_excel not in nav_df.index:
        matching_dates = nav_df.index[nav_df.index >= transfer_date_excel]
        if len(matching_dates) == 0:
            raise IndexError(
                f"Transfer date {transfer_date_excel.date()} is after last NAV date"
            )
        transfer_date = matching_dates[0]
    else:
        transfer_date = transfer_date_excel
    
    # Get mask for all rows from transfer_date onward (never touch earlier dates)
    mask = nav_df.index >= transfer_date
    affected_rows = mask.sum()
    
    # Snapshot before adjustment
    before_value = nav_df.loc[transfer_date, nav_column]
    baseline_value = nav_df.loc[nav_df.index[0], nav_column]  # Preserve starting baseline
    before_last_value = nav_df.loc[nav_df.index[-1], nav_column] if len(nav_df) > 0 else None
    
    # Debug: Show last 5 values before adjustment
    print(f"   [DEBUG] Last 5 NAV values BEFORE adjustment: {nav_df[nav_column].tail(5).tolist()}")
    print(f"   [DEBUG] Adjustment amount: ${effective_amount:+,.2f} (will be {'added' if effective_amount > 0 else 'subtracted'})")
    
    # Apply adjustment to all rows from transfer_date onward
    nav_df.loc[mask, nav_column] += effective_amount
    
    # Debug: Show last 5 values after adjustment
    print(f"   [DEBUG] Last 5 NAV values AFTER adjustment: {nav_df[nav_column].tail(5).tolist()}")
    
    # Verify baseline unchanged
    baseline_after = nav_df.loc[nav_df.index[0], nav_column]
    if abs(baseline_after - baseline_value) > 0.01:
        raise ValueError(f"Baseline NAV changed from {baseline_value:,.2f} to {baseline_after:,.2f}")
    
    after_value = nav_df.loc[transfer_date, nav_column]
    after_last_value = nav_df.loc[nav_df.index[-1], nav_column] if len(nav_df) > 0 else None
    
    # Log the adjustment
    verb = "added back" if effective_amount >= 0 else "removed"
    action = "withdrawal" if effective_amount > 0 else "deposit"
    print(
        f"✅ Cash transfer adjustment: {verb} ${abs(effective_amount):,.0f} ({action}) | "
        f"effective date: {transfer_date.date()}"
    )
    print(
        f"   Excel row {transfer_row}: NAV {before_nav_excel:,.2f} → {after_nav_excel:,.2f} (delta: {delta:+,.2f})"
    )
    print(
        f"   Adjusted {affected_rows} rows from {transfer_date.date()} to {nav_df.index[-1].date()}"
    )
    print(
        f"   NAV[{transfer_date.date()}] before: {before_value:,.2f} → after: {after_value:,.2f}"
    )
    if before_last_value is not None and after_last_value is not None:
        print(
            f"   NAV[{nav_df.index[-1].date()}] before: {before_last_value:,.2f} → after: {after_last_value:,.2f}"
        )
    print(f"   ✅ Baseline preserved: ${baseline_value:,.2f}")

def _auto_detect_cash_transfers(xlsx_path: str, min_transfer_amount: float = 50000) -> list:
    """
    Automatically detect all cash transfers (deposits/withdrawals) by scanning for large NAV jumps.
    
    Args:
        xlsx_path: Path to Excel file
        min_transfer_amount: Minimum NAV change to consider a transfer (default: $50k)
    
    Returns:
        List of tuples: (transfer_row, delta, transfer_type)
    """
    # Read actual Excel data
    df_excel = pd.read_excel(
        xlsx_path,
        sheet_name="Sheet1",
        usecols="C,N",
        header=0,
        engine="openpyxl",
    )
    df_excel.columns = ["Date", "NAV"]
    
    # Convert dates with error handling for invalid dates
    df_excel["Date"] = pd.to_datetime(df_excel["Date"], errors='coerce')
    
    # Validate and filter out invalid dates (same logic as main NAV loading)
    MIN_VALID_DATE = pd.Timestamp('1900-01-01')
    MAX_VALID_DATE = pd.Timestamp('2260-01-01')
    
    # Remove rows with invalid dates
    df_excel = df_excel.dropna(subset=["Date"])
    invalid_dates = (df_excel["Date"] < MIN_VALID_DATE) | (df_excel["Date"] > MAX_VALID_DATE)
    if invalid_dates.any():
        print(f"   [Auto-detect] Removing {invalid_dates.sum()} invalid date(s) from Excel")
        df_excel = df_excel[~invalid_dates]
    
    # Remove rows with missing NAV
    df_excel = df_excel.dropna(subset=["NAV"])
    df_excel = df_excel.reset_index(drop=True)
    
    transfers = []
    for i in range(1, len(df_excel)):  # Start from row 1 (skip header)
        before_nav = df_excel.iloc[i - 1]["NAV"]
        after_nav = df_excel.iloc[i]["NAV"]
        delta = after_nav - before_nav
        
        # Detect if this is a large jump (likely a cash transfer)
        if abs(delta) >= min_transfer_amount:
            # Calculate expected NAV change from returns (to distinguish from performance)
            # If NAV changed by more than expected trading P&L, it's likely a transfer
            # For now, just use absolute threshold - can be refined later
            transfer_row = i + 2  # Convert to 1-indexed Excel row (header is row 1)
            transfer_type = "withdrawal" if delta < 0 else "deposit"
            transfers.append((transfer_row, delta, transfer_type))
    
    return transfers

try:
    if not AUTO_DETECT_CASH_TRANSFERS:
        print("ℹ️  Cash transfer auto-detection disabled (showing raw NAV)")
        if CASH_TRANSFER_DATE is not None or CASH_TRANSFER_ROW is not None:
            print("   Manual cash transfer adjustment specified:")
    
    if CASH_TRANSFER_DATE is not None or CASH_TRANSFER_ROW is not None:
        # Manual specification: single transfer
        if CASH_TRANSFER_ROW is not None:
            # Use specified row (this is the Excel row where transfer occurs)
            transfer_row = CASH_TRANSFER_ROW
        elif CASH_TRANSFER_DATE is not None:
            # Find Excel row matching the specified date
            df_temp = pd.read_excel(
                xlsx_path,
                sheet_name="Sheet1",
                usecols="C",
                header=0,
                engine="openpyxl",
            )
            df_temp.columns = ["Date"]
            df_temp["Date"] = pd.to_datetime(df_temp["Date"])
            target_date = pd.to_datetime(CASH_TRANSFER_DATE).normalize()
            matches = df_temp[df_temp["Date"].dt.normalize() == target_date]
            if len(matches) == 0:
                raise ValueError(f"No Excel row found for date {CASH_TRANSFER_DATE}")
            transfer_row = matches.index[0] + 2  # Convert to 1-indexed Excel row
        else:
            raise ValueError("Must specify either CASH_TRANSFER_DATE or CASH_TRANSFER_ROW")
        
        # Apply adjustment using actual Excel row
        _apply_cash_transfer_adjustment(
            NAV_df, 
            "nav-x1", 
            transfer_row, 
            TRANSFER_AMOUNT, 
            xlsx_path
        )
    elif AUTO_DETECT_CASH_TRANSFERS:
        # Auto-detect all cash transfers
        print("🔍 Auto-detecting cash transfers...")
        transfers = _auto_detect_cash_transfers(xlsx_path, min_transfer_amount=50000)
        
        if len(transfers) == 0:
            print("   No cash transfers detected (no NAV jumps >= $50k)")
        else:
            print(f"   Found {len(transfers)} potential cash transfer(s):")
            for transfer_row, delta, transfer_type in transfers:
                print(f"   - Row {transfer_row}: {transfer_type} of ${abs(delta):,.0f} (delta: {delta:+,.0f})")
            
            # Apply corrections to all detected transfers
            # Process in chronological order (lowest row first)
            transfers.sort(key=lambda x: x[0])
            
            for transfer_row, delta, transfer_type in transfers:
                try:
                    # Auto-detect amount by negating the delta
                    transfer_amount = None  # Will auto-detect from delta
                    _apply_cash_transfer_adjustment(
                        NAV_df,
                        "nav-x1",
                        transfer_row,
                        transfer_amount,
                        xlsx_path
                    )
                except Exception as e:
                    print(f"   ⚠️  Failed to adjust transfer at row {transfer_row}: {e}")
                    continue
        
        print("✅ Auto-detection complete")
except Exception as e:
    print(f"⚠️  Warning: cash transfer adjustment skipped: {e}")
    import traceback
    traceback.print_exc()

print(f"📊 Final NAV data: {len(NAV_df)} rows")
print(f"   Date range: {NAV_df.index.min().date()} to {NAV_df.index.max().date()}")
nav_col_name_final = NAV_df.columns[0]  # Get column name before NAV_col is defined
print(f"   Last NAV value: ${NAV_df[nav_col_name_final].iloc[-1]:,.2f}")
print(f"   Second-to-last NAV value: ${NAV_df[nav_col_name_final].iloc[-2]:,.2f}")
print("✅ NAV data loaded successfully (Date + nav‑x1).")



# ==============================================================================
# 5) SELECT NAV VALUE COLUMN
#    Look for one of your preferred names first (e.g. "CL" or "NAV"),
#    then fall back to the first numeric column if none are present.
#    To change later, just update NAV_CANDIDATES.
# ==============================================================================
NAV_CANDIDATES = [
    "nav-x1",    # your composite Net Liquidation Value column in the sheet
       # alternative name you might use in the future
]

for cand in NAV_CANDIDATES:
    if cand in NAV_df.columns:
        NAV_col = cand
        break
else:
    # fallback: first purely numeric column
    numeric_cols = NAV_df.select_dtypes(include=[np.number]).columns
    if not numeric_cols.empty:
        NAV_col = numeric_cols[0]
    else:
        raise KeyError(
            f"None of {NAV_CANDIDATES!r} present and no numeric column found."
        )

print(f"▶️ Using NAV column: {NAV_col}")


def _canonical_records_from_secret_rows(rows):
    """Build canonical Date/NAV records from persisted Daily Returns rows (same logic as live store)."""
    if not rows:
        return []
    pairs = []
    for r in rows:
        date_str = r.get("Date", "")
        nav_str = r.get("NAV", "")
        if not date_str or not nav_str:
            continue
        try:
            dt = pd.to_datetime(date_str)
            nav_val = float(str(nav_str).replace("$", "").replace(",", "").strip())
            if nav_val > 0:
                pairs.append((dt, nav_val))
        except Exception:
            continue
    if not pairs:
        return []
    df = pd.DataFrame(pairs, columns=["Date", "NAV"]).sort_values("Date")
    df = df.drop_duplicates(subset="Date", keep="last")
    return [
        {"Date": d.strftime("%Y-%m-%d"), "NAV": float(v)}
        for d, v in zip(df["Date"], df["NAV"])
    ]


def _build_canonical_nav_records(nav_df: pd.DataFrame, nav_column: str):
    """Build clean shared Date/NAV rows for dashboard calculations."""
    if nav_df is None or nav_df.empty or nav_column not in nav_df.columns:
        return []
    df = nav_df[[nav_column]].copy().reset_index()
    # Date may be the index name ("Date") or a generic "index" after reset_index.
    if "Date" not in df.columns and "index" in df.columns:
        df = df.rename(columns={"index": "Date"})
    if "Date" not in df.columns:
        return []
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df[nav_column] = pd.to_numeric(df[nav_column], errors="coerce")
    df = df.dropna(subset=["Date", nav_column]).sort_values("Date")
    df = df.drop_duplicates(subset="Date", keep="last")
    return [
        {"Date": d.strftime("%Y-%m-%d"), "NAV": float(v)}
        for d, v in zip(df["Date"], df[nav_column])
        if float(v) > 0
    ]

CANONICAL_NAV_RECORDS_INITIAL = _build_canonical_nav_records(NAV_df, NAV_col)
if _secret_editor_restored_from_disk:
    _nav_from_secret = _canonical_records_from_secret_rows(secret_table_records)
    if _nav_from_secret:
        CANONICAL_NAV_RECORDS_INITIAL = _nav_from_secret


# ==============================================================================
# 6) NON-COMPOUNDED DAILY RETURNS
#    Compute each day's P&L as a % of starting NAV (baseline)
# ==============================================================================
baseline = NAV_df[NAV_col].iloc[0]
daily_returns = NAV_df[NAV_col].diff().div(baseline).dropna()


# ==============================================================================
# 7) DOWNLOAD & ALIGN BENCHMARKS
#    For each symbol, get daily returns, align to NAV dates, then compute cum & drawdown
# ==============================================================================
bench_map = OrderedDict([
    ("SPXTR", "^SP500TR"),
    ("AGG",   "AGG"),
    ("GLD",   "GLD"),
    ("BTC",   "BTC-USD"),
    ("ETH",   "ETH-USD"),
])
# Filter to only those you configured
bench_map = {k:v for k,v in bench_map.items() if v in BENCHMARKS}

bench_ret, bench_cum, bench_dd = {}, {}, {}
for name, sym in bench_map.items():
    try:
        # 1) get evenly-spaced returns series
        full_ret = utils.download_returns(sym)
        
        # Add safe validation (won't break existing functionality)
        if full_ret is None or full_ret.empty:
            print(f"Warning: No data for {sym}, skipping")
            continue
            
        aligned = full_ret.reindex(NAV_df.index).ffill().bfill().dropna()
        # 2) cumulative growth
        cum = (1 + aligned).cumprod()
        # 3) drawdown in %
        dd = (cum / cum.cummax() - 1) * 100

        bench_ret[name] = aligned
        bench_cum[name] = cum
        bench_dd[name]  = dd
        
    except Exception as e:
        print(f"Warning: Failed to load {sym}: {e}")
        continue

# ==============================================================================
# 8) MONTHLY SIMPLE RETURNS (NON-COMPOUNDED) 
#    + MONTH-CELLS THAT SUM TO TRUE YEAR-OVER-YEAR RETURN
# ==============================================================================

# 8a) Month-end and month-start NAV
mp          = NAV_df.index.to_period("M")
month_last  = NAV_df[NAV_col].groupby(mp).last()

# Calculate month_first: use the last NAV value from the day before each month starts
# This ensures we get the correct starting NAV even if there are gaps in months
month_first = pd.Series(index=month_last.index, dtype=float)
for period in month_last.index:
    # Get the first day of this month
    month_start = period.start_time
    # Find the last NAV value before this month starts
    nav_before_month = NAV_df[NAV_col][NAV_df.index < month_start]
    if len(nav_before_month) > 0:
        # Use the last NAV value before the month starts
        month_first.loc[period] = nav_before_month.iloc[-1]
    else:
        # For the very first month, use baseline
        month_first.loc[period] = baseline

# 8b) Compute each month's change **relative** to fixed baseline
monthly_simple = (month_last - month_first) / baseline * 100

# Hard-coded overrides for specific months (requested adjustments)
override_months = {
    pd.Period('2025-04', freq='M'): 4.58,
    pd.Period('2025-10', freq='M'): 0.58,
}
for override_period, override_value in override_months.items():
    if override_period in monthly_simple.index:
        monthly_simple.loc[override_period] = override_value

# Debug output for October 2025
oct_2025_period = pd.Period("2025-10", freq="M")
if oct_2025_period in monthly_simple.index:
    oct_last = month_last.loc[oct_2025_period]
    oct_first = month_first.loc[oct_2025_period]
    oct_return = monthly_simple.loc[oct_2025_period]
    print(f"🔍 October 2025 Debug:")
    print(f"   month_last (Oct end NAV): ${oct_last:,.2f}")
    print(f"   month_first (Sep end NAV): ${oct_first:,.2f}")
    print(f"   baseline: ${baseline:,.2f}")
    print(f"   calculated return: {oct_return:.2f}%")
    print(f"   formula: (${oct_last:,.2f} - ${oct_first:,.2f}) / ${baseline:,.2f} * 100 = {oct_return:.2f}%")

 
 
 
 
 
 
 
# 8c) Sum those 12 numbers to get the Year Total
yearly_simple = monthly_simple.groupby(monthly_simple.index.year).sum()

# 8d) Build the “calendar” table
years  = sorted(monthly_simple.index.year.unique())
months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

monthly_data = {"Year": [str(y) for y in years]}

for idx, m in enumerate(months, start=1):
    monthly_data[m] = [
        # look up period in monthly_simple; if missing, blank
        f"{monthly_simple.get(pd.Period(f'{y}-{idx:02d}'), 0):.4f}%"
        if pd.Period(f"{y}-{idx:02d}") in monthly_simple.index
        else ""
        for y in years
    ]

# 8e) Use the sum-of-months for Year Total
monthly_data["Year Total"] = [
    f"{yearly_simple.get(y, 0):.4f}%"
    for y in years
]

monthly_df = pd.DataFrame(monthly_data)

# Safe helper function for monthly calendar (alternative to inline logic)
def build_monthly_calendar_safe(monthly_simple_series):
    """Alternative monthly calendar builder with error handling"""
    try:
        years = sorted(monthly_simple_series.index.year.unique())
        months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        
        monthly_data = {"Year": [str(y) for y in years]}
        
        for idx, m in enumerate(months, start=1):
            monthly_data[m] = [
                f"{monthly_simple_series.get(pd.Period(f'{y}-{idx:02d}'), 0):.4f}%"
                if pd.Period(f'{y}-{idx:02d}') in monthly_simple_series.index
                else ""
                for y in years
            ]
        
        # Calculate year totals
        yearly_simple = monthly_simple_series.groupby(monthly_simple_series.index.year).sum()
        monthly_data["Year Total"] = [
            f"{yearly_simple.get(y, 0):.4f}%"
            for y in years
        ]
        
        return pd.DataFrame(monthly_data)
    except Exception as e:
        print(f"Warning: Error building monthly calendar: {e}")
        return pd.DataFrame()  # Return empty DataFrame as fallback

# ==============================================================================
# 9) DAILY PERFORMANCE METRICS
#    Define a helper to compute all your key stats over any period
# ==============================================================================
def calculate_period_metrics(returns: pd.Series, start_date: pd.Timestamp) -> dict:
    """
    Given a series of non-compounded daily returns,
    compute cumulative return, annualized, avg daily,
    win/loss counts & rates, top/bottom 3 days.
    """
    short_period = {
        "Cumulative Return":      "0.0%",
        "Annualized Return":      "0.0%",
        "Avg Daily Return":       "0.000%",
        "Number of Trading Days": "0",
        "% Winning Days":         "0 (0.0%)",
        "% Losing Days":          "0 (0.0%)",
        "Best 3 Days":            "0.00%, 0.00%, 0.00%",
        "Worst 3 Days":           "0.00%, 0.00%, 0.00%",
    }
    if len(returns) < 2:
        return short_period

    cum = returns.sum()
    days = len(returns)
    span_days = (returns.index.max() - start_date).days
    if span_days == 0:
        annualized = cum
    else:
        annualized = cum * 365.0 / span_days
    avg = returns.mean()

    wins = (returns > 0).sum()
    losses = (returns < 0).sum()

    top3 = returns.nlargest(3) * 100
    bot3 = returns.nsmallest(3) * 100

    return {
        "Cumulative Return":      f"{cum*100:.1f}%",
        "Annualized Return":      f"{annualized*100:.1f}%",
        "Avg Daily Return":       f"{avg*100:.3f}%",
        "Number of Trading Days": str(days),
        "% Winning Days":         f"{wins} ({wins/days*100:.1f}%)",
        "% Losing Days":          f"{losses} ({losses/days*100:.1f}%)",
        "Best 3 Days":            ", ".join(f"{v:.2f}%" for v in top3),
        "Worst 3 Days":           ", ".join(f"{v:.2f}%" for v in bot3),
    }

# ── List of the metrics in the order you want them shown ────────────────────
metric_labels = [
    "Cumulative Return",
    "Annualized Return",
    "Avg Daily Return",
    "Number of Trading Days",
    "% Winning Days",
    "% Losing Days",
    "Best 3 Days",
    "Worst 3 Days"
]

# ── Define period boundaries ────────────────────────────────────────────────
inception_start = NAV_df.index.min()
ttm_start       = NAV_df.index.max() - pd.DateOffset(years=1)

# ── Slice your strategy series ─────────────────────────────────────────────
one_year_returns  = daily_returns.loc[ttm_start:].dropna()
inception_returns = daily_returns.copy()

# ── Slice your SPXTR series ────────────────────────────────────────────────
spxtr_series       = bench_ret["SPXTR"]
spxtr_one_year     = spxtr_series.loc[ttm_start:].dropna()
spxtr_inception    = spxtr_series.loc[inception_start:].dropna()

# ── Compute metrics ─────────────────────────────────────────────────────────
one_year_metrics       = calculate_period_metrics(one_year_returns,  ttm_start)
inception_metrics      = calculate_period_metrics(inception_returns, inception_start)

# ── Assemble your DataFrame ────────────────────────────────────────────────
daily_perf_df = pd.DataFrame({
    "Metric": metric_labels,
    f"{STRATEGY_NAME} (1 Year/TTM)":    [one_year_metrics[m] for m in metric_labels],
    f"{STRATEGY_NAME} (Inception)":     [inception_metrics[m] for m in metric_labels],
})




# ──────────────────────────────────────────────────────────────────────────────
# CONFIG: choose your drawdown method
# ──────────────────────────────────────────────────────────────────────────────
# True  = standard peak-to-trough drawdown (quantstats style)
# False = custom baseline-relative drawdown
USE_QUANTSTATS_DD_STRATEGY   = False   # TKP
USE_QUANTSTATS_DD_BENCHMARKS = True  # SPXTR & others
SHOW_DD_PRICE = False   # True ⇒ “2025-02-20 (152345.67)”, False ⇒ “2025-02-20”

# ==============================================================================
# 10) Compute Worst (Max) Drawdown Profile, Inception Only
# ==============================================================================
def drawdown_profile(
    nav: pd.Series,
    baseline: float,
    use_quantstats: bool,
    show_price: bool,
    price_series: pd.Series
) -> dict:
    """
    Returns the single worst drawdown episode,
    with optional price display.
    """
    # 1) running peak & choose drawdown series
    running_max = nav.cummax()
    if use_quantstats:
        dd_series = (nav / running_max - 1) * 100
    else:
        dd_series = (nav - running_max) / baseline * 100

    # 2) find trough and its preceding peak
    trough = dd_series.idxmin()
    peak   = nav.loc[:trough].idxmax()

    # 3) format dates and optional prices
    peak_date   = peak.strftime("%Y-%m-%d")
    valley_date = trough.strftime("%Y-%m-%d")

    if show_price:
        peak_str   = f"{peak_date} - {price_series.loc[peak]:,.2f}"
        valley_str = f"{valley_date} - {price_series.loc[trough]:,.2f}"
    else:
        peak_str   = peak_date
        valley_str = valley_date

     # 4) recovery logic
    rec      = nav[trough:][nav[trough:] >= nav[peak]]
    rec_idx  = rec.index[0] if not rec.empty else None
    decline_days = (trough - peak).days

    if rec_idx:
        # fully recovered
        end_date   = rec_idx.strftime("%Y-%m-%d")
        if show_price:
            price = price_series.loc[rec_idx]
            end_str = f"{end_date} - {price:,.2f}"
        else:
            end_str = end_date

        recovery_days = (rec_idx - trough).days
        total_days    = (rec_idx - peak).days
        recovery_text = f"{recovery_days} days"
        total_text    = f"{total_days} days"

    else:
        # still in drawdown
        # inside your `else:` for “not recovered”:
        last_date     = nav.index.max()
        last_price    = price_series.loc[last_date]
        peak_price    = price_series.loc[peak]
        trough_price  = price_series.loc[trough]

        remaining_pct = (peak_price - last_price) / (peak_price - trough_price) * 100

        last_date = nav.index.max()
        if show_price:
            last_price = price_series.loc[last_date]
            end_str = (
                f"TBD - Current Price is {last_price:,.2f}, "
                f"{remaining_pct:.1f} % of the current drawdown is left for a full recovery"
            )
        else:
            end_str = "TBD"

        recovery_days = (last_date - trough).days
        total_days    = (last_date - peak).days
        recovery_text = f"Ongoing for {recovery_days} days"
        total_text    = f"Ongoing for {total_days} days"

    # 5) worst depth
    depth = dd_series.min()

    # assemble result with dynamic field names
    result = {
        "Depth":           f"{depth:.1f}%",
        "Decline Period":  f"{decline_days} days",
        "Recovery Period": recovery_text,
        "Total Duration":  total_text
    }

    if show_price:
        result["Start Date & Price"]    = peak_str
        result["Valley Date & Price"]   = valley_str
        result["End Date & Price"]      = end_str
    else:
        result["Start Date"]            = peak_str
        result["Valley Date"]           = valley_str
        result["End Date"]              = end_str

    return result

# Safe helper function for drawdown profile (alternative to main function)
def safe_drawdown_profile(
    nav: pd.Series,
    baseline: float,
    use_quantstats: bool,
    show_price: bool,
    price_series: pd.Series
) -> dict:
    """
    Safe version of drawdown profile with error handling.
    Returns empty dict if calculation fails.
    """
    try:
        return drawdown_profile(nav, baseline, use_quantstats, show_price, price_series)
    except Exception as e:
        print(f"Warning: Error calculating drawdown profile: {e}")
        return {
            "Depth": "N/A",
            "Decline Period": "N/A", 
            "Recovery Period": "N/A",
            "Total Duration": "N/A",
            "Start Date": "N/A",
            "Valley Date": "N/A",
            "End Date": "N/A"
        }

spxtr_price_series = (
    yf.download(
        "^SP500TR",
        start="2023-04-01",
        end="2025-07-01",
        auto_adjust=True,
        progress=False
    )[['Close']]
    .squeeze()
    .reindex(NAV_df.index)
    .ffill()
)

# Safe helper function for SPXTR price series download
def safe_spxtr_download():
    """Safe SPXTR price series download with error handling"""
    try:
        return (
            yf.download(
                "^SP500TR",
                start="2023-04-01",
                end="2025-07-01",
                auto_adjust=True,
                progress=False
            )[['Close']]
            .squeeze()
            .reindex(NAV_df.index)
            .ffill()
        )
    except Exception as e:
        print(f"Warning: Error downloading SPXTR data: {e}")
        # Return empty series as fallback
        return pd.Series(dtype=float, index=NAV_df.index)

# ==============================================================================
# 11) Build the “Worst Drawdown” DataFrame
# ==============================================================================
# Strategy NAV + baseline
strategy_nav      = NAV_df[NAV_col]
strategy_baseline = strategy_nav.iloc[0]

# SPXTR NAV scaled to strategy baseline
spxtr_returns    = bench_ret["SPXTR"]
spxtr_nav        = (1 + spxtr_returns.loc[inception_start:]).cumprod() * strategy_baseline
spxtr_baseline   = strategy_baseline

# Make sure you have the raw SPXTR close-price series defined as spxtr_price_series
# e.g.: spxtr_price_series = utils.download_price("^SP500TR").reindex(NAV_df.index).ffill()

period_slices = {
    f"{STRATEGY_NAME} (Inception)": (
        strategy_nav,
        strategy_baseline,
        USE_QUANTSTATS_DD_STRATEGY,
        SHOW_DD_PRICE,
        strategy_nav        # use NAV as price series for TKP
    ),
    "SPXTR (Inception)": (
        spxtr_nav,
        spxtr_baseline,
        USE_QUANTSTATS_DD_BENCHMARKS,
        SHOW_DD_PRICE,
        spxtr_price_series  # your SPXTR close-price Series
    ),
}

max_dd_df = (
    pd.DataFrame({
        name: drawdown_profile(
            nav,
            baseline,
            use_flag,
            show_price,
            price_series
        )
        for name, (nav, baseline, use_flag, show_price, price_series)
            in period_slices.items()
    })
    .rename_axis("Metric")
    .reset_index()
)


# ==============================================================================
# 12) Hard-coded “Additional Information”
# ==============================================================================
grouped_info = {
    "Account Stats": [
        ("Nominal Assets Being Traded in the Program", "200k"),
        ("Total Accounts/Tranches Opened",           "4"),
        ("Accounts/Tranches Currently Open",         "2"),
        ("Accounts/Tranches Closed Profitably",      "2"),
        ("Accounts/Tranches Closed Unprofitably",    "0"),
        ("Range of Net Returns of Accounts/Tranches Closed", "0.36% to 4.2%"),
    ],
    "Terms & Fees": [
        ("Investment Type",    "Managed Account"),
        ("Fee Structure",      "0% Annual / 20% Performance"),
        ("High Water Mark",    "Yes"),
        ("Lockup Period",      "None"),
        ("Liquidity",          "Daily"),
        ("Notional Funding",   "Yes"),
        ("Execution FCM",      "StoneX Financial"),
    ],
}

# ==============================================================================
# 13) Legal disclaimers & footer contact
# ==============================================================================
hcdisclaimer_text = (
    "UNTIL TKP IS OFFICIALLY OPENED TO OUTSIDE INVESTORS BY THE INTRODUCING BROKER, "
    "THE STRATEGY REMAINS PROPRIETARY AND THIS PAGE OR DESCRIPTION IS NOT A SOLICITATION TO INVEST. "
    "NO SUBSCRIPTION DOCUMENTS HAVE BEEN ISSUED, AND TKP WILL ONLY BECOME AVAILABLE ONCE THE IB "
    "PUBLISHES THE APPROPRIATE SUBSCRIPTION MATERIALS AND DECLARES THE PROGRAM OPEN FOR OUTSIDE INVESTMENT."
)

disclaimer_text = (
    "THE RISK OF LOSS IN COMMODITY INTEREST TRADING CAN BE SUBSTANTIAL. YOU SHOULD, THEREFORE, "
    "CAREFULLY CONSIDER WHETHER SUCH TRADING IS SUITABLE FOR YOU IN LIGHT OF YOUR FINANCIAL CONDITION. "
    "THE HIGH DEGREE OF LEVERAGE IN COMMODITY INTEREST TRADING MEANS INVESTMENTS SHOULD BE MADE WITH RISK "
    "CAPITAL ONLY. ALL INFORMATION ABOVE IS COMPILED WITH THE INTENTION OF BEING FULLY CORRECT, THOUGH THERE "
    "IS NO GUARANTEE ALL INFORMATION IS CORRECT AND COULD BE SUBJECT TO UNINTENTIONAL CLERICAL ITEMS. "
    "PAST PERFORMANCE IS NOT NECESSARILY INDICATIVE OF FUTURE RESULTS."
)
footer_contact = (
    "HUGHES & COMPANY LLC • NFA ID 0423388 • 330 Himmararshee, Ste 110, FTL, FL 33312 • 954-500-0500 • www.hughesandco.ltd"
)

# ==============================================================================
# 14) Helper: Build Plotly “NAV” figure
# ==============================================================================
def build_NAV_figure():
    fig = go.Figure(
        go.Scatter(
            x=NAV_df.index,
            y=NAV_df[NAV_col],
            mode="lines",
            line={"color": PRIMARY_COLOR},
            name="NAV"
        )
    )

    # Base layout configuration
    layout_config = {
        "title": {
            "text": "<u>Non-Compounded NAV Since Inception</u>",
            "x": 0.5,
            "xanchor": "center"
        },
        "template": "ggplot2",
        "plot_bgcolor": GREY_BG,
        "paper_bgcolor": WHITE_BG,
        "xaxis_title": "Date",
        "yaxis_title": "NAV",
        "autosize": True,
    }

    # Add secondary percentage axis if enabled
    if SHOW_PERCENTAGE_AXIS:
        # Get the primary y-axis range from the data
        nav_min = NAV_df[NAV_col].min()
        nav_max = NAV_df[NAV_col].max()
        nav_range = nav_max - nav_min
        
        # Calculate percentage range
        pct_min = ((nav_min - BASELINE_AMOUNT) / BASELINE_AMOUNT) * 100
        pct_max = ((nav_max - BASELINE_AMOUNT) / BASELINE_AMOUNT) * 100
        pct_range = pct_max - pct_min
        
        # Determine appropriate tick step based on range
        if pct_range > 50:
            tick_step = 10  # 10% increments
        elif pct_range > 20:
            tick_step = 5   # 5% increments
        elif pct_range > 10:
            tick_step = 2   # 2% increments
        else:
            tick_step = 1   # 1% increments
        
        # Generate percentage tick values
        pct_tick_start = (int(pct_min / tick_step) - 1) * tick_step
        pct_tick_end = (int(pct_max / tick_step) + 2) * tick_step
        pct_ticks = list(range(int(pct_tick_start), int(pct_tick_end) + tick_step, tick_step))
        
        # Convert percentage ticks to NAV values (for positioning)
        # NAV = BASELINE_AMOUNT * (1 + pct/100)
        nav_tick_values = [BASELINE_AMOUNT * (1 + pct / 100) for pct in pct_ticks]
        pct_tick_labels = [f"{pct:.0f}%" for pct in pct_ticks]
        
        # Create percentage axis config with explicit ticks
        layout_config["yaxis2"] = {
            "title": "Return (%)",
            "overlaying": "y",
            "side": "right",
            "tickmode": "array",
            "tickvals": nav_tick_values,  # NAV positions where ticks appear
            "ticktext": pct_tick_labels,   # Percentage labels to display
            "showgrid": False,
            "zeroline": False,
        }
        layout_config["margin"] = {"l": 40, "r": 70, "t": 40, "b": 40}
    else:
        # No secondary axis, use standard margins
        layout_config["margin"] = {"l": 40, "r": 10, "t": 40, "b": 40}

    fig.update_layout(**layout_config)
    fig.update_xaxes(showgrid=True)
    fig.update_yaxes(showgrid=True)
    return fig

# Safe helper function for NAV figure building (alternative to main function)
def safe_build_NAV_figure():
    """Safe version of NAV figure builder with error handling"""
    try:
        return build_NAV_figure()
    except Exception as e:
        print(f"Warning: Error building NAV figure: {e}")
        # Return a simple fallback figure
        fig = go.Figure()
        fig.add_annotation(
            text="Error loading NAV data",
            xref="paper", yref="paper",
            x=0.5, y=0.5, showarrow=False
        )
        fig.update_layout(
            title="Error Loading NAV Data",
            xaxis_title="Date",
            yaxis_title="Value"
        )
        return fig


# ==============================================================================
# 15) Helper: Build Plotly “Drawdown” figure
# ==============================================================================
def build_drawdown_figure():
    fig = go.Figure()
    # build NAV
    strat_nav = (1 + daily_returns).cumprod()
    # baseline-relative drawdown
    strat_max = strat_nav.cummax()
    strat_dd  = (strat_nav - strat_max) / baseline * 100

    fig.add_trace(go.Scatter(
        x=strat_dd.index,
        y=strat_dd,
        name=STRATEGY_NAME + " DD (baseline)",
        line={"color":PRIMARY_COLOR}
    ))

    for name, cum in bench_cum.items():
        bench_max = cum.cummax()
        bench_dd_baseline = (cum - bench_max) / baseline * 100
        fig.add_trace(go.Scatter(
            x=bench_dd_baseline.index,
            y=bench_dd_baseline.values,
            name=f"{name} DD (baseline)",
            opacity=0.6
        ))

    fig.update_layout(
        title="Drawdown vs Peak (baseline-denominator)",
        xaxis_title="Date",
        yaxis_title="Drawdown (%)"
    )
    return fig

# Safe helper function for drawdown figure building (alternative to main function)
def safe_build_drawdown_figure():
    """Safe version of drawdown figure builder with error handling"""
    try:
        return build_drawdown_figure()
    except Exception as e:
        print(f"Warning: Error building drawdown figure: {e}")
        # Return a simple fallback figure
        fig = go.Figure()
        fig.add_annotation(
            text="Error loading drawdown data",
            xref="paper", yref="paper",
            x=0.5, y=0.5, showarrow=False
        )
        fig.update_layout(
            title="Error Loading Drawdown Data",
            xaxis_title="Date",
            yaxis_title="Drawdown (%)"
        )
        return fig

# ==============================================================================
# 15) Construct the Dash App
# ==============================================================================
app = dash.Dash(
    __name__,
    external_stylesheets=[dbc.themes.BOOTSTRAP, "/assets/styles.css"],
    suppress_callback_exceptions=True,
    title="H&C – TKP",
)

tkp_admin_auth_manager = AdminAuthManager(load_tkp_admin_auth_settings(), session_key=TKP_SESSION_KEY)
configure_flask_session_secret(app.server, tkp_admin_auth_manager.settings)


def _tkp_admin_board_stats():
    """Latest completed date / row count for the Portal board, from the persisted secret state."""
    try:
        records = _load_fresh_secret_records()
    except Exception:
        records = None
    if not records:
        return "—", 0
    latest_date = "—"
    for row in reversed(records):
        if row.get("Date"):
            latest_date = row["Date"]
            break
    return latest_date, len(records)


def _build_tkp_date_status_label_children(latest_date: str) -> tuple[list, list]:
    """Desktop/mobile children for the top-right 'Data current to' status block."""
    return build_header_date_label_children_from_date(latest_date)


def serve_layout(records=None):
    if records is None:
        records = secret_table_records
    desktop_date_label, mobile_date_label = _build_tkp_date_status_label_children(DAILY_RETURNS_LATEST_DATE)
    return dbc.Container(
        id="page-container",
        fluid=True,        # ⇒ always 100% on xs, sm; constrained on md+ breakpoints
        className="py-4",
        children=[
            # ── Header ─────────────────────────────────────────────────────────
            *build_tearsheet_header_row(
                logo_src=logo_src,
                logo_alt=f"{tsd.HNC_LEGAL_NAME} Logo",
                firm_name=tsd.HNC_LEGAL_NAME,
                product_name="The Keymaker Program",
                desktop_label_children=desktop_date_label,
                mobile_label_children=mobile_date_label,
                grey_bg=GREY_BG,
            ),

            # ── Description ────────────────────────────────────────────────────
            html.Div(
                [
                    html.P(
                        f"{tsd.HNC_LEGAL_NAME} is an introducing brokerage firm with expertise in the futures options industry. ",
                        className="lead text-center",
                    ),
                    html.P(
                        "Principals: Daniel V. Hughes III | Inception: April 2023 | Products Traded: E-Mini Micro S&P 500 Options | Styles: Short Options",
                        className="text-center mb-5",
                    ),
                ],
                className="description",
            ),

            # ── NAV Chart ─────────────────────────────────────────────────────
            dcc.Graph(
                id="NAV-graph",
                figure=build_NAV_figure(),
                config={"displayModeBar": False, "responsive": True},
                style={
                    "width": "100%",
                    "maxWidth": "100%",  # Ensure it doesn't exceed page width
                    "maxHeight": "400px",  # Cap height to fit on page
                    "pageBreakInside": "avoid",
                    #"overflow": "hidden",  # Prevent overflow
                },
            ),
            html.P(
                "This chart visualizes the growth of a $150,000 investment from inception to today. "
                "NAV stands for Net Asset Value; it reflects the non-compounded performance, net of all fees.",
                className="text-center small",
                style={"marginTop": "4rem"}  # gives some breathing room
            ),

            html.P(
                "Please note that all percentages shown are relative to the initial amount invested. "
                "Also note that performance may vary depending on the time of entry due to the fixed-sizing nature of this strategy.",
                className="text-center small",
                style={"marginBottom": "3rem"}
            ),

            # ── Performance Summary ────────────────────────────────────────────
            html.H5("Performance Summary", className="text-center mb-2"),
            html.Div(
                id="monthly-calendar-container",
                children=dbc.Table(
                    [
                        html.Thead(
                            html.Tr([
                                html.Th(col, style={"backgroundColor": GREY_BG, "color": "#000"})
                                for col in monthly_df.columns
                            ])
                        ),
                        html.Tbody([
                            html.Tr([html.Td(monthly_df.iloc[i][col]) for col in monthly_df.columns])
                            for i in range(len(monthly_df))
                        ])
                    ],
                    bordered=True,
                    hover=True,
                    size="sm",
                    className="table-responsive mb-5",
                    style={"width": "95%", "margin": "0 auto", "pageBreakInside": "avoid"},
                ),
            ),

            # ── General and Sector Information Tables ──────────────────────────
dbc.Row(
    [
        # ── Left Side: Strategy Overview ─────────────────────────────
        dbc.Col(
            dbc.Card(
                [
                    dbc.CardHeader(html.H6("Strategy Overview", className="mb-0"), className=HEADER_ROW_CLASS),
                    dbc.CardBody(
                        dbc.Table(
                            [
                                # header
                                html.Thead(
                                    html.Tr([ html.Th("Strategy Description", colSpan=3, className=HEADER_ROW_CLASS) ]),
                                    className="bg-light"
                                ),

                                # body
                                html.Tbody([
                                    # description row
                                    html.Tr([
                                        html.Td(
                                            [
                                                html.P(
                                                    f"The Keymaker Program (TKP) is a unique offering by {tsd.HNC_LEGAL_NAME} which utilizes specific strike daily options on the S&P 500 Index. It is oriented to achieve long-biased stable returns through intraday scalping of a proprietarily selected Put strikes in the nearest expiring option chain of the Micro ES product suite, and is most active in Volatile environments. The strategy simultaneously was built to allow for Put assignments for underlying Micro Futures Contracts, writing proprietarily selected Call strikes in sequential fashion to mitigate both drawdown depth and duration regardless of market environment. TKP has been designed as a long term, positively performing, market-neutral offering, with daily visibility and liquidity."
                                                ),
                                            ],
                                            colSpan=3,
                                            style={
                                                "whiteSpace": "normal",
                                                "fontStyle": "italic",
                                            }
                                        )
                                    ]),

                                    # spacer
                                    html.Tr([
                                        html.Td("", colSpan=3, style={"height": LEFT_TABLE_GAPS})
                                    ]),
                                                html.Tr([
                                                    html.Th(
                                                    "Methodology",
                                                    colSpan=3,
                                                    #style={"backgroundColor": GREY_BG},  
                                                    className= "bg-light",
                                                    )
                                                ]),
                                                html.Tr([
                                                    html.Td("Trading Style"),
                                                    html.Td([
                                                        html.Tr([
                                                            html.Td([
                                                                html.Tr([
                                                                    html.Td(html.Span("✓ Mean Reversion", style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"}),),
                                                                ]),
                                                                html.Tr([
                                                                    html.Td(html.Span("✗ Technical", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}),),
                                                                ]),
                                                                html.Tr([
                                                                    html.Td(html.Span("✗ Fundamental", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}),),
                                                                ]),
                                                            ]),
                                                            html.Td(),
                                                        ]),
                                                    ]),
                                                    html.Td([
                                                        html.Tr([
                                                            html.Td([
                                                                html.Tr([
                                                                    html.Td(html.Span("✗ Breakout", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}),),
                                                                ]),
                                                                html.Tr([
                                                                    html.Td(html.Span("✓ Premium Collection", style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"}),),
                                                                ]),
                                                                html.Tr([
                                                                    html.Td(html.Span("✗ Arbitrage", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}),),
                                                                ]),
                                                            ]),
                                                            html.Td(),
                                                        ]),
                                                    ]),
                                                ]),
                                                html.Tr([
                                                    html.Td("Decision Making Style"),
                                                    html.Td([
                                                        html.Tr([
                                                            html.Td(html.Span("✓ Systematic", style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"})),
                                                            html.Td(),
                                                        ]),
                                                    ]),
                                                    html.Td([
                                                        html.Tr([
                                                            html.Td(html.Span("✗ Discretionary", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"})),
                                                            html.Td(),
                                                        ]),
                                                    ]),
                                                ]),
                                                html.Tr([
                                                    html.Td("Execution Style"),
                                                    html.Td([
                                                        html.Tr([
                                                            html.Td(html.Span("✓ Automated", style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"})),
                                                            html.Td(),
                                                        ]),
                                                    ]),
                                                    html.Td([
                                                        html.Tr([
                                                            html.Td(html.Span("✓ Manual", style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"})),
                                                            html.Td(),
                                                        ]),
                                                    ]),
                                                ]),
                                                html.Tr([
                                                    html.Td("Position Types"),
                                                    html.Td([
                                                        html.Tr([
                                                            html.Td([
                                                                html.Tr([
                                                                    html.Td(html.Span("✓ Straight Futures", style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"}),),
                                                                ]),
                                                                html.Tr([
                                                                    html.Td(html.Span("✗ Futures Spreads", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}),),
                                                                ]),
                                                                html.Tr([
                                                                    html.Td(html.Span("✗ Cross Product Spreads", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}),),
                                                                ]),
                                                            ]),
                                                            html.Td(),
                                                        ]),
                                                    ]),
                                                    html.Td([
                                                        html.Tr([
                                                            html.Td([
                                                                html.Tr([
                                                                    html.Td(html.Span("✓ Covered Options", style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"}),),
                                                                ]),
                                                                html.Tr([
                                                                    html.Td(html.Span("✓ Uncovered Options", style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"}),),
                                                                ]),
                                                                html.Tr([
                                                                    html.Td(html.Span("✗ Options Spreads", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}),),
                                                                ]),

                                                            ]),
                                                            html.Td(),
                                                        ]),
                                                    ]),
                                                ]),
                                                html.Tr([
                                                    html.Td("", colSpan=3, style={"height": LEFT_TABLE_GAPS}),
                                                ]),  # Blank row above Activity Profile
                                                html.Tr([
                                                    html.Th("Activity Profile", colSpan=3, className=HEADER_ROW_CLASS),
                                                ]),
                                                html.Tr([
                                                    html.Td("Trading Frequency"),
                                                    html.Td([
                                                        html.Tr([
                                                            html.Td("Contracts Traded/Year/Million", style={"text-decoration": "underline"}),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("✗ Low (<500 Contracts)", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"})),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("✓ Medium (500-2000 Contracts)", style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"})),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("✗ High (>2000 Contracts)", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"})),
                                                        ]),
                                                    ]),
                                                    html.Td([
                                                        html.Tr([
                                                            html.Td("Approximate Average", style={"text-decoration": "underline"}), 
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("--"), style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("1000"), style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"}),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("--"), style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}),
                                                        ]),
                                                    ]),
                                                ]),
                                                html.Tr([
                                                    html.Td("Holding Periods"),
                                                    html.Td([
                                                        html.Tr([
                                                            html.Td("Time Period", style={"text-decoration": "underline"}), 
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("✓ Intraday", style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"})),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("✓ 1-30 Days", style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"})),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("✗ 1-3 Months", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"})),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("✗ 4+ Months", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"})),
                                                        ]),
                                                    ]),
                                                    html.Td([
                                                        html.Tr([
                                                            html.Td("Percentage (Totals 100%)", style={"text-decoration": "underline"}), 
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("95 %"), style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"}),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("5 %"), style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"}),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("-- %"), style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("-- %"), style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}),
                                                        ]),
                                                    ]),
                                                ]),
                                            ]),
                                        ],
                                        striped=False,
                                        bordered=True,
                                        hover=True,
                                        size="sm",
                                        className="table-responsive",
                                    )
                                ),
                            ],
                            outline=True,
                            className="mb-4",
                        ),
                        width=6,
                    ),

                    # ── Right Side: Sector Information ─────────────────────
                    dbc.Col(
                        dbc.Card(
                            [
                                dbc.CardHeader(html.H6("Trading Universe & Risk Profile", className="mb-0"), className=HEADER_ROW_CLASS),
                                dbc.CardBody(
                                    dbc.Table(
                                        [
                                            html.Thead(
                                                html.Tr([
                                                    html.Th("Product Exchanges", colSpan=3, className=HEADER_ROW_CLASS),
                                                ])
                                            ),
                                            html.Tbody([
                                                html.Tr([
                                                    html.Td("North America"),
                                                    html.Td("Europe"),
                                                    html.Td("Asia/Pacific"),
                                                ]),
                                                html.Tr([
                                                    html.Td([
                                                        html.Tr([
                                                            html.Td(html.Span("✓ CME Group / MGX", style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"})),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("✗ ICE US", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"})),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("✗ CFE", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"})),
                                                        ]), 
                                                        html.Tr([
                                                            html.Td(html.Span("✗ LME", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"})),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("✗ NODAL", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"})),
                                                        ]),
                                                    ]),
                                                    html.Td([
                                                        html.Tr([
                                                            html.Td(html.Span("✗ ICE UK / Financial", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"})),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("✗ Eurex", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"})),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("✗ Euronext", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"})),
                                                        ]),
                                                        html.Tr([
                                                        html.Span("✗ DGCX", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}),
                                                        ]),
                                                    ]),
                                                    html.Td([
                                                        html.Tr([
                                                            html.Td(html.Span("✗ SGX", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"})),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("✗ HKFE", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"})),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("✗ OSE / TOCOM", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"})),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("✗ SAFEX", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"})),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("✗ Bursa Malaysia", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"})),
                                                        ]),
                                                    ]),
                                                ]),
                                                html.Tr([
                                                    html.Td("", colSpan=3, style={"height": RIGHT_TABLE_GAPS}),
                                                ]),  # Blank row above Futures Products Traded
                                                html.Tr([
                                                    html.Th("Futures Products Traded", colSpan=3, className=HEADER_ROW_CLASS),
                                                ]),
                                                html.Tr([
                                                    html.Td("Financial Instruments"),
                                                    html.Td("Agricultural Commodities"),
                                                    html.Td("Other Asset Classes"),
                                                ]),
                                                html.Tr([
                                                    html.Td([
                                                        html.Tr(html.Td(html.Span("✓ Equity Indices", style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"}))),
                                                        html.Tr(html.Td(html.Span("✗ Volatility Indices", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}))),
                                                        html.Tr(html.Td(html.Span("✗ Interest Rates", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}))),
                                                        html.Tr(html.Td(html.Span("✗ Currencies", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}))),
                                                    ]),
                                                    html.Td([
                                                        html.Tr(html.Td(html.Span("✗ Grains / Oilseeds", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}))),
                                                        html.Tr(html.Td(html.Span("✗ Softs", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}))),
                                                        html.Tr(html.Td(html.Span("✗ Dairy", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}))),
                                                        html.Tr(html.Td(html.Span("✗ Meats / Livestock", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}))),
                                                    ]),
                                                    html.Td([
                                                        html.Tr(html.Td(html.Span("✗ Metals", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}))),
                                                        html.Tr(html.Td(html.Span("✗ Renewable Fuels", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}))),
                                                        html.Tr(html.Td(html.Span("✗ Cryptocurrencies", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}))),
                                                    ]),
                                                ]),

                                                html.Tr([
                                                    html.Td("", colSpan=3, style={"height": RIGHT_TABLE_GAPS}),
                                                ]),  # Blank row above Risk Management
                                                
                                                html.Tr([
                                                    html.Th("Risk Management", colSpan=3, className=HEADER_ROW_CLASS),
                                                ]),
                                                
                                                html.Tr([
                                                    html.Td("Average Margin Usage"),
                                                    html.Td([
                                                        html.Tr([
                                                            html.Td("1.77 %"),
                                                        ]),
                                                    ]),
                                                    html.Td([
                                                        
                                                    ]),
                                                ]),

                                                html.Tr([
                                                    html.Td([
                                                        # Main label
                                                        html.Div("Exchange Margin Ratios"),
                                                        # Explanatory text in a smaller font
                                                        html.Small(
                                                            "This is not cost-bearing, but is a measure of the exchange-required minimum funds to be in the account versus the Nominal Trade Size (150 k)",
                                                            style={
                                                                "fontSize": "0.75rem",
                                                                "color": "#6c757d",
                                                                "marginTop": "0.25rem",  # a little breathing room
                                                                "display": "block"
                                                            }
                                                        ),
                                                    ]),
                                                    html.Td([
                                                        html.Div([
                                                            html.Div("Ranges", className="ratio-header"),
                                                            html.Div("% time in range (daily)", className="ratio-header"),
                                                            html.Div(html.Span("✓ 0-10 %", style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"}), className="ratio-cell"),
                                                            html.Div(html.Span("94.8 %", style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"}), className="ratio-cell"),
                                                            html.Div(html.Span("✓ 10-25 %", style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"}), className="ratio-cell"),
                                                            html.Div(html.Span("5.2 %", style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"}), className="ratio-cell"),
                                                            html.Div(html.Span("✗ 25-50 %", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}), className="ratio-cell"),
                                                            html.Div(html.Span("-- %", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}), className="ratio-cell"),
                                                            html.Div(html.Span("✗ 50 %+", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}), className="ratio-cell"),
                                                            html.Div(html.Span("-- %", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}), className="ratio-cell"),
                                                        ], className="ratio-grid"),
                                                    ], colSpan=2),
                                                ]),

                                                html.Tr([
                                                    html.Td("Risk Controls"),
                                                    html.Td([
                                                        html.Tr([
                                                            html.Td(html.Span("✗ Stop Losses", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}, id="stop-losses")),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("✓ VaR Considerations", style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"}, id="var-considerations")),
                                                        ]),
                                                        #html.Tr([
                                                        #    html.Td(html.Span("Fixed Stop Losses", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"})),
                                                        #]),
                                                    ]),
                                                    html.Td([
                                                        #html.Tr([
                                                        #    html.Td(html.Span("Trailing Stops", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"})),
                                                        #]),
                                                        #html.Tr([
                                                        #    html.Td(html.Span("Anti-Martingale", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"})),
                                                        #]),
                                                        html.Tr([
                                                            html.Td(html.Span("✗ Position Reductions", style={"color": SECONDARY_COLOR, "marginRight": "0.5rem"}, id="position-reductions")),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("✓ Position Offsets (Hedges)", style={"color": PRIMARY_COLOR, "marginRight": "0.5rem"}, id="position-hedges")),
                                                        ]),
                                                    ]),
                                                ]),

                                                # Tooltips for Risk Controls
                                                dbc.Tooltip(
                                                    "Mechanisms to limit potential losses in volatile markets.",
                                                    target="risk-controls",
                                                    placement="top",
                                                ),

                                                # Tooltip for Stop Losses
                                                dbc.Tooltip(
                                                    "Orders that close a position at a predefined price to cap losses.",
                                                    target="stop-losses",
                                                    placement="top",
                                                ),

                                                # Tooltip for VaR Considerations
                                                dbc.Tooltip(
                                                    "Statistical estimate of potential loss over a given period at a chosen confidence level.",
                                                    target="var-considerations",
                                                    placement="top",
                                                ),

                                                # Tooltip for Position Reductions
                                                dbc.Tooltip(
                                                    "Gradual decrease in position size to reduce exposure as risk increases.",
                                                    target="position-reductions",
                                                    placement="top",
                                                ),

                                                # Tooltip for Position Offsets (Hedges)
                                                dbc.Tooltip(
                                                    "Taking opposite or correlated positions to hedge against adverse moves.",
                                                    target="position-hedges",
                                                    placement="top",
                                                ),
                                                
                                                html.Tr([
                                                    html.Td("", colSpan=3, style={"height": RIGHT_TABLE_GAPS}),
                                                ]),  # Blank row above Transaction Fees
                                                html.Tr([
                                                    html.Th("Transaction Fees (per Contract)", colSpan=3, className=HEADER_ROW_CLASS),
                                                ]),
                                                html.Tr([
                                                    html.Td("Commission"),
                                                    html.Td("$0.20"),
                                                    html.Td(),
                                                ]),
                                                html.Tr([
                                                    html.Td("Exchange Fee"),
                                                    html.Td("$0.10"),
                                                    html.Td(),
                                                ]),
                                                html.Tr([
                                                    html.Td("NFA Fee"),
                                                    html.Td("$0.00"),
                                                    html.Td(),
                                                ]),
                                                html.Tr([
                                                    html.Td("Give Up Fee"),
                                                    html.Td("$0.00"),
                                                    html.Td(),
                                                ]),
                                                html.Tr([
                                                    html.Td(html.Strong("Total All-In Fees")),
                                                    html.Td(html.Strong("$0.30")),
                                                    html.Td(),
                                                ]),
                                                html.Tr([
                                                    html.Td(
                                                        html.Small(
                                                            "* Give up fee is waived if account is traded at StoneX Financial.",
                                                            style={"fontStyle": "italic", "color": "#6c757d"}
                                                        ),
                                                        colSpan=3
                                                    ),
                                                ]),
                                            ]),
                                        ],
                                        striped=False,
                                        bordered=True,
                                        hover=True,
                                        size="sm",
                                        className="table-responsive",
                                    )
                                ),
                            ],
                            outline=True,
                            className="mb-4",
                        ),
                        width=6,
                    ),
                ],
                justify="start",
                className="mb-2",
            ),

            # ── Metrics & Info ─────────────────────────────────────────────────────────────────────
            # CONDITIONAL LAYOUT: Toggle USE_SIDE_BY_SIDE_LAYOUT flag at top of file to switch layouts
            # Current: Side-by-side (Metrics left, Info right) vs New: Stacked (Metrics top, Info bottom)
        ] + (
            # LAYOUT OPTION 1: SIDE-BY-SIDE (current layout, width=6 each)
            [dbc.Row(
                [
                    # ── LEFT SIDE: Metrics + Drawdown ─────────────────────
                    dbc.Col(
                        [
                            # Performance Metrics
                            html.Div(
                                id="daily-perf-container",
                                children=dbc.Card(
                                    [
                                        dbc.CardHeader(html.H6("Performance Metrics", className="mb-0")),
                                        dbc.CardBody(
                                            dbc.Table.from_dataframe(
                                                daily_perf_df,
                                                striped=False,
                                                bordered=True,
                                                hover=True,
                                                size="sm",
                                                className="fixed-cols",
                                            )
                                        ),
                                    ],
                                    outline=True,
                                    className="mb-4",
                                ),
                            ),

                            # Maximum Drawdown Profile
                            dbc.Card(
                                [
                                    dbc.CardHeader(html.H6("Maximum Drawdown Profile", className="mb-0")),
                                    dbc.CardBody(
                                        dbc.Table.from_dataframe(
                                            max_dd_df,
                                            striped=False,
                                            bordered=True,
                                            hover=True,
                                            size="sm",
                                            className="fixed-cols",
                                        )
                                    ),
                                    dbc.CardFooter(
                                        html.Small(
                                            "Both TKP & SPXTR drawdown stats are reflective of the same $150,000 fixed nominal exposure at start of drawdown period.",
                                            className="text-muted fst-italic"
                                        )
                                    ),
                                ],
                                outline=True,
                                className="mb-4",
                            ),
                        ],
                        width=6,
                    ),

                    # ── RIGHT SIDE: Additional Information ─────────────────
                    dbc.Col(
                        dbc.Card(
                            [
                                dbc.CardHeader(html.H6("Investor Information", className="mb-0")),
                                dbc.CardBody(
                                    html.Div(
                                        [
                                            dbc.Table(
                                                [
                                                    html.Thead(
                                                        html.Tr([
                                                            html.Th("Terms & Fees"),
                                                            html.Th("Details"),
                                                        ])
                                                    ),
                                                    html.Tbody([
                                                        html.Tr([
                                                            html.Td(label),
                                                            html.Td(value)
                                                        ])
                                                        for label, value in grouped_info["Terms & Fees"]
                                                    ]),
                                                ],
                                                striped=False,
                                                bordered=True,
                                                hover=True,
                                                size="sm",
                                                className="mb-3"
                                            ),
                                            dbc.Table(
                                                [
                                                    html.Thead(
                                                        html.Tr([
                                                            html.Th("Account Stats"),
                                                            html.Th("Proprietary"),
                                                            html.Th("Client"),
                                                        ])
                                                    ),
                                                    html.Tbody([
                                                        html.Tr([
                                                            html.Td("Nominal Assets Being Traded in the Program"),
                                                            html.Td("$300,000"),
                                                            html.Td("0")
                                                        ]),
                                                        html.Tr([
                                                            html.Td("Total Accounts/Tranches Opened"),
                                                            html.Td("4"),
                                                            html.Td("0")
                                                        ]),
                                                        html.Tr([
                                                            html.Td("Accounts/Tranches Currently Open"),
                                                            html.Td("2"),
                                                            html.Td("0")
                                                        ]),
                                                        html.Tr([
                                                            html.Td("Accounts/Tranches Closed Profitably"),
                                                            html.Td("2"),
                                                            html.Td("0")
                                                        ]),
                                                        html.Tr([
                                                            html.Td("Accounts/Tranches Closed Unprofitably"),
                                                            html.Td("0"),
                                                            html.Td("0")
                                                        ]),
                                                        html.Tr([
                                                            html.Td("Range of Net Returns of Accounts/Tranches Closed"),
                                                            html.Td("0.36% to 4.2%"),
                                                            html.Td("N/A")
                                                        ]),
                                                    ]),
                                                ],
                                                striped=False,
                                                bordered=True,
                                                hover=True,
                                                size="sm",
                                                className="mb-3"
                                            ),
                                            html.P(
                                                "Other Notes:",
                                                className="small fw-bold mb-1 mt-2"
                                            ),
                                            html.P(
                                                "TKP allows for efficient, opportunistic deployments of capital in and out of the program in fixed nominal trading levels of $150,000 per tranche. The program will remain perpetually funded with permanent capital of the Introducing Broker in the form of a minimum of two tranches ($300,000 Nominal). The IB itself also has historically allocated more tranches, and closed tranches profitably, and plans on continuing in doing so, in what it considers opportunities for additional capital deployment based on drawdowns of the program itself, with expected recoveries. This capability is allowed for investors as well, with the announcement of any tranche opening or closure by/of the IB shared for complete disclosure and additional visibility for the benefit of all potential participants.",
                                                className="mt-2",
                                                style={"fontSize": "0.9rem"}
                                            ),
                                        ]
                                    )
                                ),
                            ],
                            outline=True,
                            className="mb-4",
                        ),
                        width=6,
                    ),
                ],
                justify="start",
                className="mb-2",
            )] if USE_SIDE_BY_SIDE_LAYOUT else
            # LAYOUT OPTION 2: STACKED (new layout, full width blocks)
            [
                # Row 1: Metrics + Drawdown (full width)
                dbc.Row(
                    dbc.Col(
                        [
                            # Performance Metrics
                            html.Div(
                                id="daily-perf-container",
                                children=dbc.Card(
                                    [
                                        dbc.CardHeader(html.H6("Performance Metrics", className="mb-0")),
                                        dbc.CardBody(
                                            dbc.Table.from_dataframe(
                                                daily_perf_df,
                                                striped=False,
                                                bordered=True,
                                                hover=True,
                                                size="sm",
                                                className="fixed-cols",
                                            )
                                        ),
                                    ],
                                    outline=True,
                                    className="mb-4",
                                ),
                            ),

                            # Maximum Drawdown Profile
                            dbc.Card(
                                [
                                    dbc.CardHeader(html.H6("Maximum Drawdown Profile", className="mb-0")),
                                    dbc.CardBody(
                                        dbc.Table.from_dataframe(
                                            max_dd_df,
                                            striped=False,
                                            bordered=True,
                                            hover=True,
                                            size="sm",
                                            className="fixed-cols",
                                        )
                                    ),
                                    dbc.CardFooter(
                                        html.Small(
                                            "Both TKP & SPXTR drawdown stats are reflective of the same $150,000 fixed nominal exposure at start of drawdown period.",
                                            className="text-muted fst-italic"
                                        )
                                    ),
                                ],
                                outline=True,
                                className="mb-4",
                            ),
                        ],
                        width=12,
                    ),
                    justify="start",
                    className="mb-2",
                ),
                # Row 2: Investor Information (full width)
                dbc.Row(
                    dbc.Col(
                        dbc.Card(
                            [
                                dbc.CardHeader(html.H6("Investor Information", className="mb-0")),
                                dbc.CardBody(
                                    html.Div(
                                        [
                                            dbc.Table(
                                                [
                                                    html.Thead(
                                                        html.Tr([
                                                            html.Th("Terms & Fees"),
                                                            html.Th("Details"),
                                                        ])
                                                    ),
                                                    html.Tbody([
                                                        html.Tr([
                                                            html.Td(label),
                                                            html.Td(value)
                                                        ])
                                                        for label, value in grouped_info["Terms & Fees"]
                                                    ]),
                                                ],
                                                striped=False,
                                                bordered=True,
                                                hover=True,
                                                size="sm",
                                                className="mb-3"
                                            ),
                                            dbc.Table(
                                                [
                                                    html.Thead(
                                                        html.Tr([
                                                            html.Th("Account Stats"),
                                                            html.Th("Proprietary"),
                                                            html.Th("Client"),
                                                        ])
                                                    ),
                                                    html.Tbody([
                                                        html.Tr([
                                                            html.Td("Nominal Assets Being Traded in the Program"),
                                                            html.Td("$300,000"),
                                                            html.Td("0")
                                                        ]),
                                                        html.Tr([
                                                            html.Td("Total Accounts/Tranches Opened"),
                                                            html.Td("4"),
                                                            html.Td("0")
                                                        ]),
                                                        html.Tr([
                                                            html.Td("Accounts/Tranches Currently Open"),
                                                            html.Td("2"),
                                                            html.Td("0")
                                                        ]),
                                                        html.Tr([
                                                            html.Td("Accounts/Tranches Closed Profitably"),
                                                            html.Td("2"),
                                                            html.Td("0")
                                                        ]),
                                                        html.Tr([
                                                            html.Td("Accounts/Tranches Closed Unprofitably"),
                                                            html.Td("0"),
                                                            html.Td("0")
                                                        ]),
                                                        html.Tr([
                                                            html.Td("Range of Net Returns of Accounts/Tranches Closed"),
                                                            html.Td("0.36% to 4.2%"),
                                                            html.Td("N/A")
                                                        ]),
                                                    ]),
                                                ],
                                                striped=False,
                                                bordered=True,
                                                hover=True,
                                                size="sm",
                                                className="mb-3"
                                            ),
                                            html.P(
                                                "Other Notes:",
                                                className="small fw-bold mb-1 mt-2"
                                            ),
                                            html.P(
                                                "TKP allows for efficient, opportunistic deployments of capital in and out of the program in fixed nominal trading levels of $150,000 per tranche. The program will remain perpetually funded with permanent capital of the Introducing Broker in the form of a minimum of two tranches ($300,000 Nominal). The IB itself also has historically allocated more tranches, and closed tranches profitably, and plans on continuing in doing so, in what it considers opportunities for additional capital deployment based on drawdowns of the program itself, with expected recoveries. This capability is allowed for investors as well, with the announcement of any tranche opening or closure by/of the IB shared for complete disclosure and additional visibility for the benefit of all potential participants.",
                                                className="mt-2",
                                                style={"fontSize": "0.9rem"}
                                            ),
                                        ]
                                    )
                                ),
                            ],
                            outline=True,
                            className="mb-4",
                        ),
                        width=12,
                    ),
                    justify="start",
                    className="mb-2",
                ),
            ]
        ) + [

            # ── Notional Funding Disclosure ────────────────────────────────────────────────────
            dbc.Row(
                dbc.Col(
                    dbc.Card(
                        [
                            dbc.CardHeader(html.H6("Special Notional Funding Disclosure", className="mb-0")),
                            dbc.CardBody([
                                html.P("Notional funding is allowed (partially funded accounts). This increases leverage and magnifies volatility, margin calls, and percentage returns (and losses).", className="mb-2"),
                                html.P("Performance fees are computed on Nominal Account Size (not cash equity).", className="mb-2"),
                                html.H6("Funding Level Impact on Returns:", className="mb-2"),
                                dbc.Table([
                                    html.Thead([
                                        html.Tr([
                                            html.Th("Funding Level"),
                                            html.Th("Leverage"),
                                            html.Th("Example: +10% Strategy Return"),
                                            html.Th("Example: -10% Strategy Return"),
                                        ])
                                    ]),
                                    html.Tbody([
                                        html.Tr([
                                            html.Td("100% Funded"),
                                            html.Td("1:1"),
                                            html.Td("+10.0%"),
                                            html.Td("-10.0%"),
                                        ]),
                                        html.Tr([
                                            html.Td("50% Funded"),
                                            html.Td("2:1"),
                                            html.Td("+20.0%"),
                                            html.Td("-20.0%"),
                                        ]),
                                        html.Tr([
                                            html.Td("25% Funded"),
                                            html.Td("4:1"),
                                            html.Td("+40.0%"),
                                            html.Td("-40.0%"),
                                        ]),
                                    ])
                                ], striped=True, bordered=True, hover=True, size="sm"),
                                html.P("Higher leverage increases both potential gains and losses. Consider your risk tolerance carefully.", className="mt-2 text-muted small"),
                                html.P(
                                    "For more detailed information regarding notional funding, please refer to the complete disclosure document.",
                                    className="mt-2 text-muted small fw-bold"
                                ),
                            ]),
                        ],
                        outline=True,
                        className="mb-4",
                    ),
                    width=12
                ),
                className="mb-4",
            ),

            # ── Secret Table (only visible when "e" clicked) ────────────────────────────────────
            html.Div(
                id="secret-table-container",
                children=[
                    # Daily Returns rows; fresh from JSON on every page load
                    dcc.Store(id="secret-data-store", data=records),

                    dbc.Row(
                        dbc.Col(
                            dbc.Card(
                                [
                                    dbc.CardHeader(html.H6("Daily Returns", className="mb-0")),
                                    dbc.CardBody([
                                        html.Label("Visible Columns", className="fw-bold small mb-1"),
                                        dcc.Dropdown(
                                            id="secret-col-picker",
                                            options=[{"label": c, "value": c} for c in secret_all_columns],
                                            value=[c for c in SECRET_DEFAULT_VISIBLE if c in secret_all_columns],
                                            multi=True,
                                            clearable=False,
                                            placeholder="Select columns\u2026",
                                            style={"marginBottom": "12px"},
                                        ),
                                        html.Div([
                                            html.Div([
                                                dbc.Button("Add Row", id="secret-add-btn", color="success", size="sm", className="me-2"),
                                                dbc.Button("Delete Last Row", id="secret-delete-last-btn", color="danger", size="sm", className="me-2"),
                                                dbc.Button("Show Calculations", id="secret-calc-btn", color="info", size="sm", className="me-2"),
                                                html.Span("View per page:", className="me-2 small", style={"lineHeight": "31px"}),
                                                dcc.Dropdown(
                                                    id="secret-page-size-picker",
                                                    options=[{"label": str(v), "value": v} for v in [50, 100, 150, 200, 250, 300, 350, 400]],
                                                    value=50,
                                                    clearable=False,
                                                    style={"width": "80px", "display": "inline-block"},
                                                ),
                                            ], style={"display": "inline-flex", "alignItems": "center"}),
                                            html.Div([
                                                dbc.Button("Export Excel", id="secret-export-btn", color="secondary", size="sm"),
                                                dcc.Download(id="secret-export-download"),
                                            ], style={"float": "right"}),
                                        ], className="mb-3", style={"display": "flex", "justifyContent": "space-between"}),
                                        dash_table.DataTable(
                                            id="secret-daily-table",
                                            columns=_secret_table_columns(
                                                [c for c in SECRET_DEFAULT_VISIBLE if c in secret_all_columns]
                                            ),
                                            data=records,
                                            sort_action="native",
                                            sort_mode="single",
                                            sort_by=[{"column_id": "Date", "direction": "desc"}],
                                            page_size=50,
                                            row_selectable=False,
                                            style_table={"overflowX": "auto"},
                                            style_cell={
                                                "textAlign": "right",
                                                "padding": "4px 8px",
                                                "fontSize": "12px",
                                                "fontFamily": "monospace",
                                                "whiteSpace": "nowrap",
                                            },
                                            style_cell_conditional=[
                                                {"if": {"column_id": "Date"}, "textAlign": "left"},
                                                {"if": {"column_id": "#Day"}, "textAlign": "center"},
                                                {"if": {"column_id": "# Trades"}, "textAlign": "center"},
                                            ],
                                            style_header={
                                                "backgroundColor": "#1a2a3a",
                                                "color": "white",
                                                "fontWeight": "bold",
                                                "fontSize": "11px",
                                                "textAlign": "center",
                                            },
                                            style_data_conditional=[
                                                {"if": {"row_index": "odd"}, "backgroundColor": "#f8f9fa"},
                                                {"if": {"filter_query": "{Perc. Net} > 0", "column_id": "Perc. Net"}, "color": "#198754"},
                                                {"if": {"filter_query": "{Perc. Net} < 0", "column_id": "Perc. Net"}, "color": "#dc3545"},
                                                {"if": {"filter_query": "{Cumm Perc. Net} > 0", "column_id": "Cumm Perc. Net"}, "color": "#198754"},
                                                {"if": {"filter_query": "{Cumm Perc. Net} < 0", "column_id": "Cumm Perc. Net"}, "color": "#dc3545"},
                                            ],
                                        ),
                                    ]),
                                ],
                                outline=True,
                                className="mb-4",
                            ),
                            width=12
                        ),
                        className="mb-4",
                    ),

                    # ── Add Row Modal ──
                    dbc.Modal([
                        dbc.ModalHeader("Add Row"),
                        dbc.ModalBody([
                            dbc.Label("Date"),
                            dbc.Input(id="secret-add-date", type="date",
                                      value=_default_add_row_date_str()),
                            dbc.Label("Plus500 Balance", className="mt-2"),
                            dbc.Input(id="secret-add-plus500", type="number", value=0),
                            dbc.Label("StoneX Balance", className="mt-2"),
                            dbc.Input(id="secret-add-balance", type="number", value=0),
                            dbc.Label("Deposit / Withdrawal", className="mt-2"),
                            dbc.Input(id="secret-add-deposit", type="number", value=0),
                            html.P(
                                "(negative number = withdrawal)",
                                className="small text-muted fst-italic mt-1 mb-0",
                            ),
                        ]),
                        dbc.ModalFooter([
                            dbc.Button("Save", id="secret-add-save", color="primary", size="sm"),
                            dbc.Button("Cancel", id="secret-add-cancel", color="secondary", size="sm"),
                        ]),
                    ], id="secret-add-modal", is_open=False, centered=True, size="sm"),

                    # ── Delete Confirm Modal ──────────────────────────────────
                    dbc.Modal([
                        dbc.ModalHeader(dbc.ModalTitle("Confirm Delete")),
                        dbc.ModalBody(
                            html.P(id="secret-delete-confirm-body",
                                   className="mb-0")
                        ),
                        dbc.ModalFooter([
                            dbc.Button("Delete", id="secret-delete-confirm-btn",
                                       color="danger", size="sm", className="me-2"),
                            dbc.Button("Cancel", id="secret-delete-cancel-btn",
                                       color="secondary", size="sm"),
                        ]),
                    ], id="secret-delete-confirm-modal", is_open=False, centered=True, size="sm"),

                    # ── Show Calculations Modal ──
                    dbc.Modal([
                        dbc.ModalHeader("Monthly Calculation Inspector"),
                        dbc.ModalBody([
                            # SECTION 1: inputs
                            dbc.Row([
                                dbc.Col([
                                    dbc.Label("Month", className="small fw-bold"),
                                    dcc.Dropdown(
                                        id="calc-month-picker",
                                        options=[{"label": m, "value": i} for i, m in enumerate(
                                            ["Jan","Feb","Mar","Apr","May","Jun",
                                             "Jul","Aug","Sep","Oct","Nov","Dec"], start=1)],
                                        value=datetime.today().month,
                                        clearable=False,
                                    ),
                                ], width=5),
                                dbc.Col([
                                    dbc.Label("Year", className="small fw-bold"),
                                    dbc.Input(id="calc-year-input", type="number",
                                              value=datetime.today().year, min=2020, max=2040),
                                ], width=4),
                                dbc.Col([
                                    dbc.Button("Show", id="calc-show-btn", color="primary",
                                               size="sm", className="mt-4"),
                                ], width=3),
                            ], className="mb-3"),
                            html.Hr(),
                            # SECTION 2: static formulas
                            html.H6("Formulas", className="fw-bold small"),
                            dbc.Row([
                                dbc.Col([
                                    html.Small("NAV-Based (Net)", className="fw-bold text-primary d-block mb-1"),
                                    html.Pre(
                                        "(month_end_NAV - prior_month_end_NAV)\n/ BASELINE * 100",
                                        className="bg-light p-2 rounded",
                                        style={"fontFamily": "monospace", "fontSize": "11px", "whiteSpace": "pre-wrap"},
                                    ),
                                ], width=4),
                                dbc.Col([
                                    html.Small("StoneX Gross (excl fees)", className="fw-bold d-block mb-1",
                                               style={"color": "#4a86c8"}),
                                    html.Pre(
                                        "(end_StoneX - start_StoneX - deposits)\n/ BASELINE * 100",
                                        className="bg-light p-2 rounded",
                                        style={"fontFamily": "monospace", "fontSize": "11px", "whiteSpace": "pre-wrap"},
                                    ),
                                ], width=4),
                                dbc.Col([
                                    html.Small("StoneX Net (incl fees)", className="fw-bold d-block mb-1",
                                               style={"color": "#2e7d32"}),
                                    html.Pre(
                                        "(end_StoneX - start_StoneX - deposits - fees)\n/ BASELINE * 100",
                                        className="bg-light p-2 rounded",
                                        style={"fontFamily": "monospace", "fontSize": "11px", "whiteSpace": "pre-wrap"},
                                    ),
                                ], width=4),
                            ], className="mb-1"),
                            html.Small(
                                "prior_month_end = last value before the 1st of the selected month "
                                "(or BASELINE if none). NAV result may be overridden. "
                                "StoneX balances are pre-fee; fees are applied separately to derive net performance.",
                                className="text-muted d-block mb-3",
                            ),
                            html.Hr(),
                            # SECTION 3 + 4: dynamic (populated by callback)
                            html.Div(id="calc-results-container"),
                        ]),
                        dbc.ModalFooter(
                            dbc.Button("Close", id="calc-close-btn", color="secondary", size="sm"),
                        ),
                    ], id="secret-calc-modal", is_open=False, centered=True, size="lg"),

                ],
                style={"display": "none"}
            ),

            # ── Public Daily Returns Table (collapsed by default, expand via header) ──
            dbc.Row(
                dbc.Col(
                    dbc.Card([
                        dbc.CardHeader(
                            html.Div([
                                html.H6("Daily Returns", className="mb-0 d-inline"),
                                dbc.Button(
                                    "Show ▾",
                                    id="public-daily-toggle-btn",
                                    color="link",
                                    size="sm",
                                    className="float-end p-0 text-decoration-none fw-bold",
                                    n_clicks=0,
                                ),
                            ]),
                        ),
                        dbc.Collapse(
                            id="public-daily-collapse",
                            is_open=False,
                            children=dbc.CardBody([
                                html.Div([
                                    html.Div([
                                        html.Span("View per page:", className="me-2 small",
                                                  style={"lineHeight": "31px"}),
                                        dcc.Dropdown(
                                            id="public-page-size-picker",
                                            options=[{"label": str(v), "value": v}
                                                     for v in [50, 100, 150, 200, 250, 300, 350, 400]],
                                            value=50,
                                            clearable=False,
                                            style={"width": "80px", "display": "inline-block"},
                                        ),
                                    ], style={"display": "inline-flex", "alignItems": "center"}),
                                    html.Div([
                                        dbc.Button("Export Excel", id="public-export-btn",
                                                   color="secondary", size="sm"),
                                        dcc.Download(id="public-export-download"),
                                    ], style={"float": "right"}),
                                ], className="mb-3",
                                   style={"display": "flex", "justifyContent": "space-between"}),
                                dash_table.DataTable(
                                    id="public-daily-table",
                                    columns=_build_table_columns(PUBLIC_DAILY_COLUMNS),
                                    data=[
                                        {c: r.get(c, "") for c in PUBLIC_DAILY_COLUMNS}
                                        for r in records
                                    ],
                                    sort_action="native",
                                    sort_mode="single",
                                    sort_by=[{"column_id": "Date", "direction": "desc"}],
                                    page_size=50,
                                    row_selectable=False,
                                    editable=False,
                                    style_table={"overflowX": "auto"},
                                    style_cell={
                                        "textAlign": "right",
                                        "padding": "4px 8px",
                                        "fontSize": "12px",
                                        "fontFamily": "monospace",
                                        "whiteSpace": "nowrap",
                                    },
                                    style_cell_conditional=[
                                        {"if": {"column_id": "Date"}, "textAlign": "left"},
                                        {"if": {"column_id": "#Day"}, "textAlign": "center"},
                                    ],
                                    style_header={
                                        "backgroundColor": "#1a2a3a",
                                        "color": "white",
                                        "fontWeight": "bold",
                                        "fontSize": "11px",
                                        "textAlign": "center",
                                    },
                                    style_data_conditional=[
                                        {"if": {"filter_query": "{Perc. Net} > 0",
                                                "column_id": "Perc. Net"},
                                         "color": "green"},
                                        {"if": {"filter_query": "{Perc. Net} < 0",
                                                "column_id": "Perc. Net"},
                                         "color": "red"},
                                    ],
                                ),
                            ]),
                        ),
                    ]),
                    width=12,
                ),
                className="mb-4",
            ),

            # ── H&C Disclaimer ────────────────────────────────────────────────────
            dbc.Row(
                dbc.Col(
                    html.P(hcdisclaimer_text, className="text-muted small"),
                    width=12
                ),
                className="mb-4",
            ),
            
            # ── General Disclaimer ────────────────────────────────────────────────────
            dbc.Row(
                dbc.Col(
                    html.P(disclaimer_text, className="text-muted small"),
                    width=12
                ),
                className="mb-4",
            ),

            # ── Debug / Data Provenance (only when DEBUG_PROVENANCE) ───────────
            *([] if not DEBUG_PROVENANCE else [
                dbc.Row(
                    dbc.Col(
                        html.Div([
                            html.H6("Debug / Data Provenance", className="text-muted mb-2"),
                            dbc.Table(
                                [
                                    html.Thead(html.Tr([html.Th("Field name"), html.Th("Source")])),
                                    html.Tbody([
                                        html.Tr([html.Td("Monthly returns"), html.Td("computed from NAV unless in override_months")]),
                                        html.Tr([html.Td("Daily perf metrics"), html.Td("computed from daily_returns")]),
                                        html.Tr([html.Td("Drawdown profile"), html.Td("computed from strategy_nav and spxtr_nav")]),
                                        html.Tr([html.Td("Terms & Fees / Account Stats"), html.Td("hard-coded (from grouped_info)")]),
                                        html.Tr([html.Td("Risk Management (margin usage, exchange margin ratios)"), html.Td("hard-coded in layout")]),
                                    ]),
                                ],
                                bordered=True,
                                size="sm",
                                className="mb-0",
                            ),
                        ], style={"fontSize": "0.85rem"}),
                    width=12),
                    className="mb-4",
                ),
            ]),

            # Important Disclosure section (proprietary tier — bottom panel)
            dbc.Row(
                dbc.Col(
                    html.Div(
                        tsd.proprietary_bottom_disclosure_children("TKP"),
                        className=tsd.DISCLOSURE_PANEL_CLASS,
                        style=tsd.DISCLOSURE_PANEL_STYLE,
                    ),
                    width=12,
                ),
                className="mb-4",
            ),

            # ── Toggle & Footer ───────────────────────────────────────────────
            dbc.Row(
                dbc.Col(html.P(footer_contact, className="text-center small text-muted"), width=12),
                className="mb-2",
            ),
        ],
    )

dcc_store = dcc.Store(id="disclaimer-accepted", storage_type="session")
# "standard" = Accept & Continue; "secret" = last letter of "Notice" (same UI for now; branch later via this store)
access_mode_store = dcc.Store(id="access-mode", storage_type="session", data=None)

# Accept gate — proprietary tier (no strategy inquiry contact on gate)
disclaimer_screen = build_sibling_accept_gate(
    "TKP", extra_children=[build_gate_password_row(portal_enabled=False)]
)

# Make layout dynamic - this function is called on every page load
# This ensures fresh data is loaded when the app restarts
def dynamic_layout():
    """Generate layout with fresh data on each page load.

    Reads the JSON persistence file every time so that rows added/deleted since
    server startup are immediately visible after a browser refresh.
    """
    fresh_records = _load_fresh_secret_records()
    fresh_canonical = _canonical_records_from_secret_rows(fresh_records) or CANONICAL_NAV_RECORDS_INITIAL
    return html.Div([
        dcc_store,
        access_mode_store,
        dcc.Store(id="canonical-nav-store", storage_type="memory", data=fresh_canonical),
        dcc.Store(id=GATE_PASSWORD_VISIBLE_STORE_ID, storage_type="memory", data=False),
        dcc.Location(id="url", refresh=False),
        disclaimer_screen,
        html.Div(
            id="main-app",
            style={"display": "none"},
            children=serve_layout(records=fresh_records)
        )
    ])

# Set layout as a function for dynamic data loading
app.layout = dynamic_layout

@app.callback(
    Output("disclaimer-screen", "style"),
    Output("main-app", "style"),
    Output("access-mode", "data"),
    Input("accept-button", "n_clicks"),
)
def show_main(n_accept):
    if n_accept:
        return {"display": "none"}, {"display": "block"}, "standard"
    return tsd.GATE_SCREEN_STYLE, {"display": "none"}, None


# ── Hidden admin reveal: "e" click opens the password row (no access granted yet) ──
@app.callback(
    Output(GATE_PASSWORD_VISIBLE_STORE_ID, "data"),
    Input("secret-notice-e", "n_clicks"),
    State(GATE_PASSWORD_VISIBLE_STORE_ID, "data"),
    prevent_initial_call=True,
)
def _toggle_gate_password_row(n_clicks, visible):
    if n_clicks:
        return not bool(visible)
    return dash.no_update


@app.callback(
    Output(GATE_PASSWORD_ROW_ID, "style"),
    Input(GATE_PASSWORD_VISIBLE_STORE_ID, "data"),
)
def _render_gate_password_row(visible):
    return gate_password_row_style(bool(visible))


@app.callback(
    Output(GATE_PASSWORD_INPUT_ID, "value", allow_duplicate=True),
    Input(GATE_PASSWORD_VISIBLE_STORE_ID, "data"),
    prevent_initial_call=True,
)
def _clear_password_when_hidden(visible):
    if not visible:
        return ""
    return dash.no_update


@app.callback(
    Output(GATE_PASSWORD_ERROR_ID, "children"),
    Output(GATE_PASSWORD_VISIBLE_STORE_ID, "data", allow_duplicate=True),
    Output(GATE_PASSWORD_INPUT_ID, "value", allow_duplicate=True),
    Output("disclaimer-screen", "style", allow_duplicate=True),
    Output("main-app", "style", allow_duplicate=True),
    Output("access-mode", "data", allow_duplicate=True),
    Input(GATE_PASSWORD_SUBMIT_ID, "n_clicks"),
    Input(GATE_PASSWORD_INPUT_ID, "n_submit"),
    State(GATE_PASSWORD_INPUT_ID, "value"),
    prevent_initial_call=True,
)
def _gate_admin_tearsheet_login(_submit_clicks, _n_submit, password):
    ok, _msg = tkp_admin_auth_manager.login(session, password or "")
    if not ok:
        return INVALID_PASSWORD_MESSAGE, dash.no_update, "", dash.no_update, dash.no_update, dash.no_update
    return "", False, "", {"display": "none"}, {"display": "block"}, "secret"


@app.callback(
    Output(GATE_PASSWORD_ERROR_ID, "children", allow_duplicate=True),
    Output(GATE_PASSWORD_VISIBLE_STORE_ID, "data", allow_duplicate=True),
    Output(GATE_PASSWORD_INPUT_ID, "value", allow_duplicate=True),
    Output("url", "href"),
    Output("url", "refresh"),
    Input(GATE_PASSWORD_PORTAL_ID, "n_clicks"),
    State(GATE_PASSWORD_INPUT_ID, "value"),
    prevent_initial_call=True,
)
def _gate_admin_portal_login(_portal_clicks, password):
    ok, _msg = tkp_admin_auth_manager.login(session, password or "")
    if not ok:
        return INVALID_PASSWORD_MESSAGE, dash.no_update, "", dash.no_update, dash.no_update
    return "", False, "", ADMIN_PORTAL_PATH, True


@app.server.route("/admin")
def tkp_admin_portal():
    if not tkp_admin_auth_manager.is_authenticated(session):
        return redirect("/")
    latest_date, row_count = _tkp_admin_board_stats()
    diagnostics_html = render_legacy_diagnostics_table(
        program_name="TKP",
        latest_date=latest_date,
        row_count=row_count,
        daily_entry_href="/",
    )
    return render_portal_page(
        program_name="TKP",
        accounts=[],  # no participating-account registry for TKP yet -> Pending
        diagnostics_html=diagnostics_html,
    )


@app.server.route("/admin/logout")
def tkp_admin_logout():
    tkp_admin_auth_manager.logout(session)
    return redirect("/")


@app.server.route("/healthz")
def tkp_healthz():
    ready = full_daily_df is not None and not full_daily_df.empty
    return jsonify({
        "app": "tkp",
        "status": "ready" if ready else "error",
        "rows_loaded": int(len(full_daily_df)) if full_daily_df is not None else 0,
        "admin_auth": "configured" if tkp_admin_auth_manager.is_configured else "not_configured",
    })


@app.callback(
    Output("secret-table-container", "style"),
    Input("access-mode", "data"),
)
def toggle_secret_table(access_mode):
    if access_mode == "secret":
        return {"display": "block"}
    return {"display": "none"}

# ── Public Daily Returns: sync data from admin store ──────────────────────
@app.callback(
    Output("public-daily-table", "data"),
    Input("secret-data-store", "data"),
)
def sync_public_table(store_data):
    if not store_data:
        return []
    return [{c: r.get(c, "") for c in PUBLIC_DAILY_COLUMNS} for r in store_data]


# ── Public Daily Returns: page size ───────────────────────────────────────
@app.callback(
    Output("public-daily-table", "page_size"),
    Input("public-page-size-picker", "value"),
)
def update_public_page_size(page_size):
    return page_size or 50


# ── Public Daily Returns: toggle expand/collapse ──────────────────────────
@app.callback(
    Output("public-daily-collapse", "is_open"),
    Output("public-daily-toggle-btn", "children"),
    Input("public-daily-toggle-btn", "n_clicks"),
    State("public-daily-collapse", "is_open"),
    prevent_initial_call=True,
)
def toggle_public_daily(n_clicks, is_open):
    new_open = not is_open
    label = "Hide ▴" if new_open else "Show ▾"
    return new_open, label


# ── Public Daily Returns: export Excel ────────────────────────────────────
@app.callback(
    Output("public-export-download", "data"),
    Input("public-export-btn", "n_clicks"),
    State("public-daily-table", "data"),
    prevent_initial_call=True,
)
def export_public_excel(n_clicks, table_data):
    if not n_clicks or not table_data:
        return dash.no_update
    export_df = pd.DataFrame(table_data)
    return dcc.send_data_frame(export_df.to_excel, "daily_returns.xlsx", index=False)


# ── Column visibility ──────────────────────────────────────────────────────
@app.callback(
    Output("secret-daily-table", "columns"),
    Input("secret-col-picker", "value"),
)
def update_secret_columns(selected):
    if not selected:
        selected = list(SECRET_DEFAULT_VISIBLE)
    ordered = [c for c in SECRET_DEFAULT_VISIBLE if c in selected]
    ordered += [c for c in secret_all_columns if c in selected and c not in ordered]
    return _secret_table_columns(ordered)

# ── Sync table data from store ─────────────────────────────────────────────
@app.callback(
    Output("secret-daily-table", "data"),
    Input("secret-data-store", "data"),
)
def sync_table_data(store_data):
    return store_data or []

# ── Page size control ──────────────────────────────────────────────────────
@app.callback(
    Output("secret-daily-table", "page_size"),
    Input("secret-page-size-picker", "value"),
)
def update_page_size(page_size):
    return page_size or 50

# ── Add Row: open / close modal ───────────────────────────────────────────
@app.callback(
    Output("secret-add-modal", "is_open"),
    Input("secret-add-btn", "n_clicks"),
    Input("secret-add-cancel", "n_clicks"),
    Input("secret-add-save", "n_clicks"),
    State("secret-add-modal", "is_open"),
    prevent_initial_call=True,
)
def toggle_add_modal(n_open, n_cancel, n_save, is_open):
    ctx = dash.callback_context
    tid = ctx.triggered[0]["prop_id"].split(".")[0]
    if tid == "secret-add-btn":
        return True
    return False

# ── Add Row: save to store ────────────────────────────────────────────────
@app.callback(
    Output("secret-data-store", "data", allow_duplicate=True),
    Input("secret-add-save", "n_clicks"),
    State("secret-add-date", "value"),
    State("secret-add-plus500", "value"),
    State("secret-add-balance", "value"),
    State("secret-add-deposit", "value"),
    State("secret-data-store", "data"),
    prevent_initial_call=True,
)
def add_row(n_clicks, date_val, plus500_val, balance_val, deposit_val, current_data):
    if not n_clicks or not balance_val:
        return dash.no_update
    rows = list(current_data) if current_data else []
    max_id  = max((r.get("_row_id", 0) for r in rows), default=-1) + 1
    max_day = max((int(r["#Day"]) for r in rows if str(r.get("#Day", "")).isdigit()), default=0) + 1

    new_balance = float(balance_val)
    deposit     = float(deposit_val) if deposit_val else 0.0
    plus500     = float(plus500_val) if plus500_val else 0.0

    sorted_rows = sorted(rows, key=lambda r: int(r["#Day"]) if str(r.get("#Day", "")).isdigit() else 0)
    prev_row    = sorted_rows[-1] if sorted_rows else {}

    computed = _compute_new_row(prev_row, new_balance, deposit)

    new_row = {c: "" for c in secret_all_columns}
    new_row.update(computed)
    new_row["_row_id"]    = max_id
    new_row["#Day"]       = str(max_day)
    new_row["Date"]       = date_val or ""
    new_row["Plus500"]    = f"${plus500:,.2f}" if plus500 != 0 else ""
    new_row["# Trades"]   = ""
    rows.append(new_row)
    _save_secret_editor_state(rows)
    return rows

# ── Delete Last Row: open confirm modal ───────────────────────────────────
@app.callback(
    Output("secret-delete-confirm-modal", "is_open"),
    Output("secret-delete-confirm-body", "children"),
    Input("secret-delete-last-btn", "n_clicks"),
    Input("secret-delete-cancel-btn", "n_clicks"),
    Input("secret-delete-confirm-btn", "n_clicks"),
    State("secret-data-store", "data"),
    State("secret-delete-confirm-modal", "is_open"),
    prevent_initial_call=True,
)
def toggle_delete_confirm_modal(n_open, n_cancel, n_confirm, current_data, is_open):
    ctx = dash.callback_context
    tid = ctx.triggered[0]["prop_id"].split(".")[0]
    if tid == "secret-delete-last-btn":
        # Build a description of the row about to be deleted
        body = "Are you sure you want to delete the last row?"
        if current_data:
            valid = [r for r in current_data if r.get("Date")]
            if valid:
                latest = max(valid, key=lambda r: r["Date"])
                date = latest.get("Date", "?")
                nav  = latest.get("NAV", "?")
                body = f"Delete row for {date} (NAV: {nav})? This cannot be undone."
        return True, body
    return False, dash.no_update


# ── Delete Last Row: execute on confirmation ───────────────────────────────
@app.callback(
    Output("secret-data-store", "data", allow_duplicate=True),
    Input("secret-delete-confirm-btn", "n_clicks"),
    State("secret-data-store", "data"),
    prevent_initial_call=True,
)
def delete_last_row(n_clicks, current_data):
    if not n_clicks or not current_data:
        return dash.no_update
    rows = list(current_data)
    valid = [r for r in rows if r.get("Date")]
    if not valid:
        return dash.no_update
    latest_id = max(valid, key=lambda r: r["Date"]).get("_row_id")
    updated = [r for r in rows if r.get("_row_id") != latest_id]
    _save_secret_editor_state(updated)
    return updated

# ── Show Calculations: open / close modal ─────────────────────────────────
@app.callback(
    Output("secret-calc-modal", "is_open"),
    Input("secret-calc-btn", "n_clicks"),
    Input("calc-close-btn", "n_clicks"),
    State("secret-calc-modal", "is_open"),
    prevent_initial_call=True,
)
def toggle_calc_modal(n_open, n_close, is_open):
    ctx = dash.callback_context
    tid = ctx.triggered[0]["prop_id"].split(".")[0]
    if tid == "secret-calc-btn":
        return True
    return False

# ── Show Calculations: compute and display ────────────────────────────────
@app.callback(
    Output("calc-results-container", "children"),
    Input("calc-show-btn", "n_clicks"),
    State("calc-month-picker", "value"),
    State("calc-year-input", "value"),
    State("canonical-nav-store", "data"),
    State("secret-data-store", "data"),
    prevent_initial_call=True,
)
def show_monthly_calc(n_clicks, month, year, canonical_rows, secret_rows):
    if not n_clicks or not month or not year or not canonical_rows:
        return dash.no_update

    month_names = ["Jan","Feb","Mar","Apr","May","Jun",
                   "Jul","Aug","Sep","Oct","Nov","Dec"]
    period_label = f"{month_names[int(month)-1]} {int(year)}"
    target_period = pd.Period(f"{int(year)}-{int(month):02d}", freq="M")
    bl = BASELINE_AMOUNT

    # ── NAV section ───────────────────────────────────────────────────────
    nav_s = _rebuild_nav_series(canonical_rows)
    if nav_s.empty:
        return html.P("No NAV data available.", className="text-danger")

    in_month_nav = nav_s[(nav_s.index >= target_period.start_time) &
                         (nav_s.index <= target_period.end_time)]
    if in_month_nav.empty:
        return html.P(f"No data found for {period_label}.", className="text-warning")

    month_end_nav = in_month_nav.iloc[-1]
    before_nav = nav_s[nav_s.index < target_period.start_time]
    prior_nav = before_nav.iloc[-1] if len(before_nav) > 0 else bl
    nav_computed = (month_end_nav - prior_nav) / bl * 100

    is_overridden = target_period in override_months
    nav_final = override_months[target_period] if is_overridden else nav_computed

    nav_values = dbc.Table([html.Tbody([
        html.Tr([html.Td("Period", className="fw-bold"), html.Td(period_label)]),
        html.Tr([html.Td("BASELINE", className="fw-bold"), html.Td(f"${bl:,.2f}")]),
        html.Tr([html.Td("Prior month end NAV", className="fw-bold"),
                  html.Td(f"${prior_nav:,.2f}  ({before_nav.index[-1].strftime('%Y-%m-%d') if len(before_nav) > 0 else 'N/A — using baseline'})")]),
        html.Tr([html.Td("Month end NAV", className="fw-bold"),
                  html.Td(f"${month_end_nav:,.2f}  ({in_month_nav.index[-1].strftime('%Y-%m-%d')})")]),
        html.Tr([html.Td("Data points", className="fw-bold"), html.Td(str(len(in_month_nav)))]),
    ])], bordered=True, size="sm", className="mb-2",
        style={"fontFamily": "monospace", "fontSize": "12px"})

    nav_calc = html.Pre(
        f"({month_end_nav:,.1f} - {prior_nav:,.1f}) / {bl:,.1f} * 100 = {nav_computed:.4f}%",
        className="bg-light p-2 rounded mb-1",
        style={"fontFamily": "monospace", "fontSize": "12px"})

    nav_override = html.P(
        f"Override active: forced to {nav_final:.4f}% (computed {nav_computed:.4f}%)",
        className="text-info small fw-bold") if is_overridden else None

    nav_result = dbc.Alert(
        f"NAV Result:  {nav_final:.4f}%",
        color="success" if nav_final >= 0 else "danger",
        className="text-center fw-bold mb-0")

    # ── StoneX sections (Gross and Net) ─────────────────────────────────────
    sx_s, dep_s, fee_s = _extract_stonex_deposit_fee_series(secret_rows)
    sx_gross_section = []
    sx_net_section = []

    if not sx_s.empty:
        in_month_sx = sx_s[(sx_s.index >= target_period.start_time) &
                           (sx_s.index <= target_period.end_time)]
        in_month_dep = dep_s[(dep_s.index >= target_period.start_time) &
                             (dep_s.index <= target_period.end_time)]
        in_month_fee = fee_s[(fee_s.index >= target_period.start_time) &
                             (fee_s.index <= target_period.end_time)]
        before_sx = sx_s[sx_s.index < target_period.start_time]

        if not in_month_sx.empty:
            sx_end = in_month_sx.iloc[-1]
            sx_start = before_sx.iloc[-1] if len(before_sx) > 0 else bl
            net_dep = in_month_dep.sum() if not in_month_dep.empty else 0.0
            total_fee = in_month_fee.sum() if not in_month_fee.empty else 0.0
            gross_pnl = sx_end - sx_start - net_dep
            net_pnl = gross_pnl - total_fee
            sx_gross = gross_pnl / bl * 100
            sx_net_val = net_pnl / bl * 100

            # Shared values table
            sx_values = dbc.Table([html.Tbody([
                html.Tr([html.Td("Prior month end StoneX", className="fw-bold"),
                          html.Td(f"${sx_start:,.2f}  ({before_sx.index[-1].strftime('%Y-%m-%d') if len(before_sx) > 0 else 'N/A — using baseline'})")]),
                html.Tr([html.Td("Month end StoneX", className="fw-bold"),
                          html.Td(f"${sx_end:,.2f}  ({in_month_sx.index[-1].strftime('%Y-%m-%d')})")]),
                html.Tr([html.Td("Net deposits in month", className="fw-bold"),
                          html.Td(f"${net_dep:,.2f}")]),
                html.Tr([html.Td("Fees in month", className="fw-bold"),
                          html.Td(f"${total_fee:,.2f}")]),
                html.Tr([html.Td("Data points", className="fw-bold"),
                          html.Td(str(len(in_month_sx)))]),
            ])], bordered=True, size="sm", className="mb-2",
                style={"fontFamily": "monospace", "fontSize": "12px"})

            # Gross (excl fees) — 1 decimal place for cleaner display
            sx_gross_calc = html.Pre(
                f"Gross P&L = {sx_end:,.1f} - {sx_start:,.1f} - {net_dep:,.1f} = ${gross_pnl:,.1f}\n"
                f"Gross % = {gross_pnl:,.1f} / {bl:,.1f} * 100 = {sx_gross:.4f}%",
                className="bg-light p-2 rounded mb-1",
                style={"fontFamily": "monospace", "fontSize": "11px"})

            sx_gross_result = dbc.Alert(
                f"StoneX Gross (excl fees):  {sx_gross:.4f}%",
                color="success" if sx_gross >= 0 else "danger",
                className="text-center fw-bold mb-1", style={"fontSize": "13px"})

            # Net (incl fees)
            sx_net_calc = html.Pre(
                f"Net P&L = {gross_pnl:,.1f} - {total_fee:,.1f} = ${net_pnl:,.1f}\n"
                f"Net % = {net_pnl:,.1f} / {bl:,.1f} * 100 = {sx_net_val:.4f}%",
                className="bg-light p-2 rounded mb-1",
                style={"fontFamily": "monospace", "fontSize": "11px"})

            sx_net_result = dbc.Alert(
                f"StoneX Net (incl fees):  {sx_net_val:.4f}%",
                color="success" if sx_net_val >= 0 else "danger",
                className="text-center fw-bold mb-1", style={"fontSize": "13px"})

            # Difference line vs NAV — flags reconciliation issues immediately
            diff = abs(nav_final - sx_net_val)
            MATCH_THRESHOLD = 0.05  # within 0.05% is considered a match
            if diff <= MATCH_THRESHOLD:
                diff_label = f"Difference (NAV vs StoneX Net):  {nav_final - sx_net_val:+.4f}%  ✓ Matches NAV"
                diff_color = "success"
            else:
                diff_label = f"Difference (NAV vs StoneX Net):  {nav_final - sx_net_val:+.4f}%  ⚠ Divergence detected"
                diff_color = "danger"
            diff_badge = dbc.Alert(
                diff_label,
                color=diff_color,
                className="text-center small mb-0",
                style={"fontSize": "11px", "opacity": "0.85"})

            sx_shared_values = sx_values
            sx_gross_section = [sx_gross_calc, sx_gross_result]
            sx_net_section = [sx_net_calc, sx_net_result, diff_badge]
        else:
            sx_shared_values = None
            sx_gross_section = [html.P(f"No StoneX data for {period_label}.", className="text-muted small")]
            sx_net_section = []
    else:
        sx_shared_values = None
        sx_gross_section = [html.P("StoneX data not available.", className="text-muted small")]
        sx_net_section = []

    # ── Assemble in 3 equal columns ─────────────────────────────────────────
    # Left: NAV-Based (Net)
    nav_col_children = [
        html.H6("NAV-Based (Net)", className="fw-bold small text-primary"),
        nav_values, nav_calc,
    ]
    if nav_override:
        nav_col_children.append(nav_override)
    nav_col_children.append(nav_result)

    # Center: StoneX Net (incl fees)
    sx_net_col_children = [
        html.H6("StoneX Net (incl fees)", className="fw-bold small", style={"color": "#2e7d32"}),
    ] + sx_net_section

    # Right: StoneX Gross (excl fees)
    sx_gross_col_children = [
        html.H6("StoneX Gross (excl fees)", className="fw-bold small", style={"color": "#4a86c8"}),
    ] + sx_gross_section

    # Shared StoneX values table above the 3 columns (if available)
    shared_row = []
    if sx_shared_values is not None:
        shared_row = [
            html.H6("StoneX Data for Period", className="fw-bold small text-secondary mt-2"),
            sx_shared_values,
            html.Hr(className="my-2"),
        ]

    return html.Div(shared_row + [
        dbc.Row([
            dbc.Col(nav_col_children, md=4, className="mb-3"),
            dbc.Col(sx_net_col_children, md=4, className="mb-3"),
            dbc.Col(sx_gross_col_children, md=4, className="mb-3"),
        ])
    ])

# ── Helper functions for auto-calculation (from CURSOR_PATCH.md) ──────────
def _parse_money(s):
    """'$1,234.56' → 1234.56; '' or None → 0.0"""
    try:
        return float(str(s).replace("$", "").replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0

def _parse_pct(s):
    """Stored as display % (e.g. 1.2345) → return decimal 0.012345 for calculations."""
    if isinstance(s, (int, float)) and not (s != s):  # excludes NaN
        return float(s) / 100
    try:
        cleaned = str(s).replace("%", "").strip()
        if not cleaned:
            return 0.0
        return float(cleaned) / 100
    except (ValueError, TypeError):
        return 0.0

def _compute_new_row(prev_row: dict, new_balance: float, deposit: float) -> dict:
    prev_balance    = _parse_money(prev_row.get("StoneX"))
    prev_nav        = _parse_money(prev_row.get("NAV"))
    prev_hwm_str    = str(prev_row.get("HWM", "")).replace(" *", "")
    prev_hwm        = _parse_money(prev_hwm_str) if prev_hwm_str else prev_nav
    prev_loss_carry = _parse_money(prev_row.get("Loss Carry"))
    prev_cumm_fee   = _parse_money(prev_row.get("Cumm Fee"))
    prev_cumm_pct   = _parse_pct(prev_row.get("Cumm Perc. Net"))

    pl         = new_balance - prev_balance - deposit
    fee        = max(0.0, (pl - prev_loss_carry) * 0.2) if pl > prev_loss_carry else 0.0
    cumm_fee   = prev_cumm_fee + fee
    net_pl     = pl - fee
    nav        = prev_nav + net_pl
    loss_carry = max(0.0, prev_hwm - nav)
    pct_net    = net_pl / BASELINE_AMOUNT
    cumm_pct   = prev_cumm_pct + pct_net
    hwm        = max(prev_hwm, nav)
    hwm_new_high = hwm > (prev_hwm + 0.01)

    return {
        "StoneX": f"${new_balance:,.2f}",
        "$PL":              f"${pl:,.2f}",
        "Fee (20%)":        f"${fee:,.2f}",
        "Cumm Fee":         f"${cumm_fee:,.2f}",
        "Net P&L":          f"${net_pl:,.2f}",
        "Net P&L / Unit":   f"${net_pl:,.2f}",
        "NAV":              f"${nav:,.2f}",
        "Loss Carry":       f"${loss_carry:,.2f}",
        "Perc. Net":        round(pct_net * 100, 6),
        "Cumm Perc. Net":   round(cumm_pct * 100, 6),
        "HWM":              f"${hwm:,.2f}" + (" *" if hwm_new_high else ""),
        "Deposit":          f"${deposit:,.0f}" if deposit != 0 else "",
    }

# ══════════════════════════════════════════════════════════════════════════
# RECALCULATION ENGINE — rebuild dashboard metrics from Daily Returns store
# ══════════════════════════════════════════════════════════════════════════

def _rebuild_nav_series(rows):
    """Extract (Date, NAV) pairs from store rows and return a date-indexed pd.Series."""
    pairs = []
    for r in rows:
        date_str = r.get("Date", "")
        nav_str = r.get("NAV", "")
        if not date_str or not nav_str:
            continue
        try:
            dt = pd.to_datetime(date_str)
            nav_val = _parse_money(nav_str)
            if nav_val > 0:
                pairs.append((dt, nav_val))
        except Exception:
            continue
    if not pairs:
        return pd.Series(dtype=float)
    df = pd.DataFrame(pairs, columns=["Date", "NAV"]).sort_values("Date")
    df = df.drop_duplicates(subset="Date", keep="last")
    return df.set_index("Date")["NAV"]

@app.callback(
    Output("canonical-nav-store", "data"),
    Input("secret-data-store", "data"),
    prevent_initial_call=True,
)
def sync_canonical_nav_store(secret_store_rows):
    if not secret_store_rows:
        return dash.no_update
    nav_s = _rebuild_nav_series(secret_store_rows)
    if nav_s.empty:
        return []
    return [
        {"Date": dt.strftime("%Y-%m-%d"), "NAV": float(v)}
        for dt, v in nav_s.items()
    ]


def _extract_stonex_deposit_fee_series(store_rows):
    """Extract date-indexed StoneX balance, Deposit, and Fee series from secret-data-store rows.
    
    Data model notes:
    - StoneX = actual brokerage balance (GROSS — does not reflect advisor fees)
    - Fee (20%) = performance fee charged that day (stored separately)
    - Deposit = cash transfer that day
    - NAV = net asset value (reflects fees already deducted)
    
    Therefore:
    - StoneX Excl. Fees (gross) = StoneX change minus deposits
    - StoneX Incl. Fees (net) = StoneX change minus deposits minus fees
    """
    if not store_rows:
        return pd.Series(dtype=float), pd.Series(dtype=float), pd.Series(dtype=float)
    rows = []
    for r in store_rows:
        d = r.get("Date", "")
        sx = r.get("StoneX", "")
        dep = r.get("Deposit", "")
        fee = r.get("Fee (20%)", "")
        if not d or not sx:
            continue
        try:
            dt = pd.to_datetime(d)
            sx_val = _parse_money(sx)
            dep_val = _parse_money(dep) if dep else 0.0
            fee_val = _parse_money(fee) if fee else 0.0
            rows.append((dt, sx_val, dep_val, fee_val))
        except Exception:
            continue
    if not rows:
        return pd.Series(dtype=float), pd.Series(dtype=float), pd.Series(dtype=float)
    df = pd.DataFrame(rows, columns=["Date", "StoneX", "Deposit", "Fee"]).sort_values("Date")
    df = df.drop_duplicates(subset="Date", keep="last").set_index("Date")
    return df["StoneX"], df["Deposit"], df["Fee"]


def _compute_stonex_monthly_gross_net(stonex_s, deposit_s, fee_s, bl):
    """Compute monthly returns from StoneX balances, returning both gross and net.
    
    Gross (Excl. Fees):
        (month_end_StoneX - prior_month_end_StoneX - net_deposits) / BASELINE * 100
        
    Net (Incl. Fees):
        (month_end_StoneX - prior_month_end_StoneX - net_deposits - fees_in_month) / BASELINE * 100
    
    Returns: (gross_series, net_series) — both indexed by Period
    """
    if stonex_s.empty or bl == 0:
        return pd.Series(dtype=float), pd.Series(dtype=float)
    mp = stonex_s.index.to_period("M")
    m_last = stonex_s.groupby(mp).last()
    m_first = pd.Series(index=m_last.index, dtype=float)
    m_dep = pd.Series(index=m_last.index, dtype=float)
    m_fee = pd.Series(index=m_last.index, dtype=float)
    for period in m_last.index:
        before = stonex_s[stonex_s.index < period.start_time]
        m_first.loc[period] = before.iloc[-1] if len(before) > 0 else bl
        in_month_dep = deposit_s[(deposit_s.index >= period.start_time) &
                                  (deposit_s.index <= period.end_time)]
        in_month_fee = fee_s[(fee_s.index >= period.start_time) &
                              (fee_s.index <= period.end_time)]
        m_dep.loc[period] = in_month_dep.sum()
        m_fee.loc[period] = in_month_fee.sum()
    gross = (m_last - m_first - m_dep) / bl * 100
    net = (m_last - m_first - m_dep - m_fee) / bl * 100
    return gross, net


def _recompute_monthly_records(nav_series, bl):
    """Rebuild monthly calendar records from a NAV series.
    
    Produces 1 row per year with official NAV-based results only.
    StoneX comparison is available in Show Calculations modal, not in main table.
    """
    if nav_series.empty or bl == 0:
        return []
    mp = nav_series.index.to_period("M")
    m_last = nav_series.groupby(mp).last()
    m_first = pd.Series(index=m_last.index, dtype=float)
    for period in m_last.index:
        before = nav_series[nav_series.index < period.start_time]
        m_first.loc[period] = before.iloc[-1] if len(before) > 0 else bl
    m_simple = (m_last - m_first) / bl * 100

    for op, ov in override_months.items():
        if op in m_simple.index:
            m_simple.loc[op] = ov

    yr_simple = m_simple.groupby(m_simple.index.year).sum()
    years = sorted(m_simple.index.year.unique())
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    rows = []
    for y in years:
        row = {"Year": str(y)}
        for idx, m in enumerate(month_names, start=1):
            p = pd.Period(f"{y}-{idx:02d}")
            row[m] = f"{m_simple.get(p, 0):.4f}%" if p in m_simple.index else ""
        row["Year Total"] = f"{yr_simple.get(y, 0):.4f}%"
        rows.append(row)
    return rows


def _recompute_daily_perf_records(nav_series, bl):
    """Rebuild daily performance metrics table from a NAV series. Returns list[dict]."""
    if nav_series.empty or bl == 0:
        return []
    d_returns = nav_series.diff().div(bl).dropna()
    i_start = nav_series.index.min()
    ttm = nav_series.index.max() - pd.DateOffset(years=1)
    oy_ret = d_returns.loc[ttm:].dropna()
    inc_ret = d_returns.copy()
    oy_met = calculate_period_metrics(oy_ret, ttm)
    inc_met = calculate_period_metrics(inc_ret, i_start)
    rows = []
    for m in metric_labels:
        rows.append({
            "Metric": m,
            f"{STRATEGY_NAME} (1 Year/TTM)": oy_met[m],
            f"{STRATEGY_NAME} (Inception)": inc_met[m],
        })
    return rows


def _rebuild_nav_figure(nav_series):
    """Rebuild the NAV Plotly figure from a live NAV series."""
    if nav_series.empty:
        fig = go.Figure()
        fig.add_annotation(text="No data", xref="paper", yref="paper", x=0.5, y=0.5, showarrow=False)
        return fig
    fig = go.Figure(
        go.Scatter(x=nav_series.index, y=nav_series.values,
                   mode="lines", line={"color": PRIMARY_COLOR}, name="NAV")
    )
    cfg = {
        "title": {"text": "<u>Non-Compounded NAV Since Inception</u>", "x": 0.5, "xanchor": "center"},
        "template": "ggplot2",
        "plot_bgcolor": GREY_BG,
        "paper_bgcolor": WHITE_BG,
        "xaxis_title": "Date",
        "yaxis_title": "NAV",
        "autosize": True,
    }
    if SHOW_PERCENTAGE_AXIS:
        nav_min, nav_max = nav_series.min(), nav_series.max()
        pct_min = ((nav_min - BASELINE_AMOUNT) / BASELINE_AMOUNT) * 100
        pct_max = ((nav_max - BASELINE_AMOUNT) / BASELINE_AMOUNT) * 100
        pct_range = pct_max - pct_min
        step = 10 if pct_range > 50 else (5 if pct_range > 20 else (2 if pct_range > 10 else 1))
        ps = (int(pct_min / step) - 1) * step
        pe = (int(pct_max / step) + 2) * step
        ticks = list(range(int(ps), int(pe) + step, step))
        cfg["yaxis2"] = {
            "title": "Return (%)", "overlaying": "y", "side": "right",
            "tickmode": "array",
            "tickvals": [BASELINE_AMOUNT * (1 + p / 100) for p in ticks],
            "ticktext": [f"{p:.0f}%" for p in ticks],
            "showgrid": False, "zeroline": False,
        }
        cfg["margin"] = {"l": 40, "r": 70, "t": 40, "b": 40}
    else:
        cfg["margin"] = {"l": 40, "r": 10, "t": 40, "b": 40}
    fig.update_layout(**cfg)
    fig.update_xaxes(showgrid=True)
    fig.update_yaxes(showgrid=True)
    return fig


# ── Export to Excel ───────────────────────────────────────────────────────
@app.callback(
    Output("secret-export-download", "data"),
    Input("secret-export-btn", "n_clicks"),
    State("secret-data-store", "data"),
    prevent_initial_call=True,
)
def export_excel(n_clicks, store_data):
    if not n_clicks or not store_data:
        return dash.no_update
    export_df = pd.DataFrame(store_data)
    drop_cols = [c for c in ("_row_id", "Actions") if c in export_df.columns]
    if drop_cols:
        export_df.drop(columns=drop_cols, inplace=True)
    return dcc.send_data_frame(export_df.to_excel, "daily_returns.xlsx", index=False)


# ══════════════════════════════════════════════════════════════════════════
# MASTER PROPAGATION — rebuild all downstream dashboard outputs from store
# ══════════════════════════════════════════════════════════════════════════

def _build_perf_card(perf_records):
    """Build a Performance Metrics dbc.Card from records."""
    df = pd.DataFrame(perf_records)
    return dbc.Card(
        [
            dbc.CardHeader(html.H6("Performance Metrics", className="mb-0")),
            dbc.CardBody(
                dbc.Table.from_dataframe(df, striped=False, bordered=True, hover=True, size="sm", className="fixed-cols")
            ),
        ],
        outline=True,
        className="mb-4",
    )

def _build_monthly_table(monthly_records):
    """Build a monthly calendar dbc.Table from records — one row per year, official results only."""
    df = pd.DataFrame(monthly_records)
    if df.empty:
        return html.P("No data available", className="text-muted")

    header = html.Thead(
        html.Tr([html.Th(col, style={"backgroundColor": GREY_BG, "color": "#000"}) for col in df.columns])
    )

    body_rows = []
    for i in range(len(df)):
        cells = [html.Td(df.iloc[i][col]) for col in df.columns]
        body_rows.append(html.Tr(cells))

    return dbc.Table(
        [header, html.Tbody(body_rows)],
        bordered=True, hover=True, size="sm",
        className="table-responsive mb-5",
        style={"width": "95%", "margin": "0 auto", "pageBreakInside": "avoid"},
    )

@app.callback(
    Output("monthly-calendar-container", "children"),
    Output("daily-perf-container", "children"),
    Output("NAV-graph", "figure"),
    Output("data-current-label-desktop", "children"),
    Output("data-current-label-mobile", "children"),
    Input("canonical-nav-store", "data"),
    State("secret-data-store", "data"),
)
def propagate_dashboard(canonical_nav_rows, secret_store_rows):
    if not canonical_nav_rows:
        return (dash.no_update,) * 5

    nav_s = _rebuild_nav_series(canonical_nav_rows)
    if nav_s.empty:
        return (dash.no_update,) * 5

    bl = BASELINE_AMOUNT

    monthly_recs = _recompute_monthly_records(nav_s, bl)
    perf_recs = _recompute_daily_perf_records(nav_s, bl)
    nav_fig = _rebuild_nav_figure(nav_s)

    # Use the most recent date from the Daily Returns store (the last actual entry),
    # not the canonical NAV series which is forward-filled through asfreq and may overshoot.
    latest = "unavailable"
    if secret_store_rows:
        dates = [r.get("Date", "") for r in secret_store_rows if r.get("Date")]
        if dates:
            try:
                latest = max(pd.to_datetime(dates)).strftime("%B %d, %Y")
            except Exception:
                pass
    if latest == "unavailable" and len(nav_s) > 0:
        latest = nav_s.index.max().strftime("%B %d, %Y")
    desktop_label_children, mobile_label_children = _build_tkp_date_status_label_children(latest)

    return (
        _build_monthly_table(monthly_recs),
        _build_perf_card(perf_recs),
        nav_fig,
        desktop_label_children,
        mobile_label_children,
    )


if __name__ == "__main__":
    app.run(debug=True, port=8301)