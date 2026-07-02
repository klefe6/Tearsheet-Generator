"""
TCP NAV worksheet read-only ledger adapter.

No Dash, Flask, server, or persistence side effects on import.
Does not save or modify workbooks.
"""
from __future__ import annotations

import os
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

# TCP display-column names preserved for future calculator compatibility.
REQUIRED_HEADERS: Tuple[str, ...] = (
    "Cash Transfers",
    "Trading Days",
    "Date",
    "Cash Balance",
    "NLV",
    "#",
    "$PL",
    "Inc. Fee",
    "cumm fee",
    "Day PnL",
    "nav-x1",
    "Loss Carry",
    "%Net",
    "S net cummulative %",
    "HWM",
)

# Fields stored as decimal ratios (e.g. 0.0001 == 0.01% displayed in Excel).
PERCENTAGE_HEADERS: frozenset[str] = frozenset({"%Net", "S net cummulative %"})

# Currency / P&L fields.
CURRENCY_HEADERS: frozenset[str] = frozenset(
    {
        "Cash Transfers",
        "Cash Balance",
        "NLV",
        "$PL",
        "Inc. Fee",
        "cumm fee",
        "Day PnL",
        "nav-x1",
        "Loss Carry",
        "HWM",
    }
)

INTEGER_HEADERS: frozenset[str] = frozenset({"Trading Days", "#"})

EXCEL_ERROR_PREFIX = "#"


class TCPLedgerError(Exception):
    """Base adapter error."""


class WorkbookNotFound(TCPLedgerError):
    """Configured workbook path does not exist."""


class WorksheetMissing(TCPLedgerError):
    """Worksheet name not present in workbook."""


class RequiredColumnMissing(TCPLedgerError):
    """A required TCP header is absent from the worksheet."""


class DuplicateRequiredHeader(TCPLedgerError):
    """A required TCP header appears more than once."""


class InvalidDate(TCPLedgerError):
    """Date cell could not be parsed."""


class InvalidNumericValue(TCPLedgerError):
    """Numeric cell contains an Excel error or non-numeric garbage."""


class FormulaCacheMissing(TCPLedgerError):
    """Formula cell has no cached calculated value."""


class LedgerEmpty(TCPLedgerError):
    """No completed ledger rows (nav-x1) were found."""


@dataclass(frozen=True)
class LedgerRecord:
    """One worksheet row with TCP display-column names preserved."""

    excel_row_number: int
    fields: Dict[str, Any]
    formulas: Dict[str, Optional[str]] = field(default_factory=dict)

    def public_dict(self) -> Dict[str, Any]:
        """Row data without formula strings (safe for UI)."""
        return dict(self.fields)


@dataclass(frozen=True)
class LedgerMetadata:
    source_filename: str
    sheet_name: str
    header_mapping: Dict[str, str]  # Excel column letter -> header label
    total_candidate_rows: int
    completed_row_count: int
    first_completed_date: Optional[date]
    latest_completed_date: Optional[date]
    latest_completed_excel_row: Optional[int]

    def public_dict(self) -> Dict[str, Any]:
        """Metadata safe for HTTP responses (no absolute paths)."""
        return {
            "source_filename": self.source_filename,
            "sheet_name": self.sheet_name,
            "header_mapping": dict(self.header_mapping),
            "total_candidate_rows": self.total_candidate_rows,
            "completed_row_count": self.completed_row_count,
            "first_completed_date": (
                self.first_completed_date.isoformat() if self.first_completed_date else None
            ),
            "latest_completed_date": (
                self.latest_completed_date.isoformat()
                if self.latest_completed_date
                else None
            ),
            "latest_completed_excel_row": self.latest_completed_excel_row,
        }


@dataclass(frozen=True)
class LedgerLoadResult:
    candidate_records: Tuple[LedgerRecord, ...]
    completed_records: Tuple[LedgerRecord, ...]
    metadata: LedgerMetadata

    @property
    def records(self) -> Tuple[LedgerRecord, ...]:
        """Alias for completed ledger rows (primary consumer API)."""
        return self.completed_records


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        stripped = value.strip()
        return stripped == "" or stripped.lower() == "nan"
    return False


def _is_excel_error(value: Any) -> bool:
    return isinstance(value, str) and value.startswith(EXCEL_ERROR_PREFIX)


def _normalize_date(value: Any, *, excel_row: int) -> Optional[date]:
    if _is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # openpyxl may return serial numbers in rare read paths
        try:
            from openpyxl.utils.datetime import from_excel

            return from_excel(value).date()
        except Exception as exc:
            raise InvalidDate(f"Row {excel_row}: cannot parse Excel date serial {value!r}") from exc
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%y", "%m-%d-%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    raise InvalidDate(f"Row {excel_row}: invalid Date value {value!r}")


def _normalize_numeric(
    value: Any,
    *,
    excel_row: int,
    header: str,
) -> Optional[float]:
    if _is_blank(value):
        return None
    if _is_excel_error(value):
        raise InvalidNumericValue(f"Row {excel_row} {header}: Excel error {value!r}")
    if isinstance(value, bool):
        raise InvalidNumericValue(f"Row {excel_row} {header}: unexpected boolean")
    if isinstance(value, (int, float)):
        if header in INTEGER_HEADERS:
            if float(value).is_integer():
                return float(int(value))
        return float(value)
    if isinstance(value, str):
        stripped = value.strip().replace(",", "").replace("$", "")
        if stripped == "":
            return None
        try:
            num = float(stripped)
        except ValueError as exc:
            raise InvalidNumericValue(
                f"Row {excel_row} {header}: cannot parse {value!r}"
            ) from exc
        if header in INTEGER_HEADERS and num.is_integer():
            return float(int(num))
        return num
    raise InvalidNumericValue(f"Row {excel_row} {header}: unsupported type {type(value)}")


def _normalize_field(header: str, raw_value: Any, *, excel_row: int) -> Any:
    if header == "Date":
        return _normalize_date(raw_value, excel_row=excel_row)
    if header in PERCENTAGE_HEADERS:
        return _normalize_numeric(raw_value, excel_row=excel_row, header=header)
    if header in CURRENCY_HEADERS or header in INTEGER_HEADERS:
        return _normalize_numeric(raw_value, excel_row=excel_row, header=header)
    return None if _is_blank(raw_value) else raw_value


def _nav_x1_is_completed(value: Any) -> bool:
    if _is_blank(value):
        return False
    if _is_excel_error(value):
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        try:
            float(value.strip().replace(",", ""))
            return True
        except ValueError:
            return False
    return False


def discover_header_row(
    ws_formula: openpyxl.worksheet.worksheet.Worksheet,
    *,
    header_row: int = 1,
    max_columns: int = 30,
) -> Dict[str, int]:
    """
    Map required header label -> 1-based column index.
    Raises RequiredColumnMissing or DuplicateRequiredHeader.
    """
    label_to_cols: Dict[str, List[int]] = {}
    header_mapping: Dict[str, str] = {}

    for col_idx in range(1, max_columns + 1):
        letter = get_column_letter(col_idx)
        raw = ws_formula.cell(row=header_row, column=col_idx).value
        if raw is None:
            continue
        label = str(raw).strip()
        if not label:
            continue
        header_mapping[letter] = label
        if label in REQUIRED_HEADERS:
            label_to_cols.setdefault(label, []).append(col_idx)

    missing = [h for h in REQUIRED_HEADERS if h not in label_to_cols]
    if missing:
        raise RequiredColumnMissing(
            f"Required TCP headers missing from row {header_row}: {', '.join(missing)}"
        )

    duplicates = {h: cols for h, cols in label_to_cols.items() if len(cols) > 1}
    if duplicates:
        detail = ", ".join(f"{h}@{cols}" for h, cols in duplicates.items())
        raise DuplicateRequiredHeader(f"Duplicate required headers: {detail}")

    return {header: label_to_cols[header][0] for header in REQUIRED_HEADERS}


def build_header_mapping(
    ws_formula: openpyxl.worksheet.worksheet.Worksheet,
    *,
    header_row: int = 1,
    max_columns: int = 30,
) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    for col_idx in range(1, max_columns + 1):
        letter = get_column_letter(col_idx)
        raw = ws_formula.cell(row=header_row, column=col_idx).value
        if raw is not None:
            mapping[letter] = str(raw).strip()
    return mapping


def _read_row(
    ws_values: openpyxl.worksheet.worksheet.Worksheet,
    ws_formula: openpyxl.worksheet.worksheet.Worksheet,
    excel_row: int,
    header_cols: Mapping[str, int],
    *,
    strict_formula_cache: bool = False,
) -> LedgerRecord:
    fields: Dict[str, Any] = {}
    formulas: Dict[str, Optional[str]] = {}

    for header, col_idx in header_cols.items():
        formula_cell = ws_formula.cell(row=excel_row, column=col_idx)
        value_cell = ws_values.cell(row=excel_row, column=col_idx)
        raw_formula = formula_cell.value
        raw_value = value_cell.value

        is_formula = isinstance(raw_formula, str) and raw_formula.startswith("=")
        formulas[header] = raw_formula if is_formula else None

        if is_formula and _is_blank(raw_value):
            if strict_formula_cache and header != "Date":
                raise FormulaCacheMissing(
                    f"Row {excel_row} {header}: formula present without cached value"
                )
            fields[header] = None
            continue

        if header == "Date":
            if _is_blank(raw_value):
                fields[header] = None
            else:
                fields[header] = _normalize_date(raw_value, excel_row=excel_row)
        else:
            fields[header] = _normalize_field(header, raw_value, excel_row=excel_row)

    return LedgerRecord(excel_row_number=excel_row, fields=fields, formulas=formulas)


def _row_is_candidate(record: LedgerRecord) -> bool:
    """Candidate rows have a valid Date (includes date-only seed rows)."""
    return record.fields.get("Date") is not None


def _row_is_completed(record: LedgerRecord) -> bool:
    nav = record.fields.get("nav-x1")
    return nav is not None and isinstance(nav, (int, float))


def _record_has_ledger_content(
    ws_values: openpyxl.worksheet.worksheet.Worksheet,
    excel_row: int,
    header_cols: Mapping[str, int],
) -> bool:
    for col_idx in header_cols.values():
        if not _is_blank(ws_values.cell(row=excel_row, column=col_idx).value):
            return True
    return False


def load_ledger(
    workbook_path: str,
    sheet_name: str = "NAV",
    *,
    strict_formula_cache: bool = False,
) -> LedgerLoadResult:
    """
    Load and normalize the TCP NAV ledger in read-only mode.

    Percentages are returned as decimal ratios (Excel internal representation).
    Dates are returned as datetime.date objects.
    Blank cells become None. Currency values are float (not rounded).
    """
    path = workbook_path
    if not os.path.isfile(path):
        raise WorkbookNotFound(f"Workbook not found: {path}")
    if not os.access(path, os.R_OK):
        raise WorkbookNotFound(f"Workbook not readable: {path}")

    source_filename = os.path.basename(path)

    wb_formula = None
    wb_values = None
    try:
        wb_formula = openpyxl.load_workbook(path, data_only=False, read_only=True)
        if sheet_name not in wb_formula.sheetnames:
            raise WorksheetMissing(f"Worksheet {sheet_name!r} not found in {source_filename}")

        wb_values = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws_formula = wb_formula[sheet_name]
        ws_values = wb_values[sheet_name]

        header_cols = discover_header_row(ws_formula)
        header_mapping = build_header_mapping(ws_formula)

        max_row = ws_values.max_row or 1
        candidate_records: List[LedgerRecord] = []
        last_content_row = 1

        for excel_row in range(2, max_row + 1):
            if not _record_has_ledger_content(ws_values, excel_row, header_cols):
                continue
            last_content_row = excel_row
            record = _read_row(
                ws_values,
                ws_formula,
                excel_row,
                header_cols,
                strict_formula_cache=strict_formula_cache,
            )
            if _row_is_candidate(record):
                candidate_records.append(record)

        # Trim any candidate rows beyond the last non-empty worksheet row.
        candidate_records = [
            r for r in candidate_records if r.excel_row_number <= last_content_row
        ]

        completed_records = [r for r in candidate_records if _row_is_completed(r)]
        if not completed_records:
            raise LedgerEmpty("No completed ledger rows with nav-x1 found")

        first_completed = completed_records[0]
        latest_completed = completed_records[-1]

        metadata = LedgerMetadata(
            source_filename=source_filename,
            sheet_name=sheet_name,
            header_mapping=header_mapping,
            total_candidate_rows=len(candidate_records),
            completed_row_count=len(completed_records),
            first_completed_date=first_completed.fields["Date"],
            latest_completed_date=latest_completed.fields["Date"],
            latest_completed_excel_row=latest_completed.excel_row_number,
        )

        return LedgerLoadResult(
            candidate_records=tuple(candidate_records),
            completed_records=tuple(completed_records),
            metadata=metadata,
        )
    except (PermissionError, OSError) as exc:
        raise WorkbookNotFound(f"Cannot read workbook: {path}") from exc
    finally:
        if wb_formula is not None:
            wb_formula.close()
        if wb_values is not None:
            wb_values.close()


def get_record_by_excel_row(
    result: LedgerLoadResult, excel_row_number: int
) -> Optional[LedgerRecord]:
    for record in result.candidate_records:
        if record.excel_row_number == excel_row_number:
            return record
    return None
