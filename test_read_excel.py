import openpyxl
import pandas as pd
import os

# File path
xlsx_path = r"C:\Users\H&CDanHughes\Hughes & Company\Hughes & Company - Documents\3_Advisors Marketing (Tearsheets, PitchBooks, etc)\1. Tearsheet Project\TKP\VADI\tkp_alex_old.xlsx"

print("=" * 80)
print("TEST: Reading Excel file to check column G")
print("=" * 80)
print(f"File: {xlsx_path}")
print(f"File exists: {os.path.exists(xlsx_path)}")
print()

# Method 1: Using openpyxl to check column G directly
print("Method 1: Using openpyxl to read column G directly")
print("-" * 80)
try:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Sheet1"]
    
    # Find last row with data in column G
    last_row_g = 1
    g_values = []
    for row_idx in range(2, ws.max_row + 1):
        cell_g = ws.cell(row=row_idx, column=7)  # Column G = 7
        if cell_g.value is not None:
            last_row_g = row_idx
            g_values.append((row_idx, cell_g.value))
    
    wb.close()
    
    print(f"Max row in sheet (ws.max_row): {ws.max_row}")
    print(f"Last row with data in column G: {last_row_g}")
    print(f"Total non-empty values in column G: {len(g_values)}")
    print()
    print("Last 10 values in column G (row number, value):")
    for row_num, val in g_values[-10:]:
        print(f"  Row {row_num}: {val}")
    print()
    print("Last 3 values in column G:")
    for row_num, val in g_values[-3:]:
        print(f"  Row {row_num}: {val}")
    print()
    
except Exception as e:
    print(f"ERROR with openpyxl: {e}")
    if "Permission denied" in str(e):
        print("  -> File is likely open in Excel. Please close it and try again.")
    import traceback
    traceback.print_exc()

print()
print("=" * 80)
print("Method 2: Using pandas to read column G")
print("-" * 80)

# Method 2: Using pandas
try:
    # First check with openpyxl what the last row should be
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb["Sheet1"]
    max_row_check = ws.max_row
    wb.close()
    
    print(f"Excel max_row: {max_row_check}")
    
    # Try reading with different nrows values
    for nrows in [None, max_row_check - 1, 715 - 1]:
        try:
            if nrows is None:
                print(f"\nReading with nrows=None (default - stops at empty rows):")
                df = pd.read_excel(
                    xlsx_path,
                    sheet_name="Sheet1",
                    usecols="G",
                    header=0,
                    engine="openpyxl"
                )
            else:
                print(f"\nReading with nrows={nrows} (explicit row count):")
                df = pd.read_excel(
                    xlsx_path,
                    sheet_name="Sheet1",
                    usecols="G",
                    header=0,
                    engine="openpyxl",
                    nrows=nrows
                )
            
            # Get column name (pandas might rename it)
            col_name = df.columns[0]
            print(f"  Total rows read: {len(df)}")
            print(f"  Column name: {col_name}")
            
            # Drop NaN values to see actual data
            df_clean = df.dropna()
            print(f"  Non-empty values: {len(df_clean)}")
            
            if len(df_clean) > 0:
                print(f"  Last 10 values:")
                for idx, val in df_clean.iloc[-10:].iterrows():
                    print(f"    Index {idx}: {val[col_name]}")
                print(f"  Last 3 values:")
                for idx, val in df_clean.iloc[-3:].iterrows():
                    print(f"    Index {idx}: {val[col_name]}")
            else:
                print("  WARNING: No non-empty values found!")
                
        except Exception as e:
            print(f"  ERROR reading with nrows={nrows}: {e}")
    
except Exception as e:
    print(f"ERROR with pandas: {e}")
    import traceback
    traceback.print_exc()

print()
print("=" * 80)
print("Method 3: Reading columns C and N (Date and NAV) to compare")
print("-" * 80)

try:
    # Read columns C and N like the main script does
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb["Sheet1"]
    
    last_row = 1
    for row_idx in range(2, ws.max_row + 1):
        cell_c = ws.cell(row=row_idx, column=3)  # Column C
        cell_n = ws.cell(row=row_idx, column=14)  # Column N
        if cell_c.value is not None or cell_n.value is not None:
            last_row = row_idx
    
    wb.close()
    
    print(f"Last row with data in C or N: {last_row}")
    
    read_params = {
        "sheet_name": "Sheet1",
        "usecols": "C,N",
        "header": 0,
        "parse_dates": ["Date"],
        "engine": "openpyxl",
    }
    if last_row > 1:
        read_params["nrows"] = last_row - 1
    
    df_cn = pd.read_excel(xlsx_path, **read_params)
    df_cn = df_cn.dropna(how='all')
    
    print(f"Total rows read: {len(df_cn)}")
    if len(df_cn) > 0:
        print(f"Date range: {df_cn['Date'].min()} to {df_cn['Date'].max()}")
        print(f"Last 3 dates:")
        for idx, row in df_cn.tail(3).iterrows():
            print(f"  {row['Date']}: NAV = {row.get('nav-x1', 'N/A')}")
    
except Exception as e:
    print(f"ERROR: {e}")
    import traceback
    traceback.print_exc()

print()
print("=" * 80)
print("Test complete!")
print("=" * 80)

