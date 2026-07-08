"""
mp_ts.py  –  Momentum Pacer Tearsheet
======================================
Strategy : Momentum Pacer
CTA       : Algominds Financial LLC

Visual frame matches the Y&Q tearsheet (yq_ts.py).
Data source: Momentum Fee Calculation.xlsx  (Summary + Sris Fee Calc Detail tabs)

HOW TO RUN
----------
    python mp_ts.py
Then open: http://127.0.0.1:8304

Production (Cloudflare / reboot_mp_ts.bat): set MP_TS_PRODUCTION=1 so the server runs
without Dash debug/reloader (avoids unstable behavior behind a reverse proxy).
"""

from __future__ import annotations

import base64
import json
import math
import os
import sys
import traceback
from datetime import datetime, timedelta
from pathlib import Path

_TS_ROOT = Path(__file__).resolve().parent.parent
if str(_TS_ROOT) not in sys.path:
    sys.path.insert(0, str(_TS_ROOT))
import tearsheet_disclosure as tsd
from tearsheet_gate_ui import build_manager_accept_gate
from tearsheet_gate_auth import (
    build_gate_password_row,
    gate_password_row_style,
    load_agm_admin_auth_settings,
    ADMIN_PORTAL_PATH,
    AGM_SESSION_KEY,
    GATE_PASSWORD_ERROR_ID,
    GATE_PASSWORD_INPUT_ID,
    GATE_PASSWORD_PORTAL_ID,
    GATE_PASSWORD_ROW_ID,
    GATE_PASSWORD_SUBMIT_ID,
    GATE_PASSWORD_VISIBLE_STORE_ID,
    INVALID_PASSWORD_MESSAGE,
)
from tcp_admin import AdminAuthManager, configure_flask_session_secret
from tearsheet_portal import render_portal_page
from tearsheet_date_defaults import default_add_row_date_str
from tearsheet_header import (
    build_header_date_label_children_from_date,
    build_tearsheet_header_row,
)
import algominds_portal_registry as agm_registry
import algominds_daily_balances as agm_daily
import algominds_benchmark_daily as agm_bench
import algominds_daily_fees as agm_fees
import algominds_daily_accounting as agm_accounting
import algominds_monthly_summary as agm_monthly
import algominds_fee_payment_evidence as agm_fee_evidence

import numpy as np
import pandas as pd
import plotly.graph_objs as go
import openpyxl

import dash
from dash import html, dcc, dash_table
from dash.dash_table.Format import Format, Scheme, Symbol, Sign
import dash_bootstrap_components as dbc
from dash.dependencies import Input, Output, State
from flask import jsonify, redirect, session

# ==============================================================================
# PATHS
# ==============================================================================
BASE_DIR  = Path(__file__).parent
EXCEL_PATH = BASE_DIR / "Momentum Fee Calculation.xlsx"

# ==============================================================================
# BRAND / STYLE  (mirrors Y&Q tearsheet conventions)
# ==============================================================================
WHITE_BG         = "#ffffff"
GREY_BG          = "#EBEBEB"
PRIMARY_COLOR    = "#1B4F8A"      # Algominds blue
SECONDARY_COLOR  = "#CCCCCC"
ACCENT_GREEN     = "#28a745"
ACCENT_RED       = "#dc3545"
LEFT_TABLE_GAPS  = "20px"
RIGHT_TABLE_GAPS = "30px"
HEADER_ROW_CLASS = "bg-light"

# ==============================================================================
# LOGO
# ==============================================================================
logo_src = ""
for lp in [BASE_DIR / "algominds_logo.png", BASE_DIR / "logo.png"]:
    if lp.exists():
        with open(lp, "rb") as _f:
            logo_src = f"data:image/png;base64,{base64.b64encode(_f.read()).decode()}"
        break

# ==============================================================================
# DATA LOADING  –  reads directly from Excel
# ==============================================================================
LOAD_ERROR = None

def load_summary(path: Path) -> pd.DataFrame:
    """
    Parse the 'Summary' sheet.
    Columns: Month, SPX Start, SPX End, NDX Start, NDX End,
             BOT Start, BOT End After Fees,
             SPX Returns%, NDX Returns%,
             BOT Returns Before Fees%, BOT Fees%, BOT Returns After Fees%,
             Cumulative After Fees%
    Only rows where Month is a real date AND BOT Start is not None are kept.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Summary"]

    rows = list(ws.iter_rows(values_only=True))
    # Row 0 = headers, rows 1+ = data
    data = []
    for row in rows[1:]:
        # stop at Net% / Net$ footer rows or empty month
        if row[0] is None or not isinstance(row[0], datetime):
            continue
        # skip rows where BOT Start is empty (future months)
        if row[5] is None:
            continue
        data.append({
            "date":              row[0],
            "spx_start":         row[1],
            "spx_end":           row[2],
            "ndx_start":         row[3],
            "ndx_end":           row[4],
            "bot_start":         row[5],
            "bot_end_after_fees":row[6],
            "spx_ret":           row[7],   # decimal  e.g. 0.0166
            "ndx_ret":           row[8],
            "bot_gross_ret":     row[9],   # BOT returns before fees
            "bot_fees_pct":      row[10],
            "bot_net_ret":       row[11],  # BOT returns after fees
            "cumulative_net":    row[12],
        })

    df = pd.DataFrame(data)
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values("date").reset_index(drop=True)
    return df


def load_net_totals(path: Path) -> dict:
    """Read the Net% and Net$ footer rows from Summary sheet."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Summary"]
    totals = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] == "Net%":
            totals["spx_net_pct"]          = row[7]
            totals["ndx_net_pct"]          = row[8]
            totals["bot_gross_pct"]        = row[9]
            totals["bot_fees_pct"]         = row[10]
            totals["bot_net_pct"]          = row[11]
        if row[0] == "Net$":
            totals["spx_net_dollar"]       = row[7]
            totals["ndx_net_dollar"]       = row[8]
            totals["bot_gross_dollar"]     = row[9]
            totals["bot_fees_dollar"]      = row[10]
            totals["bot_net_dollar"]       = row[11]
    return totals


# Live trading began (TradeStation); monthly summary rows are month-stubs (e.g. Nov 1).
PROGRAM_INCEPTION = datetime(2025, 11, 13, 0, 0, 0)

# Charts and table filter use this date instead of machine clock (first of its month caps rows).
TEARSHEET_AS_OF = datetime(2026, 5, 12, 0, 0, 0)

# Incentive fee slabs — AlgoMinds Financial LLC Disclosure Document (e.g. effective March 1, 2026),
# "Advisor's Fees" / Incentive Fee. Benchmark = S&P 500 monthly return (WSJ official month-end closes);
# tiers are slices of net new profits measured vs multiples of the Benchmark's *dollar* return for the month.
FEE_SLAB_BANDS_DISPLAY = (
    (
        "Slab 1",
        "Net new profits from $0 up through 100% of the Benchmark's monthly dollar return (0–1×)",
        "10%",
    ),
    (
        "Slab 2",
        "Portion exceeding 100% but not more than 200% of that Benchmark dollar return (1×–2×)",
        "20%",
    ),
    (
        "Slab 3",
        "Portion exceeding 200% but not more than 300% (2×–3×)",
        "30%",
    ),
    (
        "Slab 4",
        "Portion exceeding 300% but not more than 400% (3×–4×)",
        "40%",
    ),
    (
        "Slab 5",
        "Portion exceeding 400% of the Benchmark's monthly dollar return (>4×)",
        "50%",
    ),
)

try:
    summary_df   = load_summary(EXCEL_PATH)
    net_totals   = load_net_totals(EXCEL_PATH)
    STARTING_CAPITAL = float(summary_df["bot_start"].iloc[0])
    LATEST_DATE      = summary_df["date"].max()
    print(f"[mp_ts] Loaded {len(summary_df)} months from {EXCEL_PATH.name}")
    print(
        f"[mp_ts] Program inception: {PROGRAM_INCEPTION.strftime('%m/%d/%Y')}  "
        f"Latest month: {LATEST_DATE.strftime('%b %Y')}"
    )
except Exception as _e:
    traceback.print_exc()
    summary_df   = pd.DataFrame()
    net_totals   = {}
    STARTING_CAPITAL = 30_000.0
    LATEST_DATE      = PROGRAM_INCEPTION
    LOAD_ERROR = str(_e)


# ── Daily TradeStation balances (ADMIN-ONLY raw NLV; never client-facing) ──────
# Loaded once at import, mirroring the monthly summary_df pattern. Failure here is
# non-fatal: the daily admin table/graph fall back to an honest empty state.
DAILY_BALANCES_LOAD_ERROR = None
try:
    daily_balances_df = agm_daily.load_daily_balances()
except Exception as _de:  # pragma: no cover - defensive
    traceback.print_exc()
    daily_balances_df = pd.DataFrame()
    DAILY_BALANCES_LOAD_ERROR = str(_de)

# Logging is kept out of the load try/except above so a console-encoding hiccup
# can never blank an otherwise-good dataframe. ASCII only (Windows cp1252 safe).
if DAILY_BALANCES_LOAD_ERROR is None and not daily_balances_df.empty:
    print(
        f"[mp_ts] Loaded {len(daily_balances_df)} daily balance rows "
        f"({daily_balances_df['Date'].min().date()} to {daily_balances_df['Date'].max().date()})"
    )
elif DAILY_BALANCES_LOAD_ERROR is None:
    print("[mp_ts] Daily balances CSV not found or empty - admin daily view will show empty state.")


# ── Daily benchmarks (SPX ^GSPC / NDX ^NDX) + daily fee accrual ───────────────
# Cache-first (Momentum Pacer/data/benchmarks, committed); yfinance is touched
# only when the cache does not cover the CSV date range. See
# algominds_benchmark_daily for the ^GSPC-vs-^SP500TR rationale (the fee
# workbook's SPX levels are ^GSPC price-index closes, verified to the cent).
BENCHMARK_LOAD_ERROR = None
spx_daily_df = pd.DataFrame(columns=["Date", "Close"])
ndx_daily_df = pd.DataFrame(columns=["Date", "Close"])
try:
    if not daily_balances_df.empty:
        _bench_start = daily_balances_df["Date"].min() - pd.Timedelta(days=45)
        _bench_end = daily_balances_df["Date"].max()
        spx_daily_df = agm_bench.load_daily_benchmark(agm_bench.SPX_TICKER, _bench_start, _bench_end)
        ndx_daily_df = agm_bench.load_daily_benchmark(agm_bench.NDX_TICKER, _bench_start, _bench_end)
except Exception as _bex:  # pragma: no cover - defensive
    traceback.print_exc()
    BENCHMARK_LOAD_ERROR = str(_bex)

# Daily incentive-fee accrual (AGM daily NLV vs daily SPX; workbook Summary is
# passed only as internal payment-reconciliation reference, never displayed).
DAILY_FEES_LOAD_ERROR = None
try:
    daily_fee_accrual = agm_fees.compute_daily_fee_accrual(
        daily_balances_df,
        spx_daily_df,
        inception=pd.Timestamp(PROGRAM_INCEPTION),
        monthly_reference=summary_df if not summary_df.empty else None,
    )
except Exception as _fex:  # pragma: no cover - defensive
    traceback.print_exc()
    daily_fee_accrual = agm_fees.DailyFeeAccrual(daily=pd.DataFrame())
    DAILY_FEES_LOAD_ERROR = str(_fex)

if BENCHMARK_LOAD_ERROR is None and not spx_daily_df.empty:
    print(
        f"[mp_ts] Loaded {len(spx_daily_df)} daily SPX ({agm_bench.SPX_TICKER}) closes, "
        f"{len(ndx_daily_df)} NDX ({agm_bench.NDX_TICKER}) closes"
    )
elif BENCHMARK_LOAD_ERROR is None:
    print("[mp_ts] Daily benchmark data unavailable - charts will omit benchmark series.")
if DAILY_FEES_LOAD_ERROR is None and not daily_fee_accrual.daily.empty:
    print(
        f"[mp_ts] Daily fee accrual: {len(daily_fee_accrual.daily)} days, "
        f"{len(daily_fee_accrual.crystallized)} crystallized months, "
        f"{len(daily_fee_accrual.payments)} evidenced payments, "
        f"{len(daily_fee_accrual.outstanding)} outstanding"
    )

# ── Daily accounting table (client net value, accrued fees, SPX alignment) ─────
DAILY_ACCOUNTING_LOAD_ERROR = None
daily_accounting = agm_accounting.AgmDailyAccounting(table=pd.DataFrame())
try:
    if not daily_balances_df.empty:
        daily_accounting = agm_accounting.compute_agm_daily_accounting(
            daily_balances_df,
            spx_daily_df,
            fee_accrual=daily_fee_accrual if DAILY_FEES_LOAD_ERROR is None else None,
            inception=pd.Timestamp(PROGRAM_INCEPTION),
            monthly_reference=summary_df if not summary_df.empty else None,
        )
except Exception as _aex:  # pragma: no cover - defensive
    traceback.print_exc()
    daily_accounting = agm_accounting.AgmDailyAccounting(table=pd.DataFrame())
    DAILY_ACCOUNTING_LOAD_ERROR = str(_aex)

if DAILY_ACCOUNTING_LOAD_ERROR is None and not daily_accounting.table.empty:
    print(
        f"[mp_ts] Daily accounting: {len(daily_accounting.table)} rows, "
        f"invariant_ok={agm_accounting.verify_accounting_invariant(daily_accounting.table)}"
    )

# ── DERIVED monthly Performance Summary (supersedes the workbook for display) ──
# The workbook Summary sheet is hand-maintained and goes stale between updates
# (its last entered row froze mid-May 2026, so June never appeared). Monthly
# display rows are now derived from the accepted daily accounting model, the
# fee engine's month-end crystallizations, and the cached benchmark closes —
# only COMPLETE months are emitted, so the in-progress month (e.g. July with
# data through Jul 6) never shows as a monthly row. The workbook stays loaded
# strictly as the fee engine's internal payment-reconciliation reference.
MONTHLY_SUMMARY_LOAD_ERROR = None
monthly_summary = agm_monthly.AgmMonthlySummary(table=pd.DataFrame())
try:
    if not daily_accounting.table.empty:
        monthly_summary = agm_monthly.compute_agm_monthly_summary(
            daily_accounting.table,
            daily_fee_accrual.crystallized,
            spx_daily_df,
            ndx_daily_df,
            inception=pd.Timestamp(PROGRAM_INCEPTION),
        )
except Exception as _msex:  # pragma: no cover - defensive
    traceback.print_exc()
    monthly_summary = agm_monthly.AgmMonthlySummary(table=pd.DataFrame())
    MONTHLY_SUMMARY_LOAD_ERROR = str(_msex)

if MONTHLY_SUMMARY_LOAD_ERROR is None and not monthly_summary.table.empty:
    print(
        f"[mp_ts] Derived monthly summary: {len(monthly_summary.table)} complete months "
        f"through {monthly_summary.table['date'].max().strftime('%b %Y')}"
    )


def months_trading_elapsed_approx() -> str:
    """
    Elapsed time from live inception (Nov 13) to the *start* of the latest Summary
    month row (e.g. Jun 1), expressed as decimal months using 365.25/12 days/month.
    This is not the same as the count of monthly return rows displayed.
    """
    disp = _display_summary_df
    if disp.empty:
        return "—"
    end = pd.Timestamp(disp["date"].max()).to_pydatetime()
    delta = end - PROGRAM_INCEPTION
    if delta.days <= 0:
        return "0.0"
    days_per_month = 365.25 / 12.0
    return f"{delta.days / days_per_month:.1f}"


# ==============================================================================
# CALCULATED METRICS  (mirroring Y&Q's calculate_period_metrics_monthly)
# ==============================================================================
def calc_performance_metrics(df: pd.DataFrame) -> dict:
    """
    Compute key performance stats from the summary dataframe.
    Returns a dict of display-ready strings.

    Cumulative net return matches the spreadsheet / Net% row: it is taken from
    actual month-end BOT balances (last ÷ first − 1), not from chaining the
    monthly "% after fees" cells — those can differ slightly from a pure MoM
    NAV ratio (HWM / fee mechanics / rounding), so ∏(1+r) would not match 49.65%.
    """
    if df.empty:
        return {}

    net_rets = df["bot_net_ret"].astype(float)  # decimal per month (displayed %)
    n        = len(net_rets)

    # Dollar-accurate cumulative (aligns with Summary "Cumul. Net%" / Net% row)
    start_eq = float(df["bot_start"].iloc[0])
    end_eq   = float(df["bot_end_after_fees"].iloc[-1])
    cum      = (end_eq / start_eq) - 1.0 if start_eq > 0 else 0.0

    # CAGR using calendar span inception → start of latest summary month in *df*
    latest_row = pd.Timestamp(df["date"].max()).to_pydatetime()
    span_days = (latest_row - PROGRAM_INCEPTION).days
    years_real = span_days / 365.25 if span_days > 0 else 0.0
    ann = ((1 + cum) ** (1 / years_real) - 1) if years_real > 0 else 0.0

    avg      = net_rets.mean()
    std      = net_rets.std(ddof=1)
    sharpe   = (avg / std * np.sqrt(12)) if std > 0 else 0.0

    wins   = (net_rets > 0).sum()
    losses = (net_rets < 0).sum()
    top3   = net_rets.nlargest(3) * 100
    bot3   = net_rets.nsmallest(3) * 100

    return {
        "Cumulative Net Return":   f"{cum*100:.2f}%",
        "Annualized Net Return":   f"{ann*100:.2f}%",
        "Avg Monthly Net Return":  f"{avg*100:.3f}%",
        "Sharpe Ratio (approx)":   f"{sharpe:.2f}",
        "Number of Months":        str(n),
        "% Winning Months":        f"{wins} ({wins/n*100:.1f}%)",
        "% Losing Months":         f"{losses} ({losses/n*100:.1f}%)",
        "Best 3 Months":           ", ".join(f"{v:.2f}%" for v in top3),
        "Worst 3 Months":          ", ".join(f"{v:.2f}%" for v in bot3),
    }


def calc_monthly_stats(df: pd.DataFrame) -> dict:
    """Monthly performance statistics card (mirrors Y&Q's calculate_monthly_stats)."""
    if df.empty:
        return {}
    net_rets = df["bot_net_ret"].astype(float)
    pos  = net_rets[net_rets > 0]
    neg  = net_rets[net_rets < 0]
    n    = len(net_rets)

    # Streak calculation
    signs = (net_rets > 0).astype(int)
    win_streaks, loss_streaks = [], []
    cur, cur_sign = 1, signs.iloc[0]
    for i in range(1, len(signs)):
        if signs.iloc[i] == cur_sign:
            cur += 1
        else:
            (win_streaks if cur_sign == 1 else loss_streaks).append(cur)
            cur, cur_sign = 1, signs.iloc[i]
    (win_streaks if cur_sign == 1 else loss_streaks).append(cur)

    return {
        "Number of Positive Months": f"{len(pos)} ({len(pos)/n*100:.1f}%)",
        "Number of Negative Months": f"{len(neg)} ({len(neg)/n*100:.1f}%)",
        "Average Winning Month %":   f"{pos.mean()*100:.2f}%" if len(pos) else "—",
        "Average Losing Month %":    f"{neg.mean()*100:.2f}%" if len(neg) else "—",
        "Best Single Month %":       f"{net_rets.max()*100:.2f}%",
        "Worst Single Month %":      f"{net_rets.min()*100:.2f}%",
        "Longest Winning Streak":    f"{max(win_streaks) if win_streaks else 0} months",
        "Longest Losing Streak":     f"{max(loss_streaks) if loss_streaks else 0} months",
    }


# ==============================================================================
# FIGURE BUILDERS
# ==============================================================================


def _chart_today() -> pd.Timestamp:
    """Calendar date for chart cutoffs (tearsheet as-of; not machine clock)."""
    return pd.Timestamp(TEARSHEET_AS_OF.date())


def _first_of_calendar_month(ts: pd.Timestamp) -> pd.Timestamp:
    """Midnight on the 1st of the same calendar month as *ts*."""
    t = pd.Timestamp(ts).normalize()
    return t.replace(day=1)


def _summary_through_current_month(df: pd.DataFrame) -> pd.DataFrame:
    """
    Rows on or before the 1st of the current calendar month.
    Excludes next-month rows from Excel so mid-May never looks like June data.
    """
    if df.empty:
        return df
    cur = _first_of_calendar_month(_chart_today())
    return df[df["date"] <= cur].reset_index(drop=True)


# Displayed monthly summary: the DERIVED frame (daily accounting + fee engine
# + benchmark cache; complete months only, so June appears and in-progress July
# does not). The stale workbook slice remains only as a defensive fallback if
# the daily pipeline ever fails to load; net_totals follow the same source.
if not monthly_summary.table.empty:
    _display_summary_df = monthly_summary.table
    net_totals = dict(monthly_summary.totals)
else:  # pragma: no cover - defensive fallback to the workbook path
    _display_summary_df = _summary_through_current_month(summary_df)
if not _display_summary_df.empty:
    LATEST_DATE = _display_summary_df["date"].max()
perf_metrics  = calc_performance_metrics(_display_summary_df)
monthly_stats = calc_monthly_stats(_display_summary_df)

perf_df = pd.DataFrame({
    "Metric": list(perf_metrics.keys()),
    f"Momentum Pacer (Inception)": list(perf_metrics.values()),
}) if perf_metrics else pd.DataFrame()

stats_df = pd.DataFrame({
    "Metric": list(monthly_stats.keys()),
    "Momentum Pacer (Inception)": list(monthly_stats.values()),
}) if monthly_stats else pd.DataFrame()


def _hover_customdata(y_vals: list[float], baseline: float) -> list[list]:
    """
    Returns a list of [delta_vs_prior_str, cum_pct_vs_baseline_str] per point.
    First point shows '—' for both fields.
    """
    out: list[list] = []
    for i, y in enumerate(y_vals):
        yf = float(y)
        # Δ vs prior
        if i == 0:
            delta = "—"
        else:
            prev = float(y_vals[i - 1])
            if abs(prev) < 1e-12:
                delta = "—"
            else:
                delta = f"{(yf / prev - 1.0) * 100.0:+.2f}%"
        # Cumulative % vs baseline
        if abs(baseline) < 1e-12 or i == 0:
            cum = "—"
        else:
            cum = f"{(yf / baseline - 1.0) * 100.0:+.2f}%"
        out.append([delta, cum])
    return out


# Keep old name as thin wrapper for any other callers
def _pct_change_vs_prior(y_vals: list[float]) -> list[str]:
    """One label per point: percent change from previous point (first point → '—')."""
    out: list[str] = []
    for i, y in enumerate(y_vals):
        if i == 0:
            out.append("—")
            continue
        prev = float(y_vals[i - 1])
        yf = float(y)
        if abs(prev) < 1e-12:
            out.append("—")
        else:
            pct = (yf / prev - 1.0) * 100.0
            out.append(f"{pct:+.2f}%")
    return out


_NAV_HOVER = (
    "<b>%{fullData.name}</b><br>"
    "%{x|%b %d, %Y}<br>"
    "$%{y:,.2f}<br>"
    "Δ vs prior: %{customdata[0]}<br>"
    "Cumulative % Chg: %{customdata[1]}<extra></extra>"
)

# Client-facing NAV trace label (net of accrued unpaid fees, not raw NLV).
CLIENT_NAV_TRACE_NAME = "Momentum Pacer — Net of Accrued Fees"

# ── Shared X-axis alignment for the 3 admin-verification graphs ────────────
# Client Net Economic Value chart, Actual NLV chart, and Accrued Unpaid Fees
# chart are stacked for admin verification; giving all three the same date
# range/tick positions/margins (rather than each auto-ranging to its own
# data) makes it easy to compare a value on one chart to the same calendar
# date on another. All three series now start on live inception (the Actual
# NLV chart is trimmed to inception in build_agm_daily_nlv_figure -- the raw
# CSV's earlier pre-inception flat-$30K days are still visible in the raw
# admin daily balances table, just not on this chart), so the shared window
# is simply inception -> the latest daily-balances date.
ADMIN_XAXIS_MARGIN_LR = {"l": 110, "r": 140}
ADMIN_XAXIS_TICKS = {"dtick": "M1", "tickformat": "%b %Y"}


def _admin_shared_xaxis_range() -> tuple[pd.Timestamp, pd.Timestamp] | tuple[None, None]:
    """Shared x-axis window for the 3 admin-verification charts: live
    inception through the latest daily-balances date."""
    if daily_balances_df is None or daily_balances_df.empty:
        return None, None
    x_left = pd.Timestamp(PROGRAM_INCEPTION) - pd.Timedelta(days=7)
    x_right = pd.Timestamp(daily_balances_df["Date"].max()) + pd.Timedelta(days=10)
    return x_left, x_right


def _apply_admin_shared_xaxis(fig: go.Figure) -> go.Figure:
    """Apply the shared date range/ticks/margin so this chart's x-axis lines
    up with the other two admin-verification charts (see ADMIN_XAXIS_MARGIN_LR
    note above)."""
    x_left, x_right = _admin_shared_xaxis_range()
    if x_left is not None:
        fig.update_xaxes(range=[x_left, x_right], **ADMIN_XAXIS_TICKS)
        fig.update_layout(margin=ADMIN_XAXIS_MARGIN_LR)
    return fig


def _daily_equity_frame() -> pd.DataFrame:
    """Client daily equity rows (Date / client_net_value) from live inception onward."""
    if daily_accounting is None or daily_accounting.table.empty:
        return pd.DataFrame(columns=["Date", "client_net_value"])
    t = daily_accounting.table
    mask = t["Date"] >= pd.Timestamp(PROGRAM_INCEPTION)
    return t.loc[mask, ["Date", "client_net_value"]].reset_index(drop=True)


def build_nav_figure() -> go.Figure:
    """
    DAILY NAV chart (client-facing): client net value
    (actual NLV minus accrued unpaid fees) vs daily SPX (^GSPC) and NDX (^NDX)
    benchmarks rebased to the same starting capital at live inception.
    """
    fig = go.Figure()
    eq = _daily_equity_frame()
    if eq.empty:
        fig.add_annotation(text="No data", xref="paper", yref="paper",
                           x=0.5, y=0.5, showarrow=False)
        return fig

    bot_x = [pd.Timestamp(d) for d in eq["Date"]]
    bot_y = [float(v) for v in eq["client_net_value"]]
    base = float(bot_y[0])  # $30,000 at inception

    # ── Benchmarks: daily closes on AGM trading days, rebased to the start ───
    spx_y = [float(v) if pd.notna(v) else None
             for v in agm_bench.rebase(agm_bench.align_to_dates(spx_daily_df, bot_x), base)]
    ndx_y = [float(v) if pd.notna(v) else None
             for v in agm_bench.rebase(agm_bench.align_to_dates(ndx_daily_df, bot_x), base)]

    bot_hover_data = _hover_customdata(bot_y, base)
    fig.add_trace(go.Scatter(
        x=bot_x, y=bot_y,
        mode="lines",
        line={"color": PRIMARY_COLOR, "width": 2.5},
        name=CLIENT_NAV_TRACE_NAME,
        customdata=np.asarray(bot_hover_data, dtype=object),
        hovertemplate=_NAV_HOVER,
        yaxis="y",
    ))

    if any(v is not None for v in spx_y):
        spx_hover_data = _hover_customdata(
            [v if v is not None else base for v in spx_y], base)
        fig.add_trace(go.Scatter(
            x=bot_x, y=spx_y,
            mode="lines", line={"color": "#E67E22", "dash": "dash", "width": 1.5},
            name="S&P 500 (rebased, daily)", opacity=0.8,
            customdata=np.asarray(spx_hover_data, dtype=object),
            hovertemplate=_NAV_HOVER,
            yaxis="y",
        ))

    if any(v is not None for v in ndx_y):
        ndx_hover_data = _hover_customdata(
            [v if v is not None else base for v in ndx_y], base)
        fig.add_trace(go.Scatter(
            x=bot_x, y=ndx_y,
            mode="lines", line={"color": "#8E44AD", "dash": "dot", "width": 1.5},
            name="Nasdaq-100 (rebased, daily)", opacity=0.8,
            customdata=np.asarray(ndx_hover_data, dtype=object),
            hovertemplate=_NAV_HOVER,
            yaxis="y",
        ))

    # ── Inception annotation ──────────────────────────────────────────────────
    fig.add_annotation(
        x=pd.Timestamp(PROGRAM_INCEPTION),
        y=STARTING_CAPITAL,
        text=f"Live Inception<br>{PROGRAM_INCEPTION.strftime('%b %d')}<br>${STARTING_CAPITAL:,.0f}",
        showarrow=True,
        arrowhead=2,
        arrowsize=1,
        arrowwidth=1.5,
        arrowcolor=PRIMARY_COLOR,
        # Straight vertical: label centered above the point (ax=0)
        ax=0,
        ay=-72,
        xanchor="center",
        yanchor="bottom",
        font={"size": 9, "color": PRIMARY_COLOR},
        bgcolor="rgba(255,255,255,0.9)",
        bordercolor=PRIMARY_COLOR,
        borderwidth=1.5,
    )

    # Legend below the plot (not y≈1) so it never collides with the title
    fig.update_layout(
        title={
            "text": "<u>Compounded NAV Since Inception</u>",
            "x": 0.5,
            "xanchor": "center",
            "y": 0.97,
            "yanchor": "top",
            "pad": {"t": 8, "b": 4},
        },
        template="ggplot2",
        plot_bgcolor=GREY_BG,
        paper_bgcolor=WHITE_BG,
        xaxis_title="Date",
        autosize=True,
        margin={"l": 110, "r": 140, "t": 72, "b": 128},
        legend={
            "orientation": "h",
            "yanchor": "top",
            "y": -0.30,
            "xanchor": "center",
            "x": 0.5,
            "bgcolor": "rgba(255,255,255,0.92)",
            "bordercolor": "rgba(0,0,0,0.12)",
            "borderwidth": 1,
            "font": {"size": 11},
        },
    )
    x_left, x_right = _admin_shared_xaxis_range()
    if x_left is None:
        x_left = min(pd.Timestamp(t) for t in bot_x) - pd.Timedelta(days=7)
        x_right = max(pd.Timestamp(t) for t in bot_x) + pd.Timedelta(days=10)
    fig.update_xaxes(
        showgrid=True,
        automargin=True,
        title_standoff=12,
        range=[x_left, x_right],
        **ADMIN_XAXIS_TICKS,
    )
    # ── Compute shared y range ────────────────────────────────────────────────
    # Both axes share the same numeric dollar range. yaxis2 is just a relabelled
    # mirror: its tick POSITIONS are dollar values, its tick LABELS are % vs baseline.
    all_y = [float(v) for v in bot_y + spx_y + ndx_y if v is not None]
    y_min = min(all_y)
    y_max = max(all_y)
    pad = max((y_max - y_min) * 0.05, STARTING_CAPITAL * 0.01)
    y_lo = y_min - pad
    y_hi = y_max + pad

    # Choose a round dollar step that gives ~5-8 ticks across the visible range.
    raw_step = (y_hi - y_lo) / 6.0
    for step in [500, 1000, 2000, 2500, 5000, 10000, 20000]:
        if step >= raw_step:
            dollar_step = step
            break
    else:
        dollar_step = 10000

    first_tick = math.ceil(y_lo / dollar_step) * dollar_step
    tickvals: list[float] = []
    t = first_tick
    while t <= y_hi:
        tickvals.append(float(t))
        t += dollar_step

    # Plotly only renders yaxis2 when at least one trace is assigned to it.
    # This invisible trace spans the full dollar range and forces the axis to appear.
    fig.add_trace(go.Scatter(
        x=[bot_x[0], bot_x[-1]],
        y=[y_lo, y_hi],
        mode="lines",
        line={"color": "rgba(0,0,0,0)", "width": 0},
        showlegend=False,
        hoverinfo="skip",
        yaxis="y2",
    ))

    # Each tickval is a dollar amount; label it as % change from STARTING_CAPITAL.
    # Use "0%" for the baseline tick, "+N%" for positive, "−N%" for negative.
    def _pct_label(v: float) -> str:
        pct = (v / STARTING_CAPITAL - 1.0) * 100.0
        if abs(pct) < 0.5:
            return "0%"
        return f"+{pct:.0f}%" if pct > 0 else f"{pct:.0f}%"

    ticktext = [_pct_label(v) for v in tickvals]

    fig.update_layout(
        yaxis=dict(
            title=dict(
                text=f"NAV (${STARTING_CAPITAL:,.0f} baseline)",
                standoff=25,
            ),
            range=[y_lo, y_hi],
            showgrid=True,
            automargin=True,
            tickprefix="$",
            tickformat=",.0f",
            zeroline=False,
        ),
        yaxis2=dict(
            title=dict(
                text="Return (%)",
                standoff=15,
            ),
            overlaying="y",
            side="right",
            anchor="x",
            # Same dollar range as yaxis — ticks placed at dollar values,
            # labelled as % change vs STARTING_CAPITAL.
            range=[y_lo, y_hi],
            tickmode="array",
            tickvals=tickvals,
            ticktext=ticktext,
            showticklabels=True,
            visible=True,
            showgrid=False,
            zeroline=False,
            automargin=True,
        ),
    )
    return fig


def _empty_admin_figure(title: str, message: str) -> go.Figure:
    fig = go.Figure()
    fig.add_annotation(
        text=message, xref="paper", yref="paper", x=0.5, y=0.5,
        showarrow=False, font=dict(size=13, color="#6c757d"),
    )
    fig.update_layout(
        title=title, height=320,
        xaxis=dict(visible=False), yaxis=dict(visible=False),
        margin=dict(t=50, b=30),
    )
    return fig


def build_agm_accrued_fees_figure() -> go.Figure:
    """Admin-only. Daily accrued unpaid fees from the accounting model."""
    acc_tbl = daily_accounting.table
    if acc_tbl is None or acc_tbl.empty:
        return _empty_admin_figure(
            "Accrued Unpaid Fees (Admin)",
            "Daily accounting unavailable (balances CSV or SPX benchmark data missing).",
        )
    inception_mask = acc_tbl["Date"] >= pd.Timestamp(PROGRAM_INCEPTION)
    acc = acc_tbl.loc[inception_mask]
    if acc.empty:
        return _empty_admin_figure(
            "Accrued Unpaid Fees (Admin)",
            "No post-inception accounting rows to plot.",
        )
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=acc["Date"], y=acc["accrued_unpaid_fees"], mode="lines",
        name="Accrued Unpaid Fees (daily)",
        line=dict(color="#B02A37", width=2),
        hovertemplate="%{x|%b %d, %Y}<br>Accrued: $%{y:,.2f}<extra></extra>",
    ))
    payments = daily_accounting.payments
    if payments:
        pay_dates = [p["date"] for p in payments]
        pay_y = [
            float(acc.loc[acc["Date"] == pd.Timestamp(p["date"]), "accrued_unpaid_fees"].iloc[0])
            for p in payments
        ]
        fig.add_trace(go.Scatter(
            x=pay_dates, y=pay_y,
            mode="markers", name="Fee payment (evidenced)",
            marker=dict(symbol="triangle-down", size=10, color="#1B4F8A"),
            hovertemplate="%{x|%b %d, %Y}<br>Payment: $%{customdata:,.2f}<extra></extra>",
            customdata=[p["amount"] for p in payments],
        ))
    outstanding = daily_fee_accrual.outstanding
    if outstanding:
        note = "Outstanding (no payment evidence in CSV): " + ", ".join(
            f"{o['month']} ${o['fee']:,.2f}" for o in outstanding
        )
        fig.add_annotation(
            text=note, xref="paper", yref="paper", x=0.0, y=1.10,
            showarrow=False, font=dict(size=10, color="#6c757d"), align="left",
        )
    fig.update_layout(
        title="Accrued Unpaid Fees (Admin — daily fee liability net of evidenced payments)",
        xaxis_title="Date",
        yaxis=dict(title="Accrued Unpaid Fees ($)", tickprefix="$", tickformat=",.0f"),
        height=340,
        margin=dict(t=70, b=40),
        legend=dict(orientation="h", yanchor="bottom", y=-0.35, x=0.5, xanchor="center",
                    font=dict(size=10)),
        showlegend=True,
    )
    _apply_admin_shared_xaxis(fig)
    return fig


# ── Admin-only DAILY view builders (raw TradeStation NLV — never client-facing) ──

def build_agm_daily_nlv_figure() -> go.Figure:
    """Admin only. Actual daily account NLV straight from the TradeStation
    balances CSV ("Net Worth"), from live inception onward. This is the real
    operational account value, not the client-facing performance curve, so
    it is clearly labelled and only ever rendered inside admin TearSheet
    mode. Trimmed to inception (rather than the full CSV, which also holds
    pre-inception flat-$30K days) so this chart starts on the same date as
    the other 2 admin-verification charts -- see _apply_admin_shared_xaxis."""
    if daily_balances_df is None or daily_balances_df.empty:
        return _empty_admin_figure(
            "Admin NLV / TradeStation Net Worth",
            "Daily balances CSV not loaded — no NLV series to show.",
        )
    df = daily_balances_df[daily_balances_df["Date"] >= pd.Timestamp(PROGRAM_INCEPTION)]
    if df.empty:
        return _empty_admin_figure(
            "Admin NLV / TradeStation Net Worth",
            "No post-inception daily balances to show.",
        )
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=df["Date"], y=df["Net Worth"], mode="lines",
        name="TradeStation Net Worth",
        line=dict(color=PRIMARY_COLOR, width=2),
        hovertemplate="%{x|%b %d, %Y}<br>$%{y:,.2f}<extra></extra>",
    ))
    fig.update_layout(
        title="Actual NLV / TradeStation Net Worth",
        xaxis_title="Date",
        yaxis=dict(title="Net Worth ($)", tickprefix="$", tickformat=",.0f"),
        height=340,
        margin=dict(t=60, b=40),
        showlegend=False,
    )
    _apply_admin_shared_xaxis(fig)
    return fig


def _fmt_spx(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "—"
    return f"{float(v):,.2f}"


def _fmt_money(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "—"
    return f"${v:,.2f}"


def _fmt_pct(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "—"
    return f"{v:+.2f}%"


# ── Admin reconciliation panel (date-based spot-check, admin-only) ─────────
# TradeStation NLV / Actual NLV = Client Net Economic Value + Accrued Unpaid
# Incentive Fee. Never a new calculation — just a per-date readout of the
# already-accepted daily accounting table (algominds_daily_accounting).
AGM_RECON_DATE_PICKER_ID = "agm-recon-date-picker"
AGM_RECON_OUTPUT_ID = "agm-recon-output"

# Cent-level display tolerance for the reconciliation checkmark. Looser than
# the accounting model's own invariant tolerance (1e-6 — effectively float
# rounding noise; see algominds_daily_accounting.verify_accounting_invariant),
# but still tight enough that any real formula/data defect would fail it.
RECONCILIATION_TOLERANCE = 0.01


def _recon_date_bounds() -> tuple[pd.Timestamp, pd.Timestamp] | tuple[None, None]:
    if daily_accounting is None or daily_accounting.table.empty:
        return None, None
    dates = daily_accounting.table["Date"]
    return pd.Timestamp(dates.min()), pd.Timestamp(dates.max())


def _agm_reconciliation_lookup(requested_date) -> dict:
    """
    Look up the accounting row for *requested_date* (a date string /
    pd.Timestamp / None). Falls back to the nearest available date on or
    before the request when the exact date has no row (e.g. a weekend);
    reports honestly when no data exists at or before the request. Never
    fabricates a date or a row — always the accepted daily accounting table.
    """
    table = daily_accounting.table if daily_accounting is not None else pd.DataFrame()
    if table is None or table.empty:
        return {"available": False, "reason": "No daily accounting data loaded."}

    dates = pd.DatetimeIndex(table["Date"])
    if requested_date is None:
        target = dates.max()
    else:
        try:
            target = pd.Timestamp(requested_date).normalize()
        except (ValueError, TypeError):
            return {"available": False, "reason": f"Invalid date: {requested_date!r}."}

    exact = table.loc[dates == target]
    if not exact.empty:
        row = exact.iloc[0]
        exact_match = True
    else:
        prior = table.loc[dates <= target]
        if prior.empty:
            return {
                "available": False,
                "reason": (
                    f"No accounting data on or before {target.strftime('%Y-%m-%d')} "
                    f"(earliest available: {dates.min().strftime('%Y-%m-%d')})."
                ),
            }
        row = prior.iloc[-1]
        exact_match = False

    actual_nlv = float(row["actual_nlv"])
    client_net = float(row["client_net_value"])
    accrued = float(row["accrued_unpaid_fees"])
    residual = actual_nlv - (client_net + accrued)
    return {
        "available": True,
        "requested_date": target,
        "row_date": pd.Timestamp(row["Date"]),
        "exact_match": exact_match,
        "actual_nlv": actual_nlv,
        "client_net_value": client_net,
        "accrued_unpaid_fees": accrued,
        "residual": residual,
        "within_tolerance": abs(residual) <= RECONCILIATION_TOLERANCE,
    }


def build_agm_reconciliation_panel(selected_date=None):
    """Admin-only reconciliation widget: TradeStation NLV / Actual NLV =
    Client Net Economic Value + Accrued Unpaid Incentive Fee, spot-checked
    for one date (defaults to the latest available date)."""
    result = _agm_reconciliation_lookup(selected_date)
    if not result["available"]:
        return dbc.Alert(result["reason"], color="secondary", className="mb-0")

    status_ok = result["within_tolerance"]
    status_badge = dbc.Badge(
        "✓ Reconciles" if status_ok else "⚠ Does not reconcile",
        color="success" if status_ok else "danger",
        className="ms-2",
    )
    formula_line = (
        f"TradeStation NLV (${result['actual_nlv']:,.2f}) = "
        f"Client Net Economic Value (${result['client_net_value']:,.2f}) + "
        f"Accrued Unpaid Incentive Fee (${result['accrued_unpaid_fees']:,.2f}) "
        f"{'✓' if status_ok else '⚠'}"
    )
    rows = [
        ("Date", result["row_date"].strftime("%Y-%m-%d")),
        ("TradeStation NLV / Actual NLV", _fmt_money(result["actual_nlv"])),
        ("Client Net Economic Value", _fmt_money(result["client_net_value"])),
        ("Accrued Unpaid Incentive Fee", _fmt_money(result["accrued_unpaid_fees"])),
        ("Residual (should be ~$0.00)", f"${result['residual']:+.6f}"),
    ]
    table = dbc.Table(
        html.Tbody([
            html.Tr([html.Td(label, className="fw-bold"), html.Td(value)])
            for label, value in rows
        ]),
        bordered=True, size="sm", className="mb-2",
    )
    children = []
    if not result["exact_match"]:
        children.append(html.Div(
            f"No accounting row for {result['requested_date'].strftime('%Y-%m-%d')} "
            f"(weekend/holiday) — showing the nearest prior trading day, "
            f"{result['row_date'].strftime('%Y-%m-%d')}.",
            className="small text-muted fst-italic mb-2",
        ))
    children.append(table)
    children.append(html.Div(
        [formula_line, status_badge],
        className="d-flex align-items-center flex-wrap gap-1",
    ))
    return html.Div(children)


def build_agm_daily_kpi_cards():
    """Small KPI summary from the latest accounting row (admin only)."""
    if daily_accounting is None or daily_accounting.table.empty:
        return html.Div()
    row = daily_accounting.table.iloc[-1]
    cards = [
        ("Actual NLV (TradeStation)", _fmt_money(row["actual_nlv"])),
        ("Client Net Value", _fmt_money(row["client_net_value"])),
        ("Accrued Unpaid Fees", _fmt_money(row["accrued_unpaid_fees"])),
        ("Daily %", _fmt_pct(row.get("daily_pct"))),
        ("As of", pd.Timestamp(row["Date"]).strftime("%b %d, %Y")),
    ]
    return dbc.Row(
        [
            dbc.Col(
                dbc.Card(
                    dbc.CardBody([
                        html.Div(label, className="small text-muted"),
                        html.Div(value, className="fw-bold"),
                    ]),
                    className="text-center",
                ),
                xs=6, md=True, className="mb-2",
            )
            for label, value in cards
        ],
        className="g-2",
    )


# ── Client-facing daily table (collapsed by default, public/gate-safe) ─────
# Same trust tier as the client NAV chart above (rendered directly in
# serve_layout, gated only by the Important Notice accept screen — never the
# admin auth), so it never carries admin-only operational columns (Cash
# Balance, margin figures, Buying Power/Margin Deficit) and always uses the
# client-safe "TradeStation NLV / Statement Value" vs "Client Net Economic
# Value" terminology so the two are never confused.
CLIENT_DAILY_TOGGLE_ID = "agm-client-daily-toggle-btn"
CLIENT_DAILY_COLLAPSE_ID = "agm-client-daily-collapse"
CLIENT_DAILY_TABLE_ID = "agm-client-daily-table"
CLIENT_DAILY_PAGE_SIZE_ID = "agm-client-daily-page-size"
CLIENT_DAILY_EXPORT_BTN_ID = "agm-client-daily-export-btn"
CLIENT_DAILY_EXPORT_DOWNLOAD_ID = "agm-client-daily-export-download"

_MONEY_FMT = Format(precision=2, scheme=Scheme.fixed, symbol=Symbol.yes, symbol_prefix="$")
_MONEY_SIGNED_FMT = Format(
    precision=2, scheme=Scheme.fixed, symbol=Symbol.yes, symbol_prefix="$", sign=Sign.positive
)
_PCT_SIGNED_FMT = Format(
    precision=2, scheme=Scheme.fixed, symbol=Symbol.yes, symbol_suffix="%", sign=Sign.positive
)
_INDEX_FMT = Format(precision=2, scheme=Scheme.fixed)

# (display label, accounting-table column key, format kind). "Momentum daily %"
# and "Daily %" both read momentum_daily_pct's client-net-based value today
# (the accounting model defines daily_pct as an alias of momentum_daily_pct);
# both display names are kept as distinct columns per the approved spec.
CLIENT_DAILY_TABLE_COLUMNS: list[tuple[str, str, str]] = [
    ("Date", "Date", "date"),
    ("Client Net Economic Value", "client_net_value", "money"),
    ("TradeStation NLV / Statement Value", "actual_nlv", "money"),
    ("Accrued Unpaid Incentive Fee", "accrued_unpaid_fees", "money"),
    ("Daily $", "daily_dollar", "money_signed"),
    ("Daily %", "daily_pct", "pct_signed"),
    ("Since inception %", "since_inception_pct", "pct_signed"),
    ("SPX Close", "spx_close", "index"),
    ("SPX daily %", "spx_daily_pct", "pct_signed"),
    ("Momentum daily %", "momentum_daily_pct", "pct_signed"),
    ("Momentum vs SPX daily spread %", "momentum_vs_spx_daily_spread_pct", "pct_signed"),
    ("Fee payment", "fee_payment", "money_signed"),
]

# Two-row DataTable headers — keeps each column narrow so the Daily Returns
# table needs less horizontal scrolling (Dash ``name`` as [top, bottom]).
# The rows read naturally with merge_duplicate_headers=True: adjacent columns
# sharing a top line ("Daily" over $/%, "SPX" over Close/daily %) merge into
# one spanning group header, and short single-line labels sit on the BOTTOM
# row (blank top) so every specific label lines up along the same baseline.
_DAILY_TABLE_HEADER_ROWS: dict[str, tuple[str, str]] = {
    "Date": ("\u00a0", "Date"),
    "Client Net Economic Value": ("Client Net", "Economic Value"),
    "TradeStation NLV / Statement Value": ("TradeStation NLV", "Statement Value"),
    "Accrued Unpaid Incentive Fee": ("Accrued Unpaid", "Incentive Fee"),
    "Daily $": ("Daily", "$"),
    "Daily %": ("Daily", "%"),
    "Since inception %": ("Since", "inception %"),
    "SPX Close": ("SPX", "Close"),
    "SPX daily %": ("SPX", "daily %"),
    "Momentum daily %": ("Momentum", "daily %"),
    "Momentum vs SPX daily spread %": ("Momentum vs SPX", "daily spread %"),
    "Fee payment": ("\u00a0", "Fee payment"),
}


def _daily_table_header_name(label: str) -> list[str]:
    top, bottom = _DAILY_TABLE_HEADER_ROWS.get(label, (label, "\u00a0"))
    return [top, bottom]


def _daily_table_column_defs(column_spec: list[tuple[str, str, str]]) -> list[dict]:
    """DataTable column defs from a (label, key, kind) spec — shared by the
    client and admin Daily Returns tables so both format money/percent cells
    identically."""
    cols = []
    for label, key, kind in column_spec:
        header_name = _daily_table_header_name(label)
        if kind == "date":
            cols.append({"name": header_name, "id": key})
        elif kind == "money":
            cols.append({"name": header_name, "id": key, "type": "numeric", "format": _MONEY_FMT})
        elif kind == "money_signed":
            cols.append({"name": header_name, "id": key, "type": "numeric", "format": _MONEY_SIGNED_FMT})
        elif kind == "pct_signed":
            cols.append({"name": header_name, "id": key, "type": "numeric", "format": _PCT_SIGNED_FMT})
        elif kind == "index":
            cols.append({"name": header_name, "id": key, "type": "numeric", "format": _INDEX_FMT})
        else:  # pragma: no cover - defensive, every kind above is handled
            cols.append({"name": header_name, "id": key})
    return cols


def _build_client_daily_table_columns() -> list[dict]:
    return _daily_table_column_defs(CLIENT_DAILY_TABLE_COLUMNS)


def build_client_daily_table_rows(newest_first: bool = True, table=None) -> list[dict]:
    """Row dicts for the client-facing DataTable: client-safe columns only,
    from live inception onward (matches the client NAV/drawdown charts).
    *table* overrides the accounting frame (used by the admin Add Row flow to
    show admin-entered rows recomputed through the accepted model)."""
    t = daily_accounting.table if table is None else table
    if t is None or t.empty:
        return []
    mask = t["Date"] >= pd.Timestamp(PROGRAM_INCEPTION)
    df = t.loc[mask].reset_index(drop=True)
    if newest_first:
        df = df.iloc[::-1].reset_index(drop=True)

    def _num(r, col):
        v = r.get(col)
        return float(v) if pd.notna(v) else None

    rows = []
    for _, r in df.iterrows():
        rows.append({
            "Date": pd.Timestamp(r["Date"]).strftime("%Y-%m-%d"),
            "client_net_value": _num(r, "client_net_value"),
            "actual_nlv": _num(r, "actual_nlv"),
            "accrued_unpaid_fees": _num(r, "accrued_unpaid_fees"),
            "daily_dollar": _num(r, "daily_dollar"),
            "daily_pct": _num(r, "daily_pct"),
            "since_inception_pct": _num(r, "since_inception_pct"),
            "spx_close": _num(r, "spx_close"),
            "spx_daily_pct": _num(r, "spx_daily_pct"),
            "momentum_daily_pct": _num(r, "momentum_daily_pct"),
            "momentum_vs_spx_daily_spread_pct": _num(r, "momentum_vs_spx_daily_spread_pct"),
            "fee_payment": _num(r, "fee_payment"),
        })
    return rows


# ── Admin-only Daily Returns controls (TKP/TCP pattern, inside the client card) ──
# There is no separate bottom admin table anymore (removed 2026-07-08); instead
# the client Daily Returns card gains an admin toolbar — rendered server-side
# by an auth-gated callback into AGM_DAILY_ADMIN_SLOT_ID, so none of these
# controls ever reach a non-admin browser.
AGM_DAILY_ADMIN_SLOT_ID = "agm-daily-admin-slot"
AGM_DAILY_ADMIN_COL_PICKER_ID = "agm-daily-admin-col-picker"
AGM_DAILY_ADMIN_ADD_BTN_ID = "agm-daily-admin-add-btn"
AGM_DAILY_ADMIN_ADD_MODAL_ID = "agm-daily-admin-add-modal"
AGM_DAILY_ADMIN_ADD_DATE_ID = "agm-daily-admin-add-date"
AGM_DAILY_ADMIN_ADD_NLV_ID = "agm-daily-admin-add-nlv"
AGM_DAILY_ADMIN_ADD_DEPOSIT_ID = "agm-daily-admin-add-deposit"
AGM_DAILY_ADMIN_ADD_FEE_PAID_ID = "agm-daily-admin-add-fee-paid"
AGM_DAILY_ADMIN_ADD_SAVE_ID = "agm-daily-admin-add-save"
AGM_DAILY_ADMIN_ADD_CANCEL_ID = "agm-daily-admin-add-cancel"
AGM_DAILY_ADMIN_ADD_ERROR_ID = "agm-daily-admin-add-error"
AGM_DAILY_ADMIN_DELETE_BTN_ID = "agm-daily-admin-delete-btn"
AGM_DAILY_ADMIN_DELETE_MODAL_ID = "agm-daily-admin-delete-modal"
AGM_DAILY_ADMIN_DELETE_BODY_ID = "agm-daily-admin-delete-body"
AGM_DAILY_ADMIN_DELETE_CONFIRM_ID = "agm-daily-admin-delete-confirm"
AGM_DAILY_ADMIN_DELETE_CANCEL_ID = "agm-daily-admin-delete-cancel"
AGM_DAILY_ADMIN_CALC_BTN_ID = "agm-daily-admin-calc-btn"
AGM_DAILY_ADMIN_CALC_MODAL_ID = "agm-daily-admin-calc-modal"
AGM_DAILY_ADMIN_CALC_CLOSE_ID = "agm-daily-admin-calc-close"

# Every admin-only control id — tests assert none of these leak into the
# public/client layout (only the empty slot div ships publicly).
AGM_DAILY_ADMIN_CONTROL_IDS = (
    AGM_DAILY_ADMIN_COL_PICKER_ID,
    AGM_DAILY_ADMIN_ADD_BTN_ID,
    AGM_DAILY_ADMIN_ADD_MODAL_ID,
    AGM_DAILY_ADMIN_DELETE_BTN_ID,
    AGM_DAILY_ADMIN_DELETE_MODAL_ID,
    AGM_DAILY_ADMIN_CALC_BTN_ID,
    AGM_DAILY_ADMIN_CALC_MODAL_ID,
)

# Simplified accounting identity — the ONLY calculation the AGM Show
# Calculations modal presents (never the internal fee-engine mechanics).
AGM_ACCOUNTING_IDENTITY_TEXT = (
    "TradeStation NLV = Client Net Economic Value + Accrued Unpaid Incentive Fee"
)
AGM_ACCOUNTING_IDENTITY_INVERSE_TEXT = (
    "Client Net Economic Value = TradeStation NLV - Accrued Unpaid Incentive Fee"
)

# Admin-entered daily rows (Date + TradeStation NLV only), persisted separately
# from the TradeStation CSV export (which the UI never writes to). Mirrors
# tkp_ts.py's JSON round-trip pattern for its Daily Returns editor. Every other
# column is derived by re-running the accepted daily accounting model over the
# CSV rows plus these manual rows — nothing is hand-entered twice.
AGM_MANUAL_DAILY_ROWS_FILENAME = "momentum_pacer_manual_daily_rows.json"


def _agm_manual_daily_rows_path():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), AGM_MANUAL_DAILY_ROWS_FILENAME)


def _load_agm_manual_daily_rows():
    path = _agm_manual_daily_rows_path()
    if not os.path.isfile(path):
        return []
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError):
        return []
    return data if isinstance(data, list) else []


def _save_agm_manual_daily_rows(rows):
    path = _agm_manual_daily_rows_path()
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(rows, f, indent=2)
    except OSError as e:
        print(f"⚠️ Could not save Momentum Pacer manual daily rows to {path}: {e}")


def _agm_manual_fee_payments(manual_rows):
    """Admin-entered 'Incentive Fee Paid' amounts as FeePaymentEvidence, on top
    of the hand-confirmed EVIDENCED_FEE_PAYMENTS list — same evidence
    mechanism (algominds_fee_payment_evidence), just a second source. Reduces
    accrued_unpaid_fees for that date onward; never touches the fee formula."""
    extra = [
        agm_fee_evidence.FeePaymentEvidence(
            date=pd.Timestamp(r["date"]).normalize(),
            description="Admin-entered incentive fee payment",
            amount=float(r["incentive_fee_paid"]),
        )
        for r in manual_rows
        if float(r.get("incentive_fee_paid") or 0) > 0
    ]
    if not extra:
        return None
    return tuple(agm_fee_evidence.EVIDENCED_FEE_PAYMENTS) + tuple(extra)


def _compute_accounting_with_manual_rows(manual_rows) -> pd.DataFrame:
    """Daily accounting table for the CSV rows plus admin-entered manual rows.

    Manual rows carry Date + TradeStation NLV (deposit_withdrawal is stored
    as display-only metadata — the NLV already reflects the cash debit/credit,
    see build_agm_daily_admin_controls). Every derived column (accrued fee,
    client net value, SPX alignment, daily returns) comes from re-running the
    SAME accepted accounting model over the augmented balance frame — the fee
    formula itself is never touched here. A manually recorded incentive fee
    payment is passed through as additional evidence to the SAME
    exact-daily-match/workbook-reconciliation/cash-transaction mechanism the
    fee engine already uses, so it reduces accrued_unpaid_fees (and therefore
    increases client_net_value back to par) without double-subtracting.
    """
    if not manual_rows:
        return daily_accounting.table
    extra = pd.DataFrame(
        {
            "Date": [pd.Timestamp(r["date"]).normalize() for r in manual_rows],
            "Net Worth": [float(r["actual_nlv"]) for r in manual_rows],
        }
    )
    augmented = (
        pd.concat([daily_balances_df, extra], ignore_index=True)
        .sort_values("Date")
        .reset_index(drop=True)
    )
    acct = agm_accounting.compute_agm_daily_accounting(
        augmented,
        spx_daily_df,
        inception=pd.Timestamp(PROGRAM_INCEPTION),
        monthly_reference=summary_df if not summary_df.empty else None,
        cash_transaction_payments=_agm_manual_fee_payments(manual_rows),
    )
    return acct.table


def agm_add_manual_daily_row(date_val, nlv_val, deposit_val=0, fee_paid_val=0):
    """Validate and persist one admin-entered daily row (Date + TradeStation
    NLV / Statement Value + optional Deposit/Withdrawal + optional Incentive
    Fee Paid). Returns (ok, message, recomputed_table_or_None)."""
    if not date_val:
        return False, "Date is required.", None
    if nlv_val in (None, ""):
        return False, "TradeStation NLV / Statement Value is required.", None
    try:
        date = pd.Timestamp(date_val).normalize()
    except (ValueError, TypeError):
        return False, f"Invalid date: {date_val!r}.", None
    try:
        nlv = float(nlv_val)
    except (ValueError, TypeError):
        return False, "TradeStation NLV / Statement Value must be a number.", None
    if nlv <= 0:
        return False, "TradeStation NLV / Statement Value must be positive.", None
    try:
        deposit = float(deposit_val) if deposit_val not in (None, "") else 0.0
    except (ValueError, TypeError):
        return False, "Deposit / Withdrawal must be a number.", None
    try:
        fee_paid = float(fee_paid_val) if fee_paid_val not in (None, "") else 0.0
    except (ValueError, TypeError):
        return False, "Incentive Fee Paid must be a number.", None
    if fee_paid < 0:
        return False, "Incentive Fee Paid must not be negative.", None

    manual = _load_agm_manual_daily_rows()
    latest_known = pd.Timestamp(daily_balances_df["Date"].max())
    if manual:
        latest_known = max(
            latest_known, max(pd.Timestamp(r["date"]) for r in manual)
        )
    if date <= latest_known:
        return (
            False,
            f"Date must be after the latest existing daily row "
            f"({latest_known.strftime('%Y-%m-%d')}) — TradeStation CSV rows are "
            f"never overwritten from the tearsheet.",
            None,
        )

    manual.append({
        "date": date.strftime("%Y-%m-%d"),
        "actual_nlv": nlv,
        "deposit_withdrawal": deposit,
        "incentive_fee_paid": fee_paid,
    })
    _save_agm_manual_daily_rows(manual)
    return True, "", _compute_accounting_with_manual_rows(manual)


def agm_delete_last_manual_daily_row():
    """Remove the most recent admin-entered daily row (never a CSV row).
    Returns (ok, message, recomputed_table_or_None)."""
    manual = _load_agm_manual_daily_rows()
    if not manual:
        return (
            False,
            "No manually added daily rows to delete — TradeStation CSV rows "
            "are never deleted from the tearsheet.",
            None,
        )
    manual.sort(key=lambda r: r["date"])
    removed = manual.pop()
    _save_agm_manual_daily_rows(manual)
    return (
        True,
        f"Deleted manually added row for {removed['date']}.",
        _compute_accounting_with_manual_rows(manual),
    )


def _default_admin_add_row_date() -> str:
    """Previous business day (Mon -> Fri) as YYYY-MM-DD — mirrors tkp_ts.py's
    Add Row date default exactly (see tearsheet_date_defaults.py) instead of
    the account's own latest-known-row date."""
    return default_add_row_date_str()


def build_agm_show_calculations_body():
    """Simplified accounting identity ONLY — no fee-engine mechanics, no
    workbook internals, no monthly summary fields."""
    return dbc.ModalBody([
        html.H6("Daily accounting identity", className="fw-bold small"),
        html.Pre(
            AGM_ACCOUNTING_IDENTITY_TEXT,
            className="bg-light p-2 rounded",
            style={"fontFamily": "monospace", "fontSize": "12px", "whiteSpace": "pre-wrap"},
        ),
        html.Pre(
            AGM_ACCOUNTING_IDENTITY_INVERSE_TEXT,
            className="bg-light p-2 rounded",
            style={"fontFamily": "monospace", "fontSize": "12px", "whiteSpace": "pre-wrap"},
        ),
        html.Ul([
            html.Li([
                html.Strong("TradeStation NLV / Statement Value"),
                " — the brokerage account value shown on TradeStation statements. "
                "It still includes any incentive fee accrued but not yet paid to the "
                "CTA, so it is not the client's true net value.",
            ]),
            html.Li([
                html.Strong("Accrued Unpaid Incentive Fee"),
                " — the CTA incentive fee owed/accrued but not yet paid.",
            ]),
            html.Li([
                html.Strong("Client Net Economic Value"),
                " — the client's true economic value after the unpaid incentive "
                "fee accrual is taken into account.",
            ]),
        ], className="small mb-0"),
    ])


def build_agm_daily_admin_controls():
    """Admin toolbar + modals for the Daily Returns card (rendered only for an
    authenticated admin session): Visible Columns picker, Add Row (Date +
    TradeStation NLV / Statement Value + Deposit/Withdrawal + Incentive Fee
    Paid), Delete Last Row (manual rows only), and the Show Calculations
    accounting-identity modal."""
    column_labels = [label for label, _, _ in CLIENT_DAILY_TABLE_COLUMNS]
    toolbar = html.Div(
        [
            html.Label("Visible Columns", className="fw-bold small mb-1"),
            dcc.Dropdown(
                id=AGM_DAILY_ADMIN_COL_PICKER_ID,
                options=[{"label": c, "value": c} for c in column_labels],
                value=list(column_labels),
                multi=True,
                clearable=False,
                placeholder="Select columns…",
                style={"marginBottom": "12px"},
            ),
            html.Div(
                [
                    dbc.Button("Add Row", id=AGM_DAILY_ADMIN_ADD_BTN_ID,
                               color="success", size="sm", className="me-2"),
                    dbc.Button("Delete Last Row", id=AGM_DAILY_ADMIN_DELETE_BTN_ID,
                               color="danger", size="sm", className="me-2"),
                    dbc.Button("Show Calculations", id=AGM_DAILY_ADMIN_CALC_BTN_ID,
                               color="info", size="sm", className="me-2"),
                ],
                className="mb-3",
            ),
        ],
    )
    add_modal = dbc.Modal([
        dbc.ModalHeader("Add Row"),
        dbc.ModalBody([
            dbc.Label("Date"),
            dbc.Input(id=AGM_DAILY_ADMIN_ADD_DATE_ID, type="date",
                      value=_default_admin_add_row_date()),
            dbc.Label("TradeStation NLV / Statement Value", className="mt-2"),
            dbc.Input(id=AGM_DAILY_ADMIN_ADD_NLV_ID, type="number", step="0.01"),
            dbc.Label("Deposit / Withdrawal", className="mt-2"),
            dbc.Input(id=AGM_DAILY_ADMIN_ADD_DEPOSIT_ID, type="number", step="0.01", value=0),
            html.P(
                "(negative number = withdrawal)",
                className="small text-muted fst-italic mt-1 mb-0",
            ),
            dbc.Label("Incentive Fee Paid", className="mt-2"),
            dbc.Input(id=AGM_DAILY_ADMIN_ADD_FEE_PAID_ID, type="number", step="0.01", min=0, value=0),
            html.P(
                "(positive number; reduces accrued unpaid incentive fee — "
                "does not double-subtract from Client Net Economic Value, "
                "since TradeStation NLV already reflects the cash debit)",
                className="small text-muted fst-italic mt-1 mb-0",
            ),
            html.P(
                "All other columns (SPX close, Accrued Unpaid Incentive Fee, "
                "Client Net Economic Value, daily returns) are derived "
                "automatically from the accepted daily accounting model and the "
                "benchmark cache — never entered by hand.",
                className="small text-muted fst-italic mt-2 mb-1",
            ),
            html.P(
                "Contracts bought + sold per unit is not in the data model yet; "
                "it arrives with the exchange-fee account-model branch.",
                className="small text-muted fst-italic mb-0",
            ),
            html.Div(id=AGM_DAILY_ADMIN_ADD_ERROR_ID, className="text-danger small mt-2"),
        ]),
        dbc.ModalFooter([
            dbc.Button("Save", id=AGM_DAILY_ADMIN_ADD_SAVE_ID, color="primary", size="sm"),
            dbc.Button("Cancel", id=AGM_DAILY_ADMIN_ADD_CANCEL_ID, color="secondary", size="sm"),
        ]),
    ], id=AGM_DAILY_ADMIN_ADD_MODAL_ID, is_open=False, centered=True, size="sm")
    delete_modal = dbc.Modal([
        dbc.ModalHeader(dbc.ModalTitle("Confirm Delete")),
        dbc.ModalBody(html.P(id=AGM_DAILY_ADMIN_DELETE_BODY_ID, className="mb-0")),
        dbc.ModalFooter([
            dbc.Button("Delete", id=AGM_DAILY_ADMIN_DELETE_CONFIRM_ID,
                       color="danger", size="sm", className="me-2"),
            dbc.Button("Cancel", id=AGM_DAILY_ADMIN_DELETE_CANCEL_ID,
                       color="secondary", size="sm"),
        ]),
    ], id=AGM_DAILY_ADMIN_DELETE_MODAL_ID, is_open=False, centered=True, size="sm")
    calc_modal = dbc.Modal([
        dbc.ModalHeader("Show Calculations"),
        build_agm_show_calculations_body(),
        dbc.ModalFooter(
            dbc.Button("Close", id=AGM_DAILY_ADMIN_CALC_CLOSE_ID, color="secondary", size="sm"),
        ),
    ], id=AGM_DAILY_ADMIN_CALC_MODAL_ID, is_open=False, centered=True, size="lg")
    return [toolbar, add_modal, delete_modal, calc_modal]


def build_client_daily_table_section():
    """Titled, collapsed-by-default client-facing daily table (Show/Hide),
    similar in spirit to the sibling tearsheets' 'Daily Returns' section."""
    return dbc.Card(
        [
            dbc.CardHeader(
                html.Div([
                    html.H6("Daily Returns", className="mb-0 d-inline"),
                    dbc.Button(
                        "Show ▾", id=CLIENT_DAILY_TOGGLE_ID, color="link", size="sm",
                        className="float-end p-0 text-decoration-none fw-bold", n_clicks=0,
                    ),
                ]),
            ),
            dbc.Collapse(
                id=CLIENT_DAILY_COLLAPSE_ID,
                is_open=False,
                children=dbc.CardBody([
                    html.P(
                        "TradeStation NLV / Statement Value is the raw brokerage account value "
                        "shown on TradeStation statements — it still includes any incentive fee "
                        "accrued but not yet paid to the CTA. Client Net Economic Value is your "
                        "true economic value after that unpaid fee accrual is taken into account.",
                        className="small text-muted fst-italic mb-3",
                    ),
                    # Admin-only controls (Add Row / Delete Last Row / Show
                    # Calculations / Visible Columns) render into this slot via
                    # a server-side auth-gated callback — the public layout only
                    # ever carries this empty div (TCP admin-toolbar pattern).
                    html.Div(id=AGM_DAILY_ADMIN_SLOT_ID),
                    html.Div(
                        [
                            html.Div(
                                [
                                    html.Span("View per page:", className="me-2 small",
                                               style={"lineHeight": "31px"}),
                                    dcc.Dropdown(
                                        id=CLIENT_DAILY_PAGE_SIZE_ID,
                                        options=[{"label": str(v), "value": v}
                                                 for v in [25, 50, 100, 150, 200]],
                                        value=25,
                                        clearable=False,
                                        style={"width": "80px", "display": "inline-block"},
                                    ),
                                ],
                                style={"display": "inline-flex", "alignItems": "center"},
                            ),
                            html.Div(
                                [
                                    dbc.Button("Export Excel", id=CLIENT_DAILY_EXPORT_BTN_ID,
                                               color="secondary", size="sm"),
                                    dcc.Download(id=CLIENT_DAILY_EXPORT_DOWNLOAD_ID),
                                ],
                                style={"float": "right"},
                            ),
                        ],
                        className="mb-3",
                        style={"display": "flex", "justifyContent": "space-between"},
                    ),
                    dash_table.DataTable(
                        id=CLIENT_DAILY_TABLE_ID,
                        columns=_build_client_daily_table_columns(),
                        data=build_client_daily_table_rows(newest_first=True),
                        merge_duplicate_headers=True,
                        sort_action="native",
                        sort_mode="single",
                        sort_by=[{"column_id": "Date", "direction": "desc"}],
                        page_size=25,
                        editable=False,
                        style_table={"overflowX": "auto"},
                        style_cell={
                            "textAlign": "right",
                            "padding": "4px 6px",
                            "fontSize": "12px",
                            "fontFamily": "monospace",
                            "whiteSpace": "nowrap",
                            "minWidth": "68px",
                            "maxWidth": "96px",
                        },
                        style_cell_conditional=[
                            {"if": {"column_id": "Date"}, "textAlign": "left", "minWidth": "78px", "maxWidth": "86px"},
                            {"if": {"column_id": "spx_close"}, "minWidth": "72px", "maxWidth": "80px"},
                            {"if": {"column_id": "momentum_vs_spx_daily_spread_pct"},
                             "minWidth": "78px", "maxWidth": "92px"},
                        ],
                        style_header={
                            "backgroundColor": PRIMARY_COLOR,
                            "color": "white",
                            "fontWeight": "bold",
                            "fontSize": "10px",
                            "textAlign": "center",
                            "whiteSpace": "normal",
                            "lineHeight": "1.15",
                            "padding": "4px 4px",
                            "height": "auto",
                        },
                        style_data_conditional=[
                            {"if": {"filter_query": "{daily_pct} > 0", "column_id": "daily_pct"},
                             "color": "green"},
                            {"if": {"filter_query": "{daily_pct} < 0", "column_id": "daily_pct"},
                             "color": "red"},
                            {"if": {"filter_query": "{momentum_vs_spx_daily_spread_pct} > 0",
                                    "column_id": "momentum_vs_spx_daily_spread_pct"},
                             "color": "green"},
                            {"if": {"filter_query": "{momentum_vs_spx_daily_spread_pct} < 0",
                                    "column_id": "momentum_vs_spx_daily_spread_pct"},
                             "color": "red"},
                        ],
                    ),
                ]),
            ),
        ],
        className="mb-4",
    )


def build_drawdown_figure() -> go.Figure:
    """Drawdown from peak, DAILY — computed from the client net value curve."""
    fig = go.Figure()
    eq_df = _daily_equity_frame()
    if eq_df.empty:
        return fig

    eq = eq_df["client_net_value"].astype(float)
    pk = eq.cummax()
    dd = ((eq / pk) - 1.0) * 100.0
    dd_x = [pd.Timestamp(d) for d in eq_df["Date"]]
    dd_vals = [float(v) for v in dd.values]

    fig.add_trace(go.Scatter(
        x=dd_x, y=dd_vals,
        mode="lines", fill="tozeroy",
        line={"color": ACCENT_RED, "width": 1.5},
        name="Momentum Pacer Drawdown",
        hovertemplate=(
            "<b>%{fullData.name}</b><br>"
            "%{x|%b %d, %Y}<br>"
            "%{y:.2f}%<extra></extra>"
        ),
    ))

    fig.update_layout(
        title={"text": "<u>Drawdown from Peak</u>",
               "x": 0.5, "xanchor": "center"},
        template="ggplot2",
        plot_bgcolor=GREY_BG, paper_bgcolor=WHITE_BG,
        xaxis_title="Date", yaxis_title="Drawdown (%)",
        autosize=True,
        # Extra bottom margin so "Date" + month ticks never collide with page footer
        margin={"l": 50, "r": 20, "t": 50, "b": 88},
    )
    x_right = max(pd.Timestamp(t) for t in dd_x) + pd.Timedelta(days=10)
    x_left = min(pd.Timestamp(t) for t in dd_x) - pd.Timedelta(days=7)
    fig.update_xaxes(
        showgrid=True,
        automargin=True,
        title_standoff=22,
        range=[x_left, x_right],
    )
    fig.update_yaxes(showgrid=True)
    return fig


# ==============================================================================
# PERFORMANCE SUMMARY TABLE  (spreadsheet style – mirrors screenshot + Y&Q style)
# ==============================================================================

def _cell_style(val_str: str, extra: dict | None = None) -> dict:
    """Green/red background for % values, white otherwise."""
    style = {"fontSize": "0.78rem", "padding": "4px 8px", "whiteSpace": "nowrap",
             "textAlign": "right"}
    try:
        v = float(str(val_str).replace("%", "").replace("$", "").replace(",", ""))
        if v > 0:
            style["backgroundColor"] = "#d4edda"
        elif v < 0:
            style["backgroundColor"] = "#f8d7da"
        else:
            style["backgroundColor"] = "white"
    except (ValueError, AttributeError):
        style["textAlign"] = "left"
    if extra:
        style.update(extra)
    return style


def build_performance_summary_table():
    """
    Spreadsheet-style monthly table matching the screenshot exactly:
    Month | SPX Start | SPX End | NDX Start | NDX End |
    BOT Start | BOT End After Fees |
    SPX Returns% | NDX Returns% |
    BOT Returns Before Fees% | BOT Fees% | BOT Returns After Fees% | Cumulative Net%
    Plus Net% and Net$ footer rows.
    """
    if _display_summary_df.empty:
        return html.P("No data available.", className="text-danger")

    # Header columns — every label is deliberately TWO lines (a leading no-break
    # space where a label is naturally short) so the table stays narrow enough
    # to fit on screen while the header row keeps one uniform height with all
    # labels bottom-aligned along the same baseline.
    cols = [
        " \nMonth",
        "SPX\nStart", "SPX\nEnd",
        "NDX\nStart", "NDX\nEnd",
        "BOT\nStart", "BOT End\nAfter Fees",
        "SPX\nReturns%", "NDX\nReturns%",
        "BOT Returns\nBefore Fees%", "BOT\nFees%",
        "BOT Returns\nAfter Fees%", "Cumul.\nNet%",
    ]

    # Column groups: index | spx | ndx | bot$ | spx% ndx% | bot% fees% net% cumul%
    group_borders = {6: "3px solid #dee2e6", 7: "none", 9: "3px solid #dee2e6"}

    th_style_base = {
        "backgroundColor": GREY_BG, "color": "#000",
        "fontSize": "0.75rem", "padding": "4px 6px",
        "whiteSpace": "pre-wrap", "textAlign": "center",
        "verticalAlign": "bottom", "lineHeight": "1.25",
    }

    header_cells = []
    for i, col in enumerate(cols):
        sty = dict(th_style_base)
        if i in group_borders:
            sty["borderLeft"] = group_borders[i]
        header_cells.append(html.Th(col, style=sty))

    thead = html.Thead(html.Tr(header_cells))

    # Data rows
    body_rows = []
    for _, row in _display_summary_df.iterrows():
        month_label = row["date"].strftime("%b-%Y")

        def fmt_idx(v):
            try: return f"{float(v):,.2f}"
            except: return "—"

        def fmt_dollar(v):
            try: return f"${float(v):,.0f}"
            except: return "—"

        def fmt_pct(v):
            try: return f"{float(v)*100:.2f}%"
            except: return "—"

        vals = [
            month_label,
            fmt_idx(row["spx_start"]),  fmt_idx(row["spx_end"]),
            fmt_idx(row["ndx_start"]),  fmt_idx(row["ndx_end"]),
            fmt_dollar(row["bot_start"]), fmt_dollar(row["bot_end_after_fees"]),
            fmt_pct(row["spx_ret"]),    fmt_pct(row["ndx_ret"]),
            fmt_pct(row["bot_gross_ret"]), fmt_pct(row["bot_fees_pct"]),
            fmt_pct(row["bot_net_ret"]),   fmt_pct(row["cumulative_net"]),
        ]

        cells = []
        for i, v in enumerate(vals):
            sty = _cell_style(v)
            if i == 0:
                sty["textAlign"] = "left"
                sty["fontWeight"] = "500"
                sty["backgroundColor"] = "white"
            # dollar columns: white background, no sign colouring
            if i in (1, 2, 3, 4, 5, 6):
                sty["backgroundColor"] = "white"
                sty["textAlign"] = "right"
            if i in group_borders:
                sty["borderLeft"] = group_borders[i]
            cells.append(html.Td(v, style=sty))
        body_rows.append(html.Tr(cells))

    # Net% footer row
    if net_totals:
        def _nt(k): return f"{float(net_totals.get(k, 0))*100:.2f}%"

        def _sign_bg(val_str: str) -> str:
            """Return green/red/white based on numeric sign. Never raises."""
            try:
                v = float(str(val_str).replace("%", "").replace("$", "").replace(",", ""))
                if v > 0:   return "#d4edda"
                if v < 0:   return "#f8d7da"
                return "white"
            except (ValueError, AttributeError):
                return GREY_BG

        net_pct_vals = [
            "Net %", "", "", "", "", "", "",
            _nt("spx_net_pct"), _nt("ndx_net_pct"),
            _nt("bot_gross_pct"), _nt("bot_fees_pct"),
            _nt("bot_net_pct"), _nt("bot_net_pct"),
        ]
        net_pct_cells = []
        for i, v in enumerate(net_pct_vals):
            bg = GREY_BG if (i == 0 or v == "") else _sign_bg(v)
            sty = {
                "fontSize": "0.78rem", "padding": "4px 8px",
                "whiteSpace": "nowrap", "textAlign": "left" if i == 0 else "right",
                "fontWeight": "700", "backgroundColor": bg,
            }
            if i in group_borders:
                sty["borderLeft"] = group_borders[i]
            net_pct_cells.append(html.Td(v, style=sty))
        body_rows.append(html.Tr(net_pct_cells,
                                  style={"borderTop": "2px solid #333"}))

        # Net$ footer row
        def _nd(k):
            try: return f"${float(net_totals.get(k, 0)):,.2f}"
            except: return "—"
        net_dol_vals = [
            "Net $", "", "", "", "", "", "",
            _nd("spx_net_dollar"), _nd("ndx_net_dollar"),
            _nd("bot_gross_dollar"), _nd("bot_fees_dollar"),
            _nd("bot_net_dollar"), _nd("bot_net_dollar"),
        ]
        net_dol_cells = []
        for i, v in enumerate(net_dol_vals):
            sty = {"fontSize": "0.78rem", "padding": "4px 8px",
                   "whiteSpace": "nowrap", "textAlign": "right",
                   "fontWeight": "700", "backgroundColor": "#E8F4FD"}
            if i == 0:
                sty["textAlign"] = "left"
                sty["backgroundColor"] = GREY_BG
            if i in group_borders:
                sty["borderLeft"] = group_borders[i]
            net_dol_cells.append(html.Td(v, style=sty))
        body_rows.append(html.Tr(net_dol_cells))

    return dbc.Table(
        [thead, html.Tbody(body_rows)],
        bordered=True, hover=True, size="sm",
        className="table-responsive",
        style={"width": "100%", "margin": "0 auto",
               "pageBreakInside": "avoid"},
    )


def _fee_slab_structure_rows() -> list:
    """
    Brief summary, slab table, and a single footnote row pointing to the Disclosure Document.
    """
    explain = html.Tr([
        html.Td(
            html.Small(
                "Net new profits above the High-Water Mark are measured against a Benchmark dollar amount "
                "for the month: the S&P 500's monthly percentage return times the Program's nominal trading level. "
                "When that Benchmark return is positive, the incentive fee is graduated — each slab rate applies "
                "only to the profits falling in that slice (0–1× through >4× of the Benchmark dollar amount; see table). "
                "When the Benchmark return is zero or negative, the incentive fee is 50% of qualifying net new profits; "
                "these slabs do not apply.",
                className="text-muted fst-italic",
            ),
            colSpan=2,
        ),
    ])
    th_style = {"fontSize": "0.78rem", "verticalAlign": "bottom"}
    thead = html.Thead(html.Tr([
        html.Th("Slab", className=HEADER_ROW_CLASS, style=th_style),
        html.Th(
            "Net new profits slice (vs Benchmark dollar return for the month)",
            className=HEADER_ROW_CLASS,
            style=th_style,
        ),
        html.Th(
            "Incentive fee on profits in that slice",
            className=HEADER_ROW_CLASS,
            style=th_style,
        ),
    ]))
    tbody = html.Tbody([
        html.Tr([
            html.Td(slab, style={"fontWeight": "600", "whiteSpace": "nowrap"}),
            html.Td(band),
            html.Td(rate, style={"whiteSpace": "nowrap"}),
        ])
        for slab, band, rate in FEE_SLAB_BANDS_DISPLAY
    ])
    nested = dbc.Table(
        [thead, tbody],
        bordered=True,
        hover=True,
        size="sm",
        className="mb-0",
    )
    table_row = html.Tr([
        html.Td(
            nested,
            colSpan=2,
            style={"padding": "6px 0", "borderTop": "none"},
        ),
    ])
    doc_row = html.Tr([
        html.Td(
            html.Small(
                "Definitions, worked examples, and rounding: see the AlgoMinds Financial LLC Disclosure Document.",
                className="text-muted fst-italic",
            ),
            colSpan=2,
        ),
    ])
    return [explain, table_row, doc_row]


# ==============================================================================
# DASH APP  –  structure mirrors Y&Q serve_layout()
# ==============================================================================
# Serve the SHARED repo-root assets (the same styles.css TKP/TCP load) so the
# Important Notice gate renders with the exact sibling design system — modal
# backdrop, centered rounded card, typography scale, navy accept button.
# mp_ts.py lives in "Momentum Pacer/" which has no assets folder of its own;
# without assets_folder the gate-card CSS classes would 404 and never load.
app = dash.Dash(
    __name__,
    assets_folder=str(_TS_ROOT / "assets"),
    external_stylesheets=[dbc.themes.BOOTSTRAP],
    suppress_callback_exceptions=True,
    title="Algominds – Momentum Pacer",
)

agm_admin_auth_manager = AdminAuthManager(load_agm_admin_auth_settings(), session_key=AGM_SESSION_KEY)
configure_flask_session_secret(app.server, agm_admin_auth_manager.settings)


def serve_layout():
    agm_latest_date = (
        daily_balances_df["Date"].max()
        if daily_balances_df is not None and not daily_balances_df.empty
        else None
    )
    desktop_date_label, mobile_date_label = build_header_date_label_children_from_date(
        agm_latest_date
    )

    inception_str = PROGRAM_INCEPTION.strftime("%B %d, %Y")
    latest_str = (
        _display_summary_df["date"].max().strftime("%B %Y")
        if not _display_summary_df.empty
        else PROGRAM_INCEPTION.strftime("%B %Y")
    )
    _recon_min_ts, _recon_max_ts = _recon_date_bounds()
    # DatePickerSingle expects ISO date strings (not raw Timestamp objects)
    # for correct browser-side serialization.
    _recon_min = _recon_min_ts.strftime("%Y-%m-%d") if _recon_min_ts is not None else None
    _recon_max = _recon_max_ts.strftime("%Y-%m-%d") if _recon_max_ts is not None else None

    return dbc.Container(
        id="page-container",
        fluid=True,
        className="py-4",
        style={"maxWidth": "1400px"},
        children=[

            dcc.Store(id="access-mode", storage_type="session", data=None),
            dcc.Store(id=GATE_PASSWORD_VISIBLE_STORE_ID, storage_type="memory", data=False),
            dcc.Location(id="url", refresh=False),

            # Accept gate — MANAGER tier: Algominds Financial LLC / Momentum Pacer (port 8304)
            build_manager_accept_gate(
                "Momentum Pacer",
                extra_children=[build_gate_password_row()],
            ),

            # ── Main content ────────────────────────────────────────────────────
            html.Div(
                id="main-app",
                style={"display": "none"},
                children=[

                    # ── Header (TCP-style “Data current to” block) ─────────────
                    *build_tearsheet_header_row(
                        logo_src=logo_src,
                        logo_alt="Algominds Financial LLC Logo",
                        firm_name="Algominds Financial LLC",
                        product_name="Momentum Pacer Program",
                        desktop_label_children=desktop_date_label,
                        mobile_label_children=mobile_date_label,
                        grey_bg=GREY_BG,
                    ),

                    # ── Description ───────────────────────────────────────────
                    html.Div(
                        [
                            html.P(
                                "Algominds Financial LLC is a Commodity Trading Advisor "
                                "operating the Momentum Pacer Trading Program.",
                                className="lead text-center",
                            ),
                            html.P(
                                "Advisor: Algominds Financial LLC | Program: Momentum Pacer | "
                                f"Live Inception: {inception_str} | "
                                "Products Traded: Nasdaq-100 Futures (NQ / MNQ) | "
                                "Style: Systematic Trend Following",
                                className="text-center mb-5",
                            ),
                        ],
                        className="description",
                    ),

                    # ── Error banner ───────────────────────────────────────────
                    dbc.Alert(f"Data error: {LOAD_ERROR}", color="danger",
                              className="mb-3") if LOAD_ERROR else html.Div(),

                    # ── NAV Chart (wrapper avoids caption / x-axis label overlap)
                    html.Div(
                        [
                            dcc.Graph(
                                id="mp-nav-graph",
                                figure=build_nav_figure(),
                                config={"displayModeBar": False, "responsive": True},
                                style={
                                    "width": "100%",
                                    "minHeight": "440px",
                                    "marginBottom": "0",
                                    "pageBreakInside": "avoid",
                                },
                            ),
                        ],
                        style={
                            "marginBottom": "0.25rem",
                            "paddingBottom": "0.5rem",
                            "overflow": "visible",
                        },
                    ),
                    html.P(
                        f"Growth of a ${STARTING_CAPITAL:,.0f} investment from inception ({inception_str}), "
                        "shown at DAILY resolution (one point per trading day). NAV is client net value "
                        "(actual NLV net of accrued unpaid fees). "
                        "The strategy trades NQ / MNQ (Nasdaq-100 futures) exclusively. "
                        "S&P 500 and Nasdaq-100 daily index closes are rebased to the same starting "
                        "capital at inception for benchmark comparison only.",
                        className="text-center small text-muted fst-italic px-3",
                        style={
                            "marginTop": "1.5rem",
                            "marginBottom": "2.25rem",
                            "paddingTop": "0.75rem",
                            "lineHeight": "1.45",
                            "maxWidth": "920px",
                            "marginLeft": "auto",
                            "marginRight": "auto",
                        },
                    ),

                    # ── Admin-only: DAILY TradeStation balances (TearSheet mode only, AGM only) ──
                    # Raw NLV / Net Worth values from the daily CSV. NEVER client-facing.
                    # The sensitive content (table/graph/KPIs) is rendered into
                    # `agm-admin-daily-content` by a SERVER-SIDE, auth-gated callback so
                    # the raw NLV values are never shipped to a non-admin browser at all
                    # (not merely hidden via CSS).
                    html.Div(
                        id="agm-admin-daily-container",
                        style={"display": "none"},
                        className="mt-2 mb-4",
                        children=[
                            dbc.Alert(
                                "Admin — daily TradeStation balances (raw account NLV; admin-only, not shown to clients)",
                                color="warning",
                                className="text-center fw-bold mb-3",
                            ),
                            html.Div(id="agm-admin-daily-content"),
                        ],
                    ),

                    # ── Admin-only: Accrued Fees + NLV (TearSheet mode only, AGM only) ──
                    html.Div(
                        id="agm-admin-fee-charts-container",
                        style={"display": "none"},
                        className="mt-2 mb-4",
                        children=[
                            dbc.Alert(
                                "Admin — fee accrual detail (calculated daily from AGM vs SPX; "
                                "not shown on the client-facing tearsheet)",
                                color="warning",
                                className="text-center fw-bold mb-3",
                            ),
                            dcc.Graph(
                                id="agm-accrued-fees-graph",
                                figure=build_agm_accrued_fees_figure(),
                                config={"displayModeBar": False, "responsive": True},
                                style={"width": "100%", "minHeight": "340px"},
                            ),

                            # ── Admin reconciliation panel (date-based spot-check) ──
                            html.Hr(className="my-3"),
                            html.H6("Reconciliation Check (Admin)", className="mb-2"),
                            html.P(
                                "Spot-check any date: TradeStation NLV / Actual NLV always "
                                "equals Client Net Economic Value plus Accrued Unpaid "
                                "Incentive Fee.",
                                className="small text-muted mb-2",
                            ),
                            dcc.DatePickerSingle(
                                id=AGM_RECON_DATE_PICKER_ID,
                                min_date_allowed=_recon_min,
                                max_date_allowed=_recon_max,
                                initial_visible_month=_recon_max,
                                date=_recon_max,
                                display_format="YYYY-MM-DD",
                            ),
                            html.Div(id=AGM_RECON_OUTPUT_ID, className="mt-3"),
                        ],
                    ),

                    # ── Performance Summary Table ──────────────────────────────
                    html.H5(
                        "Performance Summary",
                        className="text-center mb-2",
                        style={"marginTop": "0.5rem", "paddingTop": "0.25rem"},
                    ),
                    html.Div(
                        build_performance_summary_table(),
                        className="table-responsive mb-5",
                        style={"overflowX": "auto",
                               "pageBreakInside": "avoid"},
                    ),

                    # ── Strategy Overview  +  Risk & Controls  ─────────────────
                    # (same 2-column card layout as Y&Q)
                    dbc.Row(
                        [
                            # LEFT: Strategy Overview
                            dbc.Col(
                                dbc.Card(
                                    [
                                        dbc.CardHeader(
                                            html.H6("Strategy Overview", className="mb-0"),
                                            className=HEADER_ROW_CLASS,
                                        ),
                                        dbc.CardBody(
                                            dbc.Table(
                                                [
                                                    html.Thead(
                                                        html.Tr([html.Th("Strategy Description",
                                                                          colSpan=3,
                                                                          className=HEADER_ROW_CLASS)]),
                                                        className="bg-light",
                                                    ),
                                                    html.Tbody([
                                                        html.Tr([
                                                            html.Td(
                                                                [
                                                                    html.P(
                                                                        "The Momentum Pacer Program is a systematic "
                                                                        "momentum strategy trading exclusively in "
                                                                        "Nasdaq-100 futures contracts, including NQ "
                                                                        "and MNQ. The program uses quantitative "
                                                                        "momentum signals to identify and capture "
                                                                        "directional moves in the Nasdaq-100, with "
                                                                        "risk managed through adaptive position "
                                                                        "sizing and stop-loss orders.",
                                                                        className="mb-3",
                                                                    ),
                                                                    html.P(
                                                                        "S&P 500 and Nasdaq-100 index levels are shown "
                                                                        "as benchmarks for comparison only. The "
                                                                        "strategy trades Nasdaq-100 futures and does "
                                                                        "not trade the S&P 500 cash index or related "
                                                                        "products for alpha. Per the Disclosure "
                                                                        "Document, the contractual monthly incentive "
                                                                        "fee uses the S&P 500 monthly return as the "
                                                                        "Benchmark reference, as described in the "
                                                                        "Fee Slab Structure, and is applied to net "
                                                                        "new trading profits subject to a high-water "
                                                                        "mark.",
                                                                    ),
                                                                ],
                                                                colSpan=3,
                                                                style={"whiteSpace": "normal",
                                                                       "fontStyle": "italic"},
                                                            )
                                                        ]),
                                                        html.Tr([html.Td("", colSpan=3,
                                                                          style={"height": LEFT_TABLE_GAPS})]),
                                                        html.Tr([
                                                            html.Th("Methodology", colSpan=3,
                                                                     className="bg-light"),
                                                        ]),
                                                        html.Tr([
                                                            html.Td("Trading Style"),
                                                            html.Td(html.Span("✓ Momentum / Trend",
                                                                               style={"color": ACCENT_GREEN})),
                                                            html.Td(html.Span("✗ Mean Reversion",
                                                                               style={"color": SECONDARY_COLOR})),
                                                        ]),
                                                        html.Tr([
                                                            html.Td("Decision Making"),
                                                            html.Td(html.Span("✓ Systematic",
                                                                               style={"color": ACCENT_GREEN})),
                                                            html.Td(html.Span("✗ Discretionary",
                                                                               style={"color": SECONDARY_COLOR})),
                                                        ]),
                                                        html.Tr([
                                                            html.Td("Execution"),
                                                            html.Td(html.Span("✓ Fully automated*",
                                                                               style={"color": ACCENT_GREEN})),
                                                            html.Td(html.Span("✗ Manual",
                                                                               style={"color": SECONDARY_COLOR})),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(
                                                                html.Small(
                                                                    "*Order generation and routing are fully systematic "
                                                                    "and automated; fills remain subject to exchange, "
                                                                    "FCM, and market conditions.",
                                                                    className="text-muted fst-italic",
                                                                ),
                                                                colSpan=3,
                                                            ),
                                                        ]),
                                                        html.Tr([
                                                            html.Td("Instruments"),
                                                            html.Td("NQ / MNQ (Nasdaq-100 E-mini & Micro)", colSpan=2),
                                                        ]),
                                                        html.Tr([
                                                            html.Td("Exchanges"),
                                                            html.Td("CME Group", colSpan=2),
                                                        ]),
                                                        html.Tr([
                                                            html.Td("Account Start Date"),
                                                            html.Td(inception_str, colSpan=2),
                                                        ]),
                                                        html.Tr([
                                                            html.Td("Initial Capital"),
                                                            html.Td(f"${STARTING_CAPITAL:,.0f}", colSpan=2),
                                                        ]),
                                                        html.Tr([html.Td("", colSpan=3,
                                                                          style={"height": LEFT_TABLE_GAPS})]),
                                                        html.Tr([
                                                            html.Th("Risk Controls", colSpan=3,
                                                                     className="bg-light"),
                                                        ]),
                                                        html.Tr([
                                                            html.Td(html.Span("✓ Stop Losses",
                                                                               style={"color": ACCENT_GREEN})),
                                                            html.Td(html.Span("✓ Position Sizing",
                                                                               style={"color": ACCENT_GREEN})),
                                                            html.Td(html.Span("✓ High-Water Mark",
                                                                               style={"color": ACCENT_GREEN})),
                                                        ]),
                                                    ]),
                                                ],
                                                striped=False, bordered=True,
                                                hover=True, size="sm",
                                            )
                                        ),
                                    ],
                                    outline=True, className="mb-4",
                                ),
                                width=6,
                            ),

                            # RIGHT: Fee Structure  (mirrors Y&Q's Transaction Fees card)
                            dbc.Col(
                                dbc.Card(
                                    [
                                        dbc.CardHeader(
                                            html.H6("Fee Structure", className="mb-0"),
                                            className=HEADER_ROW_CLASS,
                                        ),
                                        dbc.CardBody(
                                            dbc.Table(
                                                [
                                                    html.Thead(
                                                        html.Tr([html.Th("Terms & Fees",
                                                                          colSpan=2,
                                                                          className=HEADER_ROW_CLASS)])
                                                    ),
                                                    html.Tbody([
                                                        html.Tr([html.Td("Management Fee"),
                                                                 html.Td("None")]),
                                                        html.Tr([html.Td("Performance Fee"),
                                                                 html.Td(
                                                                     "Monthly incentive fee on net new trading profits "
                                                                     "(subject to High-Water Mark), determined by reference "
                                                                     "to the S&P 500 monthly return (the Benchmark) per the "
                                                                     "AlgoMinds Disclosure Document — graduated slabs when "
                                                                     "the Benchmark is positive; different rule when the "
                                                                     "Benchmark is zero or negative (see Fee Slab Structure)."
                                                                 )]),
                                                        html.Tr([html.Td("High-Water Mark"),
                                                                 html.Td("Yes — fee only on new net gains above prior HWM")]),
                                                        html.Tr([html.Td("Fee Frequency"),
                                                                 html.Td("Monthly")]),
                                                        html.Tr([html.Td(""), html.Td("")]),
                                                        html.Tr([
                                                            html.Th("Fee Slab Structure",
                                                                     colSpan=2,
                                                                     className="bg-light"),
                                                        ]),
                                                        *_fee_slab_structure_rows(),
                                                        html.Tr([html.Td(""), html.Td("")]),
                                                        html.Tr([
                                                            html.Th("Account Information",
                                                                     colSpan=2,
                                                                     className="bg-light"),
                                                        ]),
                                                        html.Tr([html.Td("Minimum Investment"),
                                                                 html.Td("$30,000")]),
                                                        html.Tr([html.Td("Notional Funding"),
                                                                 html.Td("Available")]),
                                                        html.Tr([html.Td("Broker / FCM"),
                                                                 html.Td("TradeStation")]),
                                                        html.Tr([html.Td("Liquidity"),
                                                                 html.Td("Withdrawals with 7 days' notice")]),
                                                        html.Tr([html.Td("Lockup Period"),
                                                                 html.Td("None")]),
                                                    ]),
                                                ],
                                                striped=False, bordered=True,
                                                hover=True, size="sm",
                                                className="table-responsive",
                                            )
                                        ),
                                    ],
                                    outline=True, className="mb-4",
                                ),
                                width=6,
                            ),
                        ],
                        justify="start",
                        className="mb-2",
                    ),

                    # ── Metrics row  (mirrors Y&Q: perf metrics left, stats right)
                    dbc.Row(
                        [
                            # LEFT: Performance Metrics
                            dbc.Col(
                                [
                                    dbc.Card(
                                        [
                                            dbc.CardHeader(
                                                html.H6("Performance Metrics", className="mb-0")
                                            ),
                                            dbc.CardBody(
                                                dbc.Table.from_dataframe(
                                                    perf_df,
                                                    striped=False, bordered=True,
                                                    hover=True, size="sm",
                                                    className="fixed-cols",
                                                ) if not perf_df.empty
                                                else html.P("No data.")
                                            ),
                                        ],
                                        outline=True, className="mb-2",
                                    ),
                                ],
                                width=6,
                            ),
                            # RIGHT: Monthly Performance Statistics
                            dbc.Col(
                                dbc.Card(
                                    [
                                        dbc.CardHeader(
                                            html.H6("Monthly Performance Statistics",
                                                    className="mb-0")
                                        ),
                                        dbc.CardBody(
                                            dbc.Table.from_dataframe(
                                                stats_df,
                                                striped=False, bordered=True,
                                                hover=True, size="sm",
                                                className="fixed-cols",
                                            ) if not stats_df.empty
                                            else html.P("No data.")
                                        ),
                                        dbc.CardFooter(
                                            html.Small(
                                                f"Statistics calculated from actual monthly return data "
                                                f"from {inception_str} to {latest_str}.",
                                                className="text-muted fst-italic",
                                            )
                                        ),
                                    ],
                                    outline=True, className="mb-2",
                                ),
                                width=6,
                            ),
                        ],
                        justify="start",
                        className="mb-2",
                    ),

                    # ── Investor Information  (full width, mirrors Y&Q) ─────────
                    dbc.Row(
                        [
                            dbc.Col(
                                dbc.Card(
                                    [
                                        dbc.CardHeader(
                                            html.H6("Investor Information", className="mb-0")
                                        ),
                                        dbc.CardBody(
                                            dbc.Row([
                                                dbc.Col(
                                                    dbc.Table(
                                                        [
                                                            html.Thead(html.Tr([
                                                                html.Th("Terms & Fees"),
                                                                html.Th("Details"),
                                                            ])),
                                                            html.Tbody([
                                                                html.Tr([html.Td("Performance Fee"),
                                                                         html.Td(
                                                                             "Graduated vs S&P 500 Benchmark (Disclosure Document), "
                                                                             "monthly, HWM"
                                                                         )]),
                                                                html.Tr([html.Td("High Water Mark"),
                                                                         html.Td("Yes")]),
                                                                html.Tr([html.Td("Lockup Period"),
                                                                         html.Td("None")]),
                                                                html.Tr([html.Td("Liquidity"),
                                                                         html.Td("Withdrawals with 7 days' notice")]),
                                                                html.Tr([html.Td("Minimum Investment"),
                                                                         html.Td("$30,000")]),
                                                                html.Tr([html.Td("Notional Funding"),
                                                                         html.Td("Available")]),
                                                                html.Tr([html.Td("Execution FCM"),
                                                                         html.Td("TradeStation")]),
                                                            ]),
                                                        ],
                                                        striped=False, bordered=True,
                                                        hover=True, size="sm",
                                                    ),
                                                    width=6,
                                                ),
                                                dbc.Col(
                                                    dbc.Table(
                                                        [
                                                            html.Thead(html.Tr([
                                                                html.Th("Account Stats"),
                                                                html.Th("Current"),
                                                            ])),
                                                            html.Tbody([
                                                                html.Tr([
                                                                    html.Td("Starting Capital"),
                                                                    html.Td(f"${STARTING_CAPITAL:,.0f}"),
                                                                ]),
                                                                html.Tr([
                                                                    html.Td("Current NAV (after fees)"),
                                                                    html.Td(
                                                                        f"${_display_summary_df['bot_end_after_fees'].iloc[-1]:,.2f}"
                                                                        if not _display_summary_df.empty else "—"
                                                                    ),
                                                                ]),
                                                                html.Tr([
                                                                    html.Td("Total Net Gain"),
                                                                    html.Td(
                                                                        f"${net_totals.get('bot_net_dollar', 0):,.2f}"
                                                                        if net_totals else "—"
                                                                    ),
                                                                ]),
                                                                html.Tr([
                                                                    html.Td("Total Fees Paid"),
                                                                    html.Td(
                                                                        f"${net_totals.get('bot_fees_dollar', 0):,.2f}"
                                                                        if net_totals else "—"
                                                                    ),
                                                                ]),
                                                                html.Tr([
                                                                    html.Td("Inception Date"),
                                                                    html.Td(inception_str),
                                                                ]),
                                                                html.Tr([
                                                                    html.Td("Months trading (approx.)"),
                                                                    html.Td(months_trading_elapsed_approx()),
                                                                ]),
                                                            ]),
                                                        ],
                                                        striped=False, bordered=True,
                                                        hover=True, size="sm",
                                                    ),
                                                    width=6,
                                                ),
                                            ])
                                        ),
                                    ],
                                    outline=True, className="mb-4",
                                ),
                                width=12,
                            ),
                        ],
                        className="mb-2",
                    ),

                    # ── Drawdown chart (wrapper keeps SVG from overlapping footer)
                    html.Div(
                        [
                            dcc.Graph(
                                id="drawdown-graph",
                                figure=build_drawdown_figure(),
                                config={"displayModeBar": False, "responsive": True},
                                style={
                                    "width": "100%",
                                    "minHeight": "360px",
                                    "marginBottom": "0",
                                },
                            ),
                        ],
                        className="mb-0",
                        style={
                            "marginBottom": "2.5rem",
                            "paddingBottom": "1.75rem",
                            "overflow": "visible",
                        },
                    ),

                    # ── Client-facing daily table (collapsed by default) ────────
                    build_client_daily_table_section(),

                    html.Hr(className="my-4"),

                    # ── Footer / disclaimers  (identical to Y&Q) ───────────────
                    html.Div(
                        [
                            html.P(
                                "THE MOMENTUM PACER PROGRAM IS A PROPRIETARY TRADING STRATEGY. "
                                "THIS PERFORMANCE DATA IS FOR INFORMATIONAL PURPOSES ONLY AND IS "
                                "NOT A SOLICITATION TO INVEST. "
                                "PAST PERFORMANCE IS NOT INDICATIVE OF FUTURE RESULTS.",
                                className="text-center small fw-bold",
                            ),
                            html.P(
                                "Past performance is not necessarily indicative of future results. "
                                "The risk of loss in commodity trading can be substantial. "
                                "This information is for informational purposes only and does not "
                                "constitute investment advice or a solicitation to invest.",
                                className="text-center small text-muted fst-italic",
                            ),
                            html.P(
                                "For more information, contact Algominds Financial LLC",
                                className="text-center small text-muted mb-0",
                            ),
                        ],
                        className="pt-3 pb-3 mb-0",
                        style={"marginTop": "0.5rem"},
                    ),

                    # Important Disclosure section (manager tier — bottom panel)
                    dbc.Row(
                        dbc.Col(
                            html.Div(
                                id="important-disclosure",
                                children=tsd.manager_bottom_disclosure_children(),
                                className=tsd.DISCLOSURE_PANEL_CLASS,
                                style=tsd.DISCLOSURE_PANEL_STYLE,
                            ),
                            width=12,
                        ),
                        className="mb-5 pb-4",
                    ),

                ],  # end main-app children
            ),  # end main-app Div
        ],
    )


app.layout = serve_layout


@app.callback(
    Output("disclaimer-screen", "style"),
    Output("main-app", "style"),
    Output("access-mode", "data"),
    Input("accept-button", "n_clicks"),
)
def show_main(n_clicks):
    if n_clicks and n_clicks > 0:
        return {"display": "none"}, {"display": "block"}, "standard"
    return tsd.GATE_SCREEN_STYLE, {"display": "none"}, None


# ── Hidden admin reveal: "e" click opens the password row (no access granted yet) ──
@app.callback(
    Output(GATE_PASSWORD_VISIBLE_STORE_ID, "data"),
    Input("secret-notice-e", "n_clicks"),
    State(GATE_PASSWORD_VISIBLE_STORE_ID, "data"),
    prevent_initial_call=True,
)
def _toggle_gate_password_row(n_clicks, visible):
    if n_clicks:
        return not bool(visible)
    return dash.no_update


@app.callback(
    Output(GATE_PASSWORD_ROW_ID, "style"),
    Input(GATE_PASSWORD_VISIBLE_STORE_ID, "data"),
)
def _render_gate_password_row(visible):
    return gate_password_row_style(bool(visible))


@app.callback(
    Output(GATE_PASSWORD_INPUT_ID, "value", allow_duplicate=True),
    Input(GATE_PASSWORD_VISIBLE_STORE_ID, "data"),
    prevent_initial_call=True,
)
def _clear_password_when_hidden(visible):
    if not visible:
        return ""
    return dash.no_update


@app.callback(
    Output(GATE_PASSWORD_ERROR_ID, "children"),
    Output(GATE_PASSWORD_VISIBLE_STORE_ID, "data", allow_duplicate=True),
    Output(GATE_PASSWORD_INPUT_ID, "value", allow_duplicate=True),
    Output("disclaimer-screen", "style", allow_duplicate=True),
    Output("main-app", "style", allow_duplicate=True),
    Output("access-mode", "data", allow_duplicate=True),
    Input(GATE_PASSWORD_SUBMIT_ID, "n_clicks"),
    Input(GATE_PASSWORD_INPUT_ID, "n_submit"),
    State(GATE_PASSWORD_INPUT_ID, "value"),
    prevent_initial_call=True,
)
def _gate_admin_tearsheet_login(_submit_clicks, _n_submit, password):
    ok, _msg = agm_admin_auth_manager.login(session, password or "")
    if not ok:
        return INVALID_PASSWORD_MESSAGE, dash.no_update, "", dash.no_update, dash.no_update, dash.no_update
    return "", False, "", {"display": "none"}, {"display": "block"}, "secret"


@app.callback(
    Output(GATE_PASSWORD_ERROR_ID, "children", allow_duplicate=True),
    Output(GATE_PASSWORD_VISIBLE_STORE_ID, "data", allow_duplicate=True),
    Output(GATE_PASSWORD_INPUT_ID, "value", allow_duplicate=True),
    Output("url", "href"),
    Output("url", "refresh"),
    Input(GATE_PASSWORD_PORTAL_ID, "n_clicks"),
    State(GATE_PASSWORD_INPUT_ID, "value"),
    prevent_initial_call=True,
)
def _gate_admin_portal_login(_portal_clicks, password):
    ok, _msg = agm_admin_auth_manager.login(session, password or "")
    if not ok:
        return INVALID_PASSWORD_MESSAGE, dash.no_update, "", dash.no_update, dash.no_update
    return "", False, "", ADMIN_PORTAL_PATH, True


@app.callback(
    Output("agm-admin-fee-charts-container", "style"),
    Output("agm-admin-daily-container", "style"),
    Input("access-mode", "data"),
)
def _toggle_admin_sections(access_mode):
    style = {"display": "block"} if access_mode == "secret" else {"display": "none"}
    return style, style


def _render_admin_daily_content(access_mode):
    """Render admin KPI cards + raw NLV graph ONLY when authenticated."""
    if access_mode != "secret" or not agm_admin_auth_manager.is_authenticated(session):
        return []
    return [
        build_agm_daily_kpi_cards(),
        dcc.Graph(
            id="agm-daily-nlv-graph",
            figure=build_agm_daily_nlv_figure(),
            config={"displayModeBar": False, "responsive": True},
            style={"width": "100%", "minHeight": "340px", "marginBottom": "1rem"},
        ),
    ]


@app.callback(
    Output("agm-admin-daily-content", "children"),
    Input("access-mode", "data"),
)
def _render_admin_daily_content_callback(access_mode):
    return _render_admin_daily_content(access_mode)


def _render_agm_reconciliation_panel(selected_date, access_mode):
    """Admin-only reconciliation readout for *selected_date* — verified
    server-side against a genuine authenticated session (a spoofed
    client-side access-mode store yields nothing)."""
    if access_mode != "secret" or not agm_admin_auth_manager.is_authenticated(session):
        return []
    return build_agm_reconciliation_panel(selected_date)


@app.callback(
    Output(AGM_RECON_OUTPUT_ID, "children"),
    Input(AGM_RECON_DATE_PICKER_ID, "date"),
    Input("access-mode", "data"),
)
def _render_agm_reconciliation_panel_callback(selected_date, access_mode):
    return _render_agm_reconciliation_panel(selected_date, access_mode)


# ── Client-facing daily table: toggle / page size / export ─────────────────
@app.callback(
    Output(CLIENT_DAILY_COLLAPSE_ID, "is_open"),
    Output(CLIENT_DAILY_TOGGLE_ID, "children"),
    Input(CLIENT_DAILY_TOGGLE_ID, "n_clicks"),
    State(CLIENT_DAILY_COLLAPSE_ID, "is_open"),
    prevent_initial_call=True,
)
def _toggle_client_daily_table(n_clicks, is_open):
    new_open = not is_open
    label = "Hide ▴" if new_open else "Show ▾"
    return new_open, label


@app.callback(
    Output(CLIENT_DAILY_TABLE_ID, "page_size"),
    Input(CLIENT_DAILY_PAGE_SIZE_ID, "value"),
)
def _update_client_daily_page_size(page_size):
    return page_size or 25


@app.callback(
    Output(CLIENT_DAILY_EXPORT_DOWNLOAD_ID, "data"),
    Input(CLIENT_DAILY_EXPORT_BTN_ID, "n_clicks"),
    State(CLIENT_DAILY_TABLE_ID, "data"),
    prevent_initial_call=True,
)
def _export_client_daily_excel(n_clicks, table_data):
    if not n_clicks or not table_data:
        return dash.no_update
    export_df = pd.DataFrame(table_data)
    return dcc.send_data_frame(export_df.to_excel, "agm_daily_performance.xlsx", index=False)


# ── Admin-only Daily Returns controls (auth-gated; TCP admin-toolbar pattern) ──

def _render_agm_daily_admin_controls(access_mode):
    """Toolbar + modals for the Daily Returns card, plus the table data with
    any admin-entered manual rows — rendered ONLY for a genuinely
    authenticated admin session (a spoofed access-mode store yields nothing)."""
    if access_mode != "secret" or not agm_admin_auth_manager.is_authenticated(session):
        return [], dash.no_update
    manual = _load_agm_manual_daily_rows()
    data = (
        build_client_daily_table_rows(table=_compute_accounting_with_manual_rows(manual))
        if manual
        else dash.no_update
    )
    return build_agm_daily_admin_controls(), data


@app.callback(
    Output(AGM_DAILY_ADMIN_SLOT_ID, "children"),
    Output(CLIENT_DAILY_TABLE_ID, "data"),
    Input("access-mode", "data"),
)
def _render_agm_daily_admin_controls_callback(access_mode):
    return _render_agm_daily_admin_controls(access_mode)


@app.callback(
    Output(CLIENT_DAILY_TABLE_ID, "columns"),
    Input(AGM_DAILY_ADMIN_COL_PICKER_ID, "value"),
    prevent_initial_call=True,
)
def _update_agm_daily_visible_columns(selected):
    if not agm_admin_auth_manager.is_authenticated(session):
        return dash.no_update
    labels = [label for label, _, _ in CLIENT_DAILY_TABLE_COLUMNS]
    if not selected:
        selected = labels
    spec = [c for c in CLIENT_DAILY_TABLE_COLUMNS if c[0] in set(selected)]
    return _daily_table_column_defs(spec)


@app.callback(
    Output(AGM_DAILY_ADMIN_ADD_MODAL_ID, "is_open"),
    Input(AGM_DAILY_ADMIN_ADD_BTN_ID, "n_clicks"),
    Input(AGM_DAILY_ADMIN_ADD_CANCEL_ID, "n_clicks"),
    prevent_initial_call=True,
)
def _toggle_agm_daily_add_modal(_open_clicks, _cancel_clicks):
    triggered = dash.callback_context.triggered[0]["prop_id"].split(".")[0]
    return triggered == AGM_DAILY_ADMIN_ADD_BTN_ID


@app.callback(
    Output(AGM_DAILY_ADMIN_ADD_MODAL_ID, "is_open", allow_duplicate=True),
    Output(AGM_DAILY_ADMIN_ADD_ERROR_ID, "children"),
    Output(CLIENT_DAILY_TABLE_ID, "data", allow_duplicate=True),
    Input(AGM_DAILY_ADMIN_ADD_SAVE_ID, "n_clicks"),
    State(AGM_DAILY_ADMIN_ADD_DATE_ID, "value"),
    State(AGM_DAILY_ADMIN_ADD_NLV_ID, "value"),
    State(AGM_DAILY_ADMIN_ADD_DEPOSIT_ID, "value"),
    State(AGM_DAILY_ADMIN_ADD_FEE_PAID_ID, "value"),
    prevent_initial_call=True,
)
def _agm_daily_add_row_save(n_clicks, date_val, nlv_val, deposit_val, fee_paid_val):
    if not n_clicks:
        return dash.no_update, dash.no_update, dash.no_update
    if not agm_admin_auth_manager.is_authenticated(session):
        return dash.no_update, "Not authenticated.", dash.no_update
    ok, message, table = agm_add_manual_daily_row(date_val, nlv_val, deposit_val, fee_paid_val)
    if not ok:
        return dash.no_update, message, dash.no_update
    return False, "", build_client_daily_table_rows(table=table)


@app.callback(
    Output(AGM_DAILY_ADMIN_DELETE_MODAL_ID, "is_open"),
    Output(AGM_DAILY_ADMIN_DELETE_BODY_ID, "children"),
    Input(AGM_DAILY_ADMIN_DELETE_BTN_ID, "n_clicks"),
    Input(AGM_DAILY_ADMIN_DELETE_CANCEL_ID, "n_clicks"),
    prevent_initial_call=True,
)
def _toggle_agm_daily_delete_modal(_open_clicks, _cancel_clicks):
    triggered = dash.callback_context.triggered[0]["prop_id"].split(".")[0]
    if triggered != AGM_DAILY_ADMIN_DELETE_BTN_ID:
        return False, dash.no_update
    manual = sorted(_load_agm_manual_daily_rows(), key=lambda r: r["date"])
    if not manual:
        body = (
            "No manually added daily rows to delete — TradeStation CSV rows "
            "are never deleted from the tearsheet."
        )
    else:
        last = manual[-1]
        body = (
            f"Delete manually added row for {last['date']} "
            f"(TradeStation NLV ${float(last['actual_nlv']):,.2f})? "
            f"This cannot be undone. TradeStation CSV rows are never deleted."
        )
    return True, body


@app.callback(
    Output(AGM_DAILY_ADMIN_DELETE_MODAL_ID, "is_open", allow_duplicate=True),
    Output(AGM_DAILY_ADMIN_DELETE_BODY_ID, "children", allow_duplicate=True),
    Output(CLIENT_DAILY_TABLE_ID, "data", allow_duplicate=True),
    Input(AGM_DAILY_ADMIN_DELETE_CONFIRM_ID, "n_clicks"),
    prevent_initial_call=True,
)
def _agm_daily_delete_last_row(n_clicks):
    if not n_clicks:
        return dash.no_update, dash.no_update, dash.no_update
    if not agm_admin_auth_manager.is_authenticated(session):
        return dash.no_update, "Not authenticated.", dash.no_update
    ok, message, table = agm_delete_last_manual_daily_row()
    if not ok:
        return True, message, dash.no_update
    return False, dash.no_update, build_client_daily_table_rows(table=table)


@app.callback(
    Output(AGM_DAILY_ADMIN_CALC_MODAL_ID, "is_open"),
    Input(AGM_DAILY_ADMIN_CALC_BTN_ID, "n_clicks"),
    Input(AGM_DAILY_ADMIN_CALC_CLOSE_ID, "n_clicks"),
    prevent_initial_call=True,
)
def _toggle_agm_daily_calc_modal(_open_clicks, _close_clicks):
    triggered = dash.callback_context.triggered[0]["prop_id"].split(".")[0]
    return triggered == AGM_DAILY_ADMIN_CALC_BTN_ID


@app.server.route("/admin")
def agm_admin_portal():
    if not agm_admin_auth_manager.is_authenticated(session):
        return redirect("/")
    latest_row = _display_summary_df.iloc[-1] if not _display_summary_df.empty else None
    since_inception_display = perf_metrics.get("Cumulative Net Return") if perf_metrics else None

    # Portal is admin/accounting-only, so the "current account value" here is the
    # real latest daily TradeStation Net Worth (raw NLV) when available, falling
    # back to the monthly workbook after-fee value only if the daily CSV is missing.
    daily_row = agm_daily.latest_row(daily_balances_df)
    if daily_row is not None:
        after_fee_nlv = float(daily_row["Net Worth"])
        last_updated = daily_row["Date"].strftime("%Y-%m-%d")
        since_inception_display = f"{float(daily_row['since_inception_pct']):.2f}%"
    else:
        after_fee_nlv = float(latest_row["bot_end_after_fees"]) if latest_row is not None else None
        last_updated = LATEST_DATE.strftime("%Y-%m-%d") if LATEST_DATE is not None else "—"

    accounts = agm_registry.build_participating_accounts(
        program_name="Momentum Pacer",
        inception_date=PROGRAM_INCEPTION,
        benchmark_base="S&P 500 (SPX)",
        after_fee_nlv=after_fee_nlv,
        month_pct=float(latest_row["bot_net_ret"]) * 100 if latest_row is not None else None,
        since_inception_pct_display=since_inception_display,
        last_updated=last_updated,
    )
    return render_portal_page(
        program_name="Momentum Pacer",
        accounts=accounts,
        columns=agm_registry.AGM_PORTAL_COLUMNS,
        row_fields=agm_registry.AGM_PORTAL_ROW_FIELDS,
    )


@app.server.route("/admin/logout")
def agm_admin_logout():
    agm_admin_auth_manager.logout(session)
    return redirect("/")


@app.server.route("/monthly")
def agm_monthly_backup():
    """The monthly workbook is a BACKEND source only — it is not exposed as a
    client- or admin-facing website experience. This route exists solely to make
    that decision explicit (a plain 404 instead of the Dash catch-all page)."""
    return "Not found", 404


@app.server.route("/healthz")
def agm_healthz():
    return jsonify({
        "app": "algominds-momentum-pacer",
        "status": "ready" if not _display_summary_df.empty else "error",
        "months_loaded": int(len(_display_summary_df)),
        "load_error": LOAD_ERROR,
        "daily_rows": int(len(daily_balances_df)),
        "spx_daily_rows": int(len(spx_daily_df)),
        "benchmark_ticker": agm_bench.SPX_TICKER,
        "daily_fee_days": int(len(daily_fee_accrual.daily)),
        "benchmark_load_error": BENCHMARK_LOAD_ERROR,
        "daily_fees_load_error": DAILY_FEES_LOAD_ERROR,
    })


# ==============================================================================
if __name__ == "__main__":
    # Production (e.g. Cloudflare Tunnel): set MP_TS_PRODUCTION=1 to disable debug/reloader
    # (debug mode can confuse reverse proxies and cause 502s on callbacks).
    _prod = os.environ.get("MP_TS_PRODUCTION", "").strip().lower() in ("1", "true", "yes")
    _debug = not _prod
    print(f"[mp_ts] Momentum Pacer tearsheet starting...")
    print(f"[mp_ts] Data: {EXCEL_PATH.name}")
    print(f"[mp_ts] Open: http://127.0.0.1:8304  (production={_prod})")
    app.run(
        debug=_debug,
        use_reloader=_debug,
        port=8304,
        host="127.0.0.1",
    )
