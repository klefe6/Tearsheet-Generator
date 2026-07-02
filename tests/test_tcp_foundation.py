"""Foundation tests for TCP v2 scaffold (no production module imports)."""
from __future__ import annotations

import ast
import sys
from datetime import datetime
from pathlib import Path

import pytest

from tcp_test_constants import CONTRACT_PATH, GOLDEN_FIXTURE_PATH, REQUIRED_SCENARIOS, TESTS_DIR

REPO_ROOT = TESTS_DIR.parent
PRODUCTION_MODULES = ("tcp_ts.py", "tkp_ts.py")


def _collect_imports(py_path: Path) -> set[str]:
    tree = ast.parse(py_path.read_text(encoding="utf-8"))
    modules: set[str] = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                modules.add(alias.name.split(".")[0])
        elif isinstance(node, ast.ImportFrom) and node.module:
            modules.add(node.module.split(".")[0])
    return modules


@pytest.mark.parametrize("test_file", sorted(TESTS_DIR.glob("test_*.py")))
def test_no_production_module_imports(test_file: Path):
    """Tests must not import tcp_ts or tkp_ts (Dash startup side effects)."""
    forbidden = {"tcp_ts", "tkp_ts"}
    imported = _collect_imports(test_file)
    assert not imported & forbidden, (
        f"{test_file.name} imports forbidden production modules: {imported & forbidden}"
    )


def test_conftest_does_not_import_production_modules():
    imported = _collect_imports(TESTS_DIR / "conftest.py")
    assert "tcp_ts" not in imported
    assert "tkp_ts" not in imported


def test_contract_document_exists_and_has_required_sections(contract_path: Path):
    assert contract_path.is_file(), f"Missing contract: {contract_path}"
    text = contract_path.read_text(encoding="utf-8")
    required_headings = [
        "## 4. Column contract",
        "## 5. Calculation-chain reconstruction",
        "## 6. Baseline analysis",
        "## 11. Golden-row candidates",
        "## 12. Decision register",
    ]
    for heading in required_headings:
        assert heading in text, f"Contract missing section: {heading}"


def test_golden_fixture_file_exists(golden_fixture_path: Path):
    assert golden_fixture_path.is_file()


def test_fixture_metadata_workbook_identity(golden_fixture: dict):
    meta = golden_fixture["metadata"]
    assert meta["workbook_filename"] == "tcp_alex.xlsx"
    assert meta["sheet_name"] == "NAV"
    assert meta["workbook_size_bytes"] > 0
    assert meta["workbook_last_write_time"]
    assert meta["fixture_extraction_timestamp"]
    assert meta["golden_excel_rows"] == [3, 4, 6, 7, 8, 10, 16, 17, 114]


def test_fixture_required_scenarios_present(golden_fixture: dict):
    found: set[str] = set()
    for row in golden_fixture["rows"]:
        found.update(row["scenarios"])
    missing = REQUIRED_SCENARIOS - found
    assert not missing, f"Fixture set missing scenarios: {sorted(missing)}"


def test_fixture_rows_chronological_and_unique_dates(golden_fixture: dict):
    rows = golden_fixture["rows"]
    dates = [datetime.strptime(r["date"], "%Y-%m-%d") for r in rows]
    assert dates == sorted(dates), "Golden rows must be in chronological order in fixture file"
    assert len({r["date"] for r in rows}) == len(rows)
    for row in rows:
        assert row["prior_excel_row_number"] < row["excel_row_number"]


@pytest.mark.local_workbook
def test_optional_local_workbook_matches_fixture(golden_fixture: dict):
    """When production workbook exists locally, verify observed values still match."""
    wb_path = Path(golden_fixture["metadata"]["workbook_path"])
    if not wb_path.is_file():
        pytest.skip(f"Local workbook not available: {wb_path}")

    stat = wb_path.stat()
    meta = golden_fixture["metadata"]
    if stat.st_size != meta["workbook_size_bytes"]:
        pytest.fail(
            f"Workbook size mismatch: fixture={meta['workbook_size_bytes']} "
            f"current={stat.st_size}"
        )

    import openpyxl

    wb = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)
    try:
        assert "NAV" in wb.sheetnames
        ws = wb["NAV"]
        for row in golden_fixture["rows"]:
            excel_row = row["excel_row_number"]
            for col_name, col_data in row["columns"].items():
                letter = col_data["excel_column"]
                current = ws[f"{letter}{excel_row}"].value
                if isinstance(current, datetime):
                    current = current.strftime("%Y-%m-%d")
                expected = col_data["observed_value"]
                if expected is None and current is None:
                    continue
                if isinstance(expected, (int, float)) and isinstance(current, (int, float)):
                    assert abs(float(current) - float(expected)) < 0.0001, (
                        f"Row {excel_row} {col_name}: expected {expected}, got {current}"
                    )
                else:
                    assert current == expected or str(current) == str(expected), (
                        f"Row {excel_row} {col_name}: expected {expected!r}, got {current!r}"
                    )
    finally:
        wb.close()


def test_production_files_exist_but_are_not_imported():
    """Sanity: production files on disk; tests run without importing them."""
    for name in PRODUCTION_MODULES:
        assert (REPO_ROOT / name).is_file()
    assert "tcp_ts" not in sys.modules
    assert "tkp_ts" not in sys.modules
