#!/usr/bin/env python3
"""
Read-only helper: extract TCP golden-row evidence from tcp_alex.xlsx into JSON fixtures.

Never saves or recalculates the workbook. Safe to rerun; refuses to overwrite unless --force.
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

# Default production workbook path (matches tcp_ts.py xlsx_path).
DEFAULT_WORKBOOK = (
    r"C:\Users\H&CDanHughes\Hughes & Company\Hughes & Company - Documents"
    r"\3_Advisors Marketing (Tearsheets, PitchBooks, etc)\1. Tearsheet Project\TCP\tcp_alex.xlsx"
)

SHEET_NAME = "NAV"

# Excel row numbers from Step 1 ledger contract (1-indexed, row 1 = header).
GOLDEN_EXCEL_ROWS: list[int] = [3, 4, 6, 7, 8, 10, 16, 17, 114]

SCENARIOS_BY_ROW: dict[int, list[str]] = {
    3: ["first_trading_day"],
    4: ["profitable_day", "fee"],
    6: ["rounding_small_pl"],
    7: ["losing_day", "loss_carry_start"],
    8: ["hwm_recovery"],
    10: ["under_hwm_loss_carry"],
    16: ["deposit", "tranche_change"],
    17: ["post_deposit"],
    114: ["latest_completed_row"],
}

# Ledger columns to capture (header name -> Excel column letter).
LEDGER_COLUMNS: dict[str, str] = {
    "Cash Transfers": "A",
    "Trading Days": "B",
    "Date": "C",
    "Cash Balance": "E",
    "NLV": "F",
    "#": "G",
    "$PL": "H",
    "Inc. Fee": "I",
    "cumm fee": "J",
    "Day PnL": "K",
    "nav-x1": "L",
    "Loss Carry": "N",
    "%Net": "O",
    "S net cummulative %": "P",
    "HWM": "Q",
}

NOTES_BY_ROW: dict[int, str] = {
    3: "First trading day: Day PnL=0; nav-x1 seeded from U6 (=50000). Medium uncertainty per audit.",
    4: "Profitable day with 20% fee when $PL > prior Loss Carry (0).",
    6: "Small P&L / rounding-sensitive row.",
    7: "Loss day; Loss Carry begins (MAX(0, prior HWM - nav-x1)).",
    8: "Recovery; nav-x1 reaches HWM (50056.792).",
    10: "Under HWM with Loss Carry > 0.",
    16: "Deposit +25000; tranche # changes 1→2. HWM formula on tranche change: medium uncertainty.",
    17: "First post-deposit row; $PL subtracts prior-row Cash Transfer (row 16).",
    114: "Latest completed ledger row at fixture extraction time.",
}


def _serialize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float, str, bool)):
        return value
    return str(value)


def _cell_evidence(ws_formula, ws_values, col_letter: str, row: int) -> dict[str, Any]:
    addr = f"{col_letter}{row}"
    formula_cell = ws_formula[addr]
    value_cell = ws_values[addr]
    raw = formula_cell.value
    is_formula = isinstance(raw, str) and raw.startswith("=")
    return {
        "excel_column": col_letter,
        "observed_value": _serialize_value(value_cell.value),
        "excel_formula": raw if is_formula else None,
        "number_format": formula_cell.number_format,
        "confidence": "confirmed",
    }


def extract_workbook(path: Path) -> dict[str, Any]:
    if not path.is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")

    stat = path.stat()
    wb_formula = openpyxl.load_workbook(path, data_only=False, read_only=False)
    wb_values = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        if SHEET_NAME not in wb_formula.sheetnames:
            raise ValueError(f"Sheet {SHEET_NAME!r} not found in {path.name}")
        ws_f = wb_formula[SHEET_NAME]
        ws_v = wb_values[SHEET_NAME]

        header_mapping = {}
        for col_idx in range(1, 30):
            letter = get_column_letter(col_idx)
            val = ws_f[f"{letter}1"].value
            if val is not None:
                header_mapping[letter] = str(val).strip()

        rows_out = []
        for excel_row in sorted(GOLDEN_EXCEL_ROWS):
            prior_excel_row = excel_row - 1
            dataframe_index = excel_row - 2  # 0-based data row index (header = row 1)
            columns = {
                name: _cell_evidence(ws_f, ws_v, col_letter, excel_row)
                for name, col_letter in LEDGER_COLUMNS.items()
            }
            date_val = columns["Date"]["observed_value"]
            rows_out.append(
                {
                    "excel_row_number": excel_row,
                    "dataframe_index": dataframe_index,
                    "prior_excel_row_number": prior_excel_row,
                    "date": date_val,
                    "scenarios": SCENARIOS_BY_ROW.get(excel_row, []),
                    "columns": columns,
                    "notes": NOTES_BY_ROW.get(excel_row, ""),
                }
            )

        return {
            "metadata": {
                "workbook_path": str(path),
                "workbook_filename": path.name,
                "sheet_name": SHEET_NAME,
                "workbook_size_bytes": stat.st_size,
                "workbook_last_write_time": datetime.fromtimestamp(
                    stat.st_mtime, tz=timezone.utc
                ).isoformat(),
                "fixture_extraction_timestamp": datetime.now(timezone.utc).isoformat(),
                "header_mapping": header_mapping,
                "golden_excel_rows": GOLDEN_EXCEL_ROWS,
                "hidden_parameter_cells": {
                    "U6": _serialize_value(ws_v["U6"].value),
                    "U10": _serialize_value(ws_v["U10"].value),
                },
            },
            "rows": rows_out,
        }
    finally:
        wb_formula.close()
        wb_values.close()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        default=DEFAULT_WORKBOOK,
        help="Absolute path to tcp_alex.xlsx",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().parent / "fixtures" / "tcp_golden_rows.json",
        help="Output JSON path",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Overwrite existing fixture file",
    )
    args = parser.parse_args(argv)

    out_path: Path = args.output
    if out_path.exists() and not args.force:
        print(f"Refusing to overwrite {out_path} without --force", file=sys.stderr)
        return 1

    payload = extract_workbook(Path(args.workbook))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, sort_keys=False)
        f.write("\n")

    print(f"Wrote {len(payload['rows'])} golden rows to {out_path}")
    for row in payload["rows"]:
        print(
            f"  row {row['excel_row_number']:3d}  {row['date']}  "
            f"{', '.join(row['scenarios'])}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
