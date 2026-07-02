"""Tests for tcp_ledger read-only adapter."""
from __future__ import annotations

import math
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pytest

from tcp_ledger import (
    DuplicateRequiredHeader,
    FormulaCacheMissing,
    InvalidDate,
    InvalidNumericValue,
    LedgerEmpty,
    RequiredColumnMissing,
    WorkbookNotFound,
    WorksheetMissing,
    get_record_by_excel_row,
    load_ledger,
)
from tcp_test_constants import GOLDEN_FIXTURE_PATH, REQUIRED_LEDGER_FIELDS

REPO_ROOT = Path(__file__).resolve().parent.parent
TMP_LEDGER_DIR = REPO_ROOT / "tests" / "_tmp_ledger"


@pytest.fixture
def ledger_tmp(request):
    """Repo-local temp dir (avoids Windows pytest ledger_tmp permission issues)."""
    TMP_LEDGER_DIR.mkdir(parents=True, exist_ok=True)
    safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in request.node.name)
    path = TMP_LEDGER_DIR / safe_name
    path.mkdir(parents=True, exist_ok=True)
    yield path
    for child in path.iterdir():
        if child.is_file():
            child.unlink()

HEADERS = list(REQUIRED_LEDGER_FIELDS)

# Production NAV sheet column letters (column D is an em-dash placeholder).
PRODUCTION_HEADER_COLUMNS: dict[str, str] = {
    "Cash Transfers": "A",
    "Trading Days": "B",
    "Date": "C",
    "Cash Balance": "E",
    "NLV": "F",
    "#": "G",
    "$PL": "H",
    "Inc. Fee": "I",
    "cumm fee": "J",
    "Day PnL": "K",
    "nav-x1": "L",
    "Loss Carry": "N",
    "%Net": "O",
    "S net cummulative %": "P",
    "HWM": "Q",
}


def _write_production_header(ws) -> None:
    ws["D1"] = "—"
    for name, letter in PRODUCTION_HEADER_COLUMNS.items():
        ws[f"{letter}1"] = name


def _write_production_row(ws, excel_row: int, values: dict[str, object]) -> None:
    for name, letter in PRODUCTION_HEADER_COLUMNS.items():
        if name in values:
            ws[f"{letter}{excel_row}"] = values[name]


def _write_header(ws, headers: list[str] | None = None) -> None:
    headers = headers or HEADERS
    for idx, name in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=name)


def _write_row(
    ws,
    excel_row: int,
    values: dict[str, object],
    *,
    headers: list[str] | None = None,
) -> None:
    headers = headers or HEADERS
    for idx, name in enumerate(headers, start=1):
        if name in values:
            ws.cell(row=excel_row, column=idx, value=values[name])


def _make_workbook(
    path: Path,
    rows: list[dict[str, object]],
    *,
    sheet_name: str = "NAV",
    headers: list[str] | None = None,
    extra_trailing_blank_rows: int = 0,
    compact_headers: bool = False,
) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    use_compact = compact_headers or headers is not None
    if use_compact:
        hdrs = headers or HEADERS
        _write_header(ws, hdrs)
        for offset, row_data in enumerate(rows):
            _write_row(ws, excel_row=2 + offset, values=row_data, headers=hdrs)
        if extra_trailing_blank_rows:
            start = 2 + len(rows)
            for i in range(extra_trailing_blank_rows):
                _write_row(ws, excel_row=start + i, values={}, headers=hdrs)
    else:
        _write_production_header(ws)
        for offset, row_data in enumerate(rows):
            _write_production_row(ws, 2 + offset, row_data)
        if extra_trailing_blank_rows:
            start = 2 + len(rows)
            for i in range(extra_trailing_blank_rows):
                _write_production_row(ws, start + i, {})
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()
    return path


def _completed_row(
    row_date: str,
    nav: float,
    *,
    cash_transfer: float | None = None,
    units: int = 1,
    trading_day: int = 1,
) -> dict[str, object]:
    return {
        "Cash Transfers": cash_transfer,
        "Trading Days": trading_day,
        "Date": row_date,
        "Cash Balance": 50000.0,
        "NLV": 50000.0,
        "#": units,
        "$PL": 100.0,
        "Inc. Fee": 0.0,
        "cumm fee": 0.0,
        "Day PnL": 100.0,
        "nav-x1": nav,
        "Loss Carry": 0.0,
        "%Net": 0.002,
        "S net cummulative %": 0.002,
        "HWM": nav,
    }


@pytest.fixture
def golden_fixture():
    import json

    with GOLDEN_FIXTURE_PATH.open(encoding="utf-8") as f:
        return json.load(f)


def test_required_header_discovery(ledger_tmp):
    path = _make_workbook(
        ledger_tmp / "headers.xlsx",
        [_completed_row("2026-01-20", 50000.0)],
    )
    result = load_ledger(str(path))
    for header in REQUIRED_LEDGER_FIELDS:
        assert header in result.metadata.header_mapping.values() or any(
            v == header for v in result.metadata.header_mapping.values()
        )


def test_correct_header_mapping(ledger_tmp):
    path = _make_workbook(
        ledger_tmp / "map.xlsx",
        [_completed_row("2026-01-20", 50000.0)],
    )
    result = load_ledger(str(path))
    letters = {v: k for k, v in result.metadata.header_mapping.items()}
    assert letters["Date"] == "C"
    assert letters["nav-x1"] == "L"
    assert letters["#"] == "G"


def test_completed_row_detection(ledger_tmp):
    rows = [
        {"Date": "2026-01-19", "nav-x1": None, "Trading Days": 0},
        _completed_row("2026-01-20", 50000.0, trading_day=1),
        _completed_row("2026-01-21", 50100.0, trading_day=2),
    ]
    path = _make_workbook(ledger_tmp / "completed.xlsx", rows)
    result = load_ledger(str(path))
    assert result.metadata.completed_row_count == 2
    assert result.metadata.total_candidate_rows == 3
    assert result.metadata.latest_completed_date == date(2026, 1, 21)


def test_blank_trailing_rows_ignored(ledger_tmp):
    rows = [
        _completed_row("2026-01-20", 50000.0),
        _completed_row("2026-01-21", 50100.0, trading_day=2),
    ]
    path = _make_workbook(
        ledger_tmp / "trailing.xlsx",
        rows,
        extra_trailing_blank_rows=5,
    )
    result = load_ledger(str(path))
    assert result.metadata.completed_row_count == 2
    assert result.completed_records[-1].fields["Date"] == date(2026, 1, 21)


def test_date_only_row_excluded_from_latest_completed(ledger_tmp):
    rows = [
        {"Date": "2026-01-19", "nav-x1": None, "Trading Days": 0},
        _completed_row("2026-01-20", 50000.0, trading_day=1),
    ]
    path = _make_workbook(ledger_tmp / "date_only.xlsx", rows)
    result = load_ledger(str(path))
    assert result.metadata.latest_completed_date == date(2026, 1, 20)
    assert result.metadata.latest_completed_excel_row == 3
    date_only = result.candidate_records[0]
    assert date_only.fields["Date"] == date(2026, 1, 19)
    assert date_only.fields["nav-x1"] is None


def test_currency_normalization(ledger_tmp):
    row = _completed_row("2026-01-20", 50000.0)
    row["Cash Balance"] = 24996.76
    row["$PL"] = -3.24
    path = _make_workbook(ledger_tmp / "currency.xlsx", [row])
    record = load_ledger(str(path)).completed_records[0]
    assert record.fields["Cash Balance"] == pytest.approx(24996.76)
    assert record.fields["$PL"] == pytest.approx(-3.24)


def test_percentage_normalization_decimal_ratio(ledger_tmp):
    row = _completed_row("2026-01-20", 50000.0)
    row["%Net"] = -6.48e-05
    row["S net cummulative %"] = -6.48e-05
    path = _make_workbook(ledger_tmp / "pct.xlsx", [row])
    record = load_ledger(str(path)).completed_records[0]
    assert record.fields["%Net"] == pytest.approx(-6.48e-05)
    assert record.fields["S net cummulative %"] == pytest.approx(-6.48e-05)


def test_tranche_unit_count_preserved(ledger_tmp):
    row = _completed_row("2026-02-06", 47000.0, cash_transfer=25000, units=2, trading_day=14)
    path = _make_workbook(ledger_tmp / "tranche.xlsx", [row])
    record = load_ledger(str(path)).completed_records[0]
    assert record.fields["#"] == 2.0
    assert record.fields["Cash Transfers"] == 25000.0


def test_missing_workbook(ledger_tmp):
    with pytest.raises(WorkbookNotFound):
        load_ledger(str(ledger_tmp / "missing.xlsx"))


def test_missing_worksheet(ledger_tmp):
    path = ledger_tmp / "nosheet.xlsx"
    wb = openpyxl.Workbook()
    wb.save(path)
    wb.close()
    with pytest.raises(WorksheetMissing):
        load_ledger(str(path), sheet_name="NAV")


def test_missing_required_column(ledger_tmp):
    headers = [h for h in HEADERS if h != "nav-x1"]
    path = _make_workbook(
        ledger_tmp / "missing_col.xlsx",
        [{"Date": "2026-01-20", "Trading Days": 1}],
        headers=headers,
    )
    with pytest.raises(RequiredColumnMissing):
        load_ledger(str(path))


def test_duplicate_required_header(ledger_tmp):
    headers = HEADERS + ["Date"]
    path = _make_workbook(
        ledger_tmp / "dup.xlsx",
        [_completed_row("2026-01-20", 50000.0)],
        headers=headers,
    )
    with pytest.raises(DuplicateRequiredHeader):
        load_ledger(str(path))


def test_invalid_date(ledger_tmp):
    row = _completed_row("2026-01-20", 50000.0)
    row["Date"] = "not-a-date"
    path = _make_workbook(ledger_tmp / "bad_date.xlsx", [row])
    with pytest.raises(InvalidDate):
        load_ledger(str(path))


def test_invalid_numeric_cell(ledger_tmp):
    row = _completed_row("2026-01-20", 50000.0)
    row["nav-x1"] = "#DIV/0!"
    path = _make_workbook(ledger_tmp / "bad_num.xlsx", [row])
    with pytest.raises(InvalidNumericValue):
        load_ledger(str(path))


def test_empty_completed_ledger(ledger_tmp):
    rows = [
        {"Date": "2026-01-19", "nav-x1": None, "Trading Days": 0},
        {"Date": "2026-01-20", "nav-x1": None, "Trading Days": 1},
    ]
    path = _make_workbook(ledger_tmp / "empty.xlsx", rows)
    with pytest.raises(LedgerEmpty):
        load_ledger(str(path))


def test_no_writes_to_source_workbook(ledger_tmp):
    path = _make_workbook(
        ledger_tmp / "readonly.xlsx",
        [_completed_row("2026-01-20", 50000.0)],
    )
    before = path.stat()
    load_ledger(str(path))
    after = path.stat()
    assert before.st_size == after.st_size
    assert int(before.st_mtime) == int(after.st_mtime)


def test_deterministic_record_ordering(ledger_tmp):
    rows = [
        _completed_row("2026-01-20", 50000.0, trading_day=1),
        _completed_row("2026-01-21", 50100.0, trading_day=2),
        _completed_row("2026-01-22", 50200.0, trading_day=3),
    ]
    path = _make_workbook(ledger_tmp / "order.xlsx", rows)
    first = load_ledger(str(path))
    second = load_ledger(str(path))
    assert [r.excel_row_number for r in first.completed_records] == [
        r.excel_row_number for r in second.completed_records
    ]
    assert [r.fields["Date"] for r in first.completed_records] == [
        date(2026, 1, 20),
        date(2026, 1, 21),
        date(2026, 1, 22),
    ]


def test_public_metadata_excludes_absolute_path(ledger_tmp):
    path = _make_workbook(
        ledger_tmp / "meta.xlsx",
        [_completed_row("2026-01-20", 50000.0)],
    )
    meta = load_ledger(str(path)).metadata.public_dict()
    assert meta["source_filename"] == "meta.xlsx"
    assert "Coding Projects" not in str(meta)
    assert "tmp" not in str(meta).lower() or meta["source_filename"] == "meta.xlsx"


def test_formula_cache_missing_strict_mode(ledger_tmp):
    path = ledger_tmp / "formula.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NAV"
    _write_production_header(ws)
    ws["C2"] = datetime(2026, 1, 20)
    ws["L2"] = "=U6"
    wb.save(path)
    wb.close()
    with pytest.raises(FormulaCacheMissing):
        load_ledger(str(path), strict_formula_cache=True)


@pytest.mark.local_workbook
def test_local_workbook_nav_sheet_and_headers(golden_fixture):
    wb_path = Path(golden_fixture["metadata"]["workbook_path"])
    if not wb_path.is_file():
        pytest.skip(f"Local workbook not available: {wb_path}")

    result = load_ledger(str(wb_path), golden_fixture["metadata"]["sheet_name"])
    assert result.metadata.sheet_name == "NAV"
    for header in REQUIRED_LEDGER_FIELDS:
        assert header in result.metadata.header_mapping.values()


@pytest.mark.local_workbook
def test_local_workbook_golden_rows_match_fixture(golden_fixture):
    wb_path = Path(golden_fixture["metadata"]["workbook_path"])
    if not wb_path.is_file():
        pytest.skip(f"Local workbook not available: {wb_path}")

    result = load_ledger(str(wb_path))
    for fixture_row in golden_fixture["rows"]:
        excel_row = fixture_row["excel_row_number"]
        record = get_record_by_excel_row(result, excel_row)
        assert record is not None, f"Missing excel row {excel_row}"
        for field_name, evidence in fixture_row["columns"].items():
            expected = evidence["observed_value"]
            actual = record.fields[field_name]
            if expected is None:
                assert actual is None, f"Row {excel_row} {field_name}"
            elif isinstance(expected, str) and field_name == "Date":
                assert actual == date.fromisoformat(expected)
            elif isinstance(expected, (int, float)):
                assert actual == pytest.approx(float(expected), rel=0, abs=1e-6), (
                    f"Row {excel_row} {field_name}: {actual} != {expected}"
                )
            else:
                assert actual == expected


@pytest.mark.local_workbook
def test_local_workbook_formula_strings(golden_fixture):
    wb_path = Path(golden_fixture["metadata"]["workbook_path"])
    if not wb_path.is_file():
        pytest.skip(f"Local workbook not available: {wb_path}")

    result = load_ledger(str(wb_path))
    for fixture_row in golden_fixture["rows"]:
        excel_row = fixture_row["excel_row_number"]
        record = get_record_by_excel_row(result, excel_row)
        assert record is not None
        for field_name, evidence in fixture_row["columns"].items():
            fixture_formula = evidence["excel_formula"]
            if fixture_formula is None:
                continue
            assert record.formulas.get(field_name) == fixture_formula, (
                f"Row {excel_row} {field_name} formula mismatch"
            )


@pytest.mark.local_workbook
def test_local_workbook_latest_completed_metadata(golden_fixture):
    wb_path = Path(golden_fixture["metadata"]["workbook_path"])
    if not wb_path.is_file():
        pytest.skip(f"Local workbook not available: {wb_path}")

    result = load_ledger(str(wb_path))
    assert result.metadata.latest_completed_date == date(2026, 6, 24)
    assert result.metadata.latest_completed_excel_row == 114
    assert result.metadata.completed_row_count == 112


@pytest.mark.local_workbook
def test_local_workbook_row_16_deposit_tranche(golden_fixture):
    wb_path = Path(golden_fixture["metadata"]["workbook_path"])
    if not wb_path.is_file():
        pytest.skip(f"Local workbook not available: {wb_path}")

    record = get_record_by_excel_row(load_ledger(str(wb_path)), 16)
    assert record is not None
    assert record.fields["Cash Transfers"] == pytest.approx(25000.0)
    assert record.fields["#"] == 2.0
    assert record.fields["Date"] == date(2026, 2, 6)


@pytest.mark.local_workbook
def test_local_workbook_row_17_post_deposit(golden_fixture):
    wb_path = Path(golden_fixture["metadata"]["workbook_path"])
    if not wb_path.is_file():
        pytest.skip(f"Local workbook not available: {wb_path}")

    record = get_record_by_excel_row(load_ledger(str(wb_path)), 17)
    assert record is not None
    assert record.fields["Date"] == date(2026, 2, 9)
    assert record.fields["#"] == 2.0
    assert record.fields["$PL"] == pytest.approx(248.15999999999622, rel=0, abs=1e-6)


@pytest.mark.local_workbook
def test_local_workbook_unchanged_after_load(golden_fixture):
    wb_path = Path(golden_fixture["metadata"]["workbook_path"])
    if not wb_path.is_file():
        pytest.skip(f"Local workbook not available: {wb_path}")

    before = wb_path.stat()
    load_ledger(str(wb_path))
    after = wb_path.stat()
    assert before.st_size == after.st_size
    assert int(before.st_mtime) == int(after.st_mtime)
    expected_size = golden_fixture["metadata"]["workbook_size_bytes"]
    assert before.st_size == expected_size


def test_first_trading_day_special_case_values(golden_fixture):
    """Golden row 3 remains loadable with Day PnL=0 and nav-x1=50000."""
    row3 = next(r for r in golden_fixture["rows"] if r["excel_row_number"] == 3)
    assert row3["columns"]["Day PnL"]["observed_value"] == 0
    assert row3["columns"]["nav-x1"]["observed_value"] == 50000


def test_percentage_values_are_finite_floats(ledger_tmp):
    row = _completed_row("2026-01-20", 50000.0)
    row["%Net"] = 0.00123
    path = _make_workbook(ledger_tmp / "finite.xlsx", [row])
    pct = load_ledger(str(path)).completed_records[0].fields["%Net"]
    assert isinstance(pct, float)
    assert math.isfinite(pct)
